# -*- coding: utf-8 -*-
"""
TaxAuditEngine — 通用税审底稿填充引擎

配置驱动、指令执行式。不包含任何模板特化逻辑。
所有模板结构、行列映射、校验规则均来自 YAML 配置。
"""

import sys, os, shutil, re, copy

sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path
from typing import Any, Dict, List, Optional, Union
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


# ============================================================
# 辅助函数
# ============================================================


def _r2(val):
    """取两位小数"""
    return round(float(val) if val else 0.0, 2)


def _safe_write(ws, row, col, value):
    """安全写入单元格（自动处理合并单元格）"""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                ws.cell(row=mr.min_row, column=mr.min_col).value = value
                return
    try:
        cell.value = value
    except AttributeError:
        pass


def _find_sheet(wb, sheet_name: str):
    """按名称精确或模糊查找sheet"""
    for sn in wb.sheetnames:
        if sn.strip() == sheet_name.strip():
            return wb[sn]
    for sn in wb.sheetnames:
        if sheet_name.replace(" ", "") in sn.replace(" ", ""):
            return wb[sn]
    return None


def _detect_row(ws, pattern: str, start_row=1, end_row=None) -> Optional[int]:
    """
    按文本模式检测行号
    pattern: 正则表达式
    """
    if end_row is None:
        end_row = ws.max_row
    for r in range(start_row, end_row + 1):
        for c in range(1, ws.max_column + 1):
            cv = ws.cell(row=r, column=c).value
            if cv and isinstance(cv, str) and re.search(pattern, cv):
                return r
    return None


def _xls_to_xlsx_com(xls_path: str, xlsx_path: str, timeout: int = 90):
    """
    用 Excel COM 将 xls 转换为 xlsx，完整保留格式

    在独立子进程中运行，避免 Excel COM 挂死阻塞主进程。
    超时后自动 kill 子进程 + 清理 Excel.exe 残留。
    """
    import subprocess

    proc = subprocess.Popen(
        [sys.executable, "-m", "tax_audit_engine.xls_convert", xls_path, xlsx_path],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.PIPE,
    )
    try:
        _, stderr = proc.communicate(timeout=timeout)
        if proc.returncode != 0:
            err_msg = stderr.decode("utf-8", errors="replace")
            # 杀残留 Excel
            _cleanup_excel()
            raise RuntimeError(f"Excel COM 转换失败 [{proc.returncode}]: {err_msg}")
    except subprocess.TimeoutExpired:
        proc.kill()
        _cleanup_excel()
        raise RuntimeError(f"Excel COM 超时 ({timeout}s): {xls_path}")


def _cleanup_excel():
    """清理残留的 Excel 进程"""
    import subprocess as _sp

    _sp.run(
        ["taskkill", "/f", "/im", "excel.exe"], stdout=_sp.DEVNULL, stderr=_sp.DEVNULL
    )


# ============================================================
# 值解析器
# ============================================================


class ValueResolver:
    """
    从配置指令中解析实际值

    支持的 value_from 前缀:
      tb.xxx          → result.tb.get("xxx")
      result.xxx      → getattr(result, "xxx")
      enterprise.xxx  → getattr(result.enterprise, "xxx")
      adj.xxx.field   → 按名称查找调整项，取字段
      calc:sum_tb(a,b,c)  → 多个TB科目加总
      calc:profit     → 会计利润
      calc:taxable    → 应纳税所得额
      literal:xxx     → 直接量
    """

    def __init__(self, result, assets=None):
        self.result = result
        self.tb = result.tb
        self.enterprise = result.enterprise
        self.adjustments = result.adjustments
        self.adj_by_name = {a.item_name: a for a in result.adjustments}
        self.assets = assets or []

    def resolve(self, value_spec: str):
        """解析值规格"""
        if not value_spec or not isinstance(value_spec, str):
            return value_spec

        if value_spec.startswith("tb."):
            key = value_spec[3:]
            return _r2(self.tb.get(key) or 0)

        if value_spec.startswith("result."):
            attr = value_spec[7:]
            val = getattr(self.result, attr, None)
            return _r2(val) if isinstance(val, (int, float)) else val

        if value_spec.startswith("enterprise."):
            attr = value_spec[11:]
            return getattr(self.enterprise, attr, None)

        if value_spec.startswith("adj."):
            parts = value_spec[4:].split(".")
            if len(parts) >= 2:
                name = parts[0]
                field = parts[1]
                adj = self.adj_by_name.get(name)
                if adj:
                    return _r2(getattr(adj, field, 0))
            return 0

        if value_spec.startswith("calc:"):
            return self._resolve_calc(value_spec[5:])

        if value_spec.startswith("literal:"):
            return value_spec[8:]

        return value_spec

    def _resolve_calc(self, expr: str):
        """解析 calc: 表达式"""
        if expr == "profit":
            return _r2(self.tb.accounting_profit)
        if expr == "taxable":
            return _r2(self.result.taxable_income)
        if expr == "tax_payable":
            val = getattr(self.result, "tax_payable", None) or getattr(
                self.result, "final_tax", None
            )
            return _r2(val) if val else 0
        if expr == "total_increase":
            return _r2(self.result.total_increase)
        if expr == "total_decrease":
            return _r2(self.result.total_decrease)
        if expr.startswith("sum_tb("):
            # sum_tb(key1,key2,...)
            inner = expr[7:].rstrip(")")
            keys = [k.strip() for k in inner.split(",")]
            total = sum(self.tb.get(k) or 0 for k in keys)
            return _r2(total)
        return expr

    def resolve_template(self, template_str: str, extra_vars: dict = None) -> str:
        """解析带占位符的模板字符串"""
        vars_dict = {
            "tb": self.tb,
            "enterprise": self.enterprise,
            "result": self.result,
            "total_increase": self.result.total_increase,
            "total_decrease": self.result.total_decrease,
            "taxable_income": self.result.taxable_income,
            "profit": self.tb.accounting_profit,
        }
        if extra_vars:
            vars_dict.update(extra_vars)
        return template_str.format(**vars_dict)


# ============================================================
# 指令执行器
# ============================================================


class InstructionExecutor:
    """
    执行模板配置中的指令列表

    支持的指令类型:
      cell          — 写入单个单元格
      data_block    — 从TB取值写入数据行区域（含合计行）
      row_map       — 行号→值映射写入
      adj_row_map   — 调整项→行映射写入（账载/税收/调增/调减）
      sum_row       — 合计行写入
      asset_category_rows — 资产类别→行写入
      payroll_rows  — 薪酬表多列写入
    """

    def __init__(self, wb, resolver: ValueResolver, config: dict, result, assets=None):
        self.wb = wb
        self.resolver = resolver
        self.config = config
        self.result = result
        self.assets = assets or []
        self.filled = []

    def execute_all(self, instructions: list):
        """执行全部指令"""
        for instr in instructions:
            self._execute_one(instr)
        return self.filled

    def _execute_one(self, instr: dict):
        """执行单条指令"""
        itype = instr.get("instruction")
        handler = getattr(self, f"_handle_{itype}", None)
        if handler:
            handler(instr)
        else:
            print(f"  ⚠ 未知指令类型: {itype}")

    # ---- 指令处理 ----

    def _handle_cell(self, instr: dict):
        """cell: 写入单个单元格"""
        sheet_name = instr.get("sheet")
        ws = _find_sheet(self.wb, sheet_name)
        if not ws:
            return

        row = instr.get("row")
        col = instr.get("col")

        # 行检测
        if row is None or row == "auto":
            pattern = instr.get("row_detect")
            if pattern:
                row = _detect_row(ws, pattern)

        value_from = instr.get("value_from")
        value = instr.get("value")

        if value_from:
            value = self.resolver.resolve(value_from)
        elif value_template := instr.get("value_template"):
            extra = instr.get("value_args", {})
            resolved_args = {
                k: self.resolver.resolve(v) if isinstance(v, str) else v
                for k, v in extra.items()
            }
            value = self.resolver.resolve_template(value_template, resolved_args)

        if value is not None and row and col:
            _safe_write(ws, row, col, _r2(value) if isinstance(value, float) else value)
            self.filled.append(f"{sheet_name} R{row}C{col}={value}")

    def _handle_data_block(self, instr: dict):
        """data_block: 从TB取值写入数据行区域（含合计行和审核说明）"""
        sheet_name = instr.get("sheet")
        ws = _find_sheet(self.wb, sheet_name)
        if not ws:
            return

        tb_keys = instr.get("tb_keys", [])
        data_start = instr.get("data_start", 8)
        col_book = instr.get("col_book", 3)
        col_label = instr.get("col_label", 2)
        col_seq = instr.get("col_seq", 1)

        # 合计行检测
        sum_pattern = instr.get("sum_row_detect", ".*合计.*")
        sum_row = _detect_row(ws, sum_pattern, start_row=data_start)

        # 审核说明行检测
        note_pattern = instr.get("note_row_detect", "数据来源|审核说明")
        note_row = _detect_row(ws, note_pattern, start_row=data_start)

        # 取值
        values = []
        for key in tb_keys:
            v = self.result.tb.get(key)
            if v and abs(v) > 0.01:
                rv = _r2(v)
                dup = False
                for _, ev in values:
                    if abs(ev - rv) < 1.0:
                        dup = True
                        break
                if not dup:
                    values.append((key, rv))

        if not values:
            return

        total_book = sum(v for _, v in values)

        # 写入数据行
        row = data_start
        for label, val in values:
            if sum_row and row >= sum_row:
                break
            if col_seq:
                _safe_write(ws, row, col_seq, row - data_start + 1)
            if col_label:
                _safe_write(ws, row, col_label, label)
            _safe_write(ws, row, col_book, val)
            row += 1

        # 合计行
        if sum_row:
            _safe_write(ws, sum_row, col_book, total_book)

        # 审核说明
        if note_row:
            tax_note = instr.get("tax_note", "")
            template = instr.get(
                "note_template",
                "审核说明：{tax_note}。数据来源于客户提供的年度账表及已审计的年度财务报告，经审核余额可以确认。",
            )
            note_text = template.format(tax_note=tax_note)
            _safe_write(ws, note_row, 1, note_text)

        self.filled.append(f"{sheet_name}: {len(values)} 项, 合计 {total_book:,.0f}")

    def _resolve_sheet(self, instr: dict):
        """按 sheet/sheet_fallback 查找工作表"""
        ws = _find_sheet(self.wb, instr.get("sheet", ""))
        if ws:
            return ws
        fallback = instr.get("sheet_fallback")
        if fallback:
            ws = _find_sheet(self.wb, fallback)
        return ws

    def _handle_row_map(self, instr: dict):
        """row_map: 行号→值映射"""
        sheet_name = instr.get("sheet")
        ws = self._resolve_sheet(instr)
        if not ws:
            return

        col = instr.get("col", 4)
        entries = instr.get("entries", [])
        written = 0
        for entry in entries:
            row = entry.get("row")
            value_from = entry.get("value_from")
            if value_from:
                val = self.resolver.resolve(value_from)
                if val is not None and val != "":
                    if isinstance(val, (int, float)):
                        if abs(val) > 0.01:
                            _safe_write(ws, row, col, _r2(val))
                            written += 1
                    else:
                        _safe_write(ws, row, col, val)
                        written += 1

            tb_key = entry.get("tb_key")
            if tb_key:
                val = self.resolver.resolve(f"tb.{tb_key}")
                if isinstance(val, (int, float)) and abs(val) > 0.01:
                    _safe_write(ws, row, col, _r2(val))
                    written += 1

        self.filled.append(f"{sheet_name}: {written} 行")

    def _handle_adj_row_map(self, instr: dict):
        """adj_row_map: 调整项→行写入"""
        sheet_name = instr.get("sheet")
        ws = _find_sheet(self.wb, sheet_name)
        if not ws:
            return

        entries = instr.get("entries", [])
        col_book = instr.get("col_book", 3)
        col_tax = instr.get("col_tax", 4)
        col_inc = instr.get("col_inc", 5)
        col_dec = instr.get("col_dec", 6)

        for entry in entries:
            row = entry.get("row")
            name = entry.get("name")
            adj = self.resolver.adj_by_name.get(name)
            if not adj:
                for a in self.result.adjustments:
                    if name in a.item_name or a.item_name in name:
                        adj = a
                        break
            if adj:
                _safe_write(ws, row, col_book, _r2(adj.book_amount))
                _safe_write(ws, row, col_tax, _r2(adj.tax_base))
                if adj.increase > 0:
                    _safe_write(ws, row, col_inc, _r2(adj.increase))
                elif adj.decrease > 0:
                    _safe_write(ws, row, col_dec, _r2(adj.decrease))
                self.filled.append(f"{sheet_name} R{row} {name}")

        # 合计行
        sum_row = instr.get("sum_row")
        if sum_row:
            sum_col_inc = instr.get("sum_col_inc", col_inc)
            sum_col_dec = instr.get("sum_col_dec", col_dec)
            total_inc = _r2(self.result.total_increase)
            total_dec = _r2(self.result.total_decrease)
            _safe_write(ws, sum_row, sum_col_inc, total_inc)
            _safe_write(ws, sum_row, sum_col_dec, total_dec)
            self.filled.append(f"{sheet_name} 合计 R{sum_row}")

    def _handle_sum_row(self, instr: dict):
        """sum_row: 计算行写入（利润/调整/所得等）"""
        sheet_name = instr.get("sheet")
        ws = _find_sheet(self.wb, sheet_name)
        if not ws:
            return

        entries = instr.get("entries", [])
        for entry in entries:
            row = entry.get("row")
            col = entry.get("col", 8)
            value_from = entry.get("value_from")
            label = entry.get("label", "")
            if value_from:
                val = self.resolver.resolve(value_from)
                if val and abs(val) > 0.01:
                    _safe_write(ws, row, col, _r2(val))
                    self.filled.append(f"{sheet_name} R{row} {label}={val:,.0f}")

    def _handle_asset_category_rows(self, instr: dict):
        """asset_category_rows: 按资产类别写入折旧数据"""
        sheet_name = instr.get("sheet")
        ws = _find_sheet(self.wb, sheet_name)
        if not ws:
            return

        col_orig = instr.get("col_orig", 5)
        col_acct_depr = instr.get("col_acct_depr", 6)
        col_tax_base = instr.get("col_tax_base")  # optional: 计税基础列
        col_tax_depr = instr.get("col_tax_depr", 9)
        cat_rows = instr.get("cat_rows", {})
        sum_row = instr.get("sum_row")
        note_row = instr.get("note_row")
        note_template = instr.get("note_template", "")

        assets = self.assets
        if not assets:
            return

        # 按类别汇总
        cat_data = {}
        for a in assets:
            cat = a.category
            if cat not in cat_data:
                cat_data[cat] = {"orig": 0, "acct_depr": 0, "tax_depr": 0}
            cat_data[cat]["orig"] += _r2(a.original_value)
            cat_data[cat]["acct_depr"] += _r2(a.current_accounting_depr)
            cat_data[cat]["tax_depr"] += _r2(a.current_tax_depr)

        total_orig = total_acct = total_tax = 0
        for cat, row_num in cat_rows.items():
            data = cat_data.get(cat)
            if not data:
                continue
            _safe_write(ws, row_num, col_orig, data["orig"])
            _safe_write(ws, row_num, col_acct_depr, data["acct_depr"])
            if col_tax_base:
                _safe_write(ws, row_num, col_tax_base, data["orig"])  # 计税基础=原值
            _safe_write(ws, row_num, col_tax_depr, data["tax_depr"])
            total_orig += data["orig"]
            total_acct += data["acct_depr"]
            total_tax += data["tax_depr"]
            self.filled.append(f"{sheet_name} R{row_num} {cat}")

        # 合计
        if sum_row and total_orig:
            _safe_write(ws, sum_row, col_orig, total_orig)
            _safe_write(ws, sum_row, col_acct_depr, total_acct)
            if col_tax_base:
                _safe_write(ws, sum_row, col_tax_base, total_orig)
            _safe_write(ws, sum_row, col_tax_depr, total_tax)

        # 附加合计行（如总计行R41）
        extra_rows = instr.get("total_sum_rows", [])
        for extra_row in extra_rows:
            _safe_write(ws, extra_row, col_orig, total_orig)
            if col_tax_base:
                _safe_write(ws, extra_row, col_tax_base, total_orig)

        # 审核说明
        if note_row and note_template:
            note_text = note_template.format(
                total_orig=total_orig,
                total_acct=total_acct,
                total_tax=total_tax,
                net_adj=_r2(total_tax - total_acct),
            )
            _safe_write(ws, note_row, 1, note_text)

        self.filled.append(f"{sheet_name}: {len(cat_data)} 类资产")

    def _handle_payroll_rows(self, instr: dict):
        """payroll_rows: 薪酬表多列写入"""
        sheet_name = instr.get("sheet")
        ws = _find_sheet(self.wb, sheet_name)
        if not ws:
            return

        entries = instr.get("entries", [])
        sum_row = instr.get("sum_row", 19)

        total_book = total_tax = total_adj = 0
        for entry in entries:
            row = entry.get("row")
            name = entry.get("name")
            book_val = entry.get("book_value", 0)
            adj = self.resolver.adj_by_name.get(name)

            _safe_write(ws, row, 3, _r2(book_val))
            _safe_write(ws, row, 4, _r2(book_val))
            if adj:
                _safe_write(ws, row, 7, _r2(adj.tax_base))
                _safe_write(ws, row, 8, _r2(adj.increase))
                total_book += book_val
                total_tax += adj.tax_base
                total_adj += adj.increase
            else:
                _safe_write(ws, row, 7, _r2(book_val))
                _safe_write(ws, row, 8, 0)
                total_book += book_val
                total_tax += book_val
            self.filled.append(f"{sheet_name} R{row} {name}")

        # 合计行
        _safe_write(ws, sum_row, 3, _r2(total_book))
        _safe_write(ws, sum_row, 4, _r2(total_book))
        _safe_write(ws, sum_row, 7, _r2(total_tax))
        _safe_write(ws, sum_row, 8, _r2(total_adj))


# ============================================================
# 验证器
# ============================================================


class FillValidator:
    """写入后校验器"""

    @staticmethod
    def validate_tie_out(wb, config: dict, result) -> list:
        """执行勾稽校验"""
        issues = []
        checks = config.get("checks", [])

        for check in checks:
            ctype = check.get("type")
            if ctype == "cell_equals":
                sheet_name = check.get("sheet")
                ws = _find_sheet(wb, sheet_name)
                if not ws:
                    issues.append(f"❌ {sheet_name} 未找到")
                    continue
                row = check.get("row")
                col = check.get("col")
                expected = check.get("expected")
                desc = check.get("desc", "")
                actual = ws.cell(row=row, column=col).value
                try:
                    if abs(float(actual or 0) - float(expected or 0)) > 10:
                        issues.append(f"❌ {desc}: 期望{expected}, 实际{actual}")
                    else:
                        issues.append(f"✅ {desc}: {actual} = {expected}")
                except (ValueError, TypeError):
                    issues.append(f"⚠ {desc}: 无法比较 {actual} vs {expected}")

        return issues

    @staticmethod
    def generate_report(filled_items: list, issues: list) -> str:
        """生成校验报告"""
        lines = []
        lines.append("=" * 50)
        lines.append("填充校验报告")
        lines.append("=" * 50)
        lines.append(f"\n已填充 {len(filled_items)} 项:")
        for item in filled_items:
            lines.append(f"  • {item}")
        lines.append(f"\n勾稽校验 {len(issues)} 项:")
        for issue in issues:
            lines.append(f"  {issue}")
        return "\n".join(lines)


# ============================================================
# 主引擎
# ============================================================


class TaxAuditEngine:
    """
    通用税审底稿填充引擎
    不包含任何模板特化逻辑，所有操作来自配置
    """

    def __init__(self, result, assets=None):
        self.result = result
        self.assets = assets or []
        self.resolver = ValueResolver(result, assets)

    def fill_template(self, template_path: str, output_path: str, config: dict):
        """
        填充单个模板文件

        Parameters
        ----------
        template_path : str
            模板文件路径
        output_path : str
            输出文件路径
        config : dict
            模板 YAML 配置字典（已解析）
        """
        src = Path(template_path)
        dst = Path(output_path)
        dst.parent.mkdir(parents=True, exist_ok=True)

        convert = config.get("convert", False)
        if convert:
            # xls → xlsx 格式转换
            temp_path = dst.with_suffix(".tmp.xlsx")
            _xls_to_xlsx_com(template_path, str(temp_path))
            shutil.move(str(temp_path), str(dst))
        elif str(src) != str(dst):
            shutil.copy2(str(src), str(dst))

        wb = load_workbook(str(dst))

        # 执行指令
        instructions = config.get("instructions", [])
        executor = InstructionExecutor(
            wb, self.resolver, config, self.result, self.assets
        )
        filled = executor.execute_all(instructions)

        # 执行校验
        validator = FillValidator()
        issues = validator.validate_tie_out(wb, config, self.result)

        wb.save(str(dst))
        return filled, issues

    def fill_all(self, manifest: dict, template_dir: str, output_dir: str) -> dict:
        """
        按 manifest 填充所有模板

        Parameters
        ----------
        manifest : dict
            业务 manifest 配置
        template_dir : str
            模板文件所在目录
        output_dir : str
            输出目录

        Returns
        -------
        dict
            {模板id: {"path": str, "filled": list, "issues": list}}
        """
        base = Path(output_dir)
        base.mkdir(parents=True, exist_ok=True)

        results = {}
        for tmpl in manifest.get("templates", []):
            tmpl_id = tmpl.get("id")
            filename = tmpl.get("file")
            tmpl_config = tmpl.get("config", {})
            output_name = tmpl.get("output")

            template_path = os.path.join(template_dir, filename)
            if not os.path.exists(template_path):
                print(f"  ⚠ 模板文件不存在: {template_path}")
                continue

            output_path = str(base / output_name)
            print(f"  >> 填充 {tmpl_id}...")
            filled, issues = self.fill_template(template_path, output_path, tmpl_config)

            results[tmpl_id] = {
                "path": output_path,
                "filled": filled,
                "issues": issues,
            }

            status = "✅" if not issues else "⚠"
            print(f"  {status} {output_name} — {len(filled)} 项")

        return results

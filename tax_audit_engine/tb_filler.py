# -*- coding: utf-8 -*-
"""
2-00 会计账簿（试算平衡表）填充器
将TB数据写入2-00会计报表的对应科目行
"""

import sys, os, shutil

sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path
from openpyxl import load_workbook


def _r2(val):
    return round(float(val) if val else 0.0, 2)


def _safe_write(ws, row, col, value):
    from openpyxl.cell.cell import MergedCell

    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                ws.cell(row=mr.min_row, column=mr.min_col).value = value
                return
    try:
        cell.value = value
    except AttributeError:
        pass


# ============================================================
# 2-00 科目行映射
# ============================================================
# 利润表部分
PL_ROW_MAP = {
    11: "一、营业总收入",  # C3=上期金额, C4=本期金额
    12: "营业收入",  # 其中：营业收入
    18: "减：营业总成本",
    19: "营业成本",  # 其中：营业成本
    33: "税金及附加",
    34: "销售费用",  # 销售(营业)费用
    35: "管理费用",
    36: "研发费用",
    37: "财务费用",
    42: "加：其他收益",
    43: "投资收益",
    47: "公允价值变动收益",
    48: "信用减值损失",
    49: "资产减值损失",
    50: "资产处置收益",
    53: "二、营业利润",
    55: "加：营业外收入",
    56: "政府补助",
    57: "减：营业外支出",
    59: "三、利润总额",
    61: "减：所得税费用",
    63: "四、净利润",
}

# 资产负债表部分
BS_ROW_MAP = {
    94: "货币资金",
    97: "以公允价值计量且其变动计入当期损益的金融资产",
    100: "应收票据",
    101: "应收账款",
    103: "预付款项",
    107: "应收利息",
    108: "应收股利",
    110: "其他应收款",
    112: "存货",
    113: "合同资产",
    115: "一年内到期的非流动资产",
    116: "其他流动资产",
    126: "长期股权投资",
    130: "固定资产",
    131: "在建工程",
    137: "无形资产",
    140: "长期待摊费用",
    141: "递延所得税资产",
    148: "短期借款",
    155: "应付票据",
    156: "应付账款",
    157: "预收款项",
    158: "合同负债",
    161: "应付职工薪酬",
    162: "应交税费",
    164: "应付股利",
    165: "其他应付款",
    172: "其他流动负债",
    176: "长期借款",
    177: "应付债券",
    181: "长期应付款",
    183: "预计负债",
    184: "递延收益",
    185: "递延所得税负债",
    192: "实收资本",
    195: "其他权益工具",
    198: "资本公积",
    200: "其他综合收益",
    202: "盈余公积",
    204: "未分配利润",
}

# TB科目名 → 2-00科目名（如有不同）
TB_TO_200 = {
    "主营业务收入": "营业收入",
    "其他业务收入": "营业收入",  # 合并to营业收入
    "营业外收入": "营业外收入",
    "主营业务成本": "营业成本",
    "其他业务成本": "营业成本",
    "税金及附加": "税金及附加",
    "销售费用": "销售费用",
    "管理费用": "管理费用",
    "研发费用": "研发费用",
    "财务费用": "财务费用",
    "资产减值损失": "资产减值损失",
    "公允价值变动收益": "公允价值变动收益",
    "投资收益": "投资收益",
    "资产处置收益": "资产处置收益",
    "其他收益": "其他收益",
    "货币资金": "货币资金",
    "应收账款": "应收账款",
    "预付款项": "预付款项",
    "其他应收款": "其他应收款",
    "存货": "存货",
    "固定资产": "固定资产",
    "在建工程": "在建工程",
    "无形资产": "无形资产",
    "短期借款": "短期借款",
    "应付账款": "应付账款",
    "预收款项": "预收款项",
    "应付职工薪酬": "应付职工薪酬",
    "应交税费": "应交税费",
    "应付股利": "应付股利",
    "其他应付款": "其他应付款",
    "长期借款": "长期借款",
    "长期应付款": "长期应付款",
    "实收资本": "实收资本",
    "资本公积": "资本公积",
    "盈余公积": "盈余公积",
    "未分配利润": "未分配利润",
    "营业外支出": "营业外支出",
}


def find_tb_value(tb, names):
    """从TB中取第一个匹配的值"""
    for name in names:
        v = tb.get(name)
        if v and abs(v) > 0.01:
            return _r2(v)
    return None


def fill_trial_balance(template_path, output_path, result):
    """
    填充2-00 会计账簿（试算平衡表）

    将TB数据写入模板的对应科目行
    - C4列: 本期金额/期末余额 (未审数)
    - C5列: 审定数
    """
    src = Path(template_path)
    dst = Path(output_path)
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(src), str(dst))

    wb = load_workbook(str(dst))
    tb = result.tb

    # 找2-00 sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "2-00" in name or "会计报表" in name:
            sheet_name = name
            break
    if not sheet_name:
        print("  [错误] 未找到2-00 会计报表 sheet")
        return []

    ws = wb[sheet_name]
    filled = []

    # 1. 利润表科目
    # 注意：有些行需要汇总（如营业收入=主营业务收入+其他业务收入）

    # 先构建科目查找映射
    tb_vals = {}
    for k, v in tb.items.items():
        if v and abs(v) > 0.01:
            # 找到2-00里的目标科目
            target = TB_TO_200.get(k, k)
            tb_vals[target] = tb_vals.get(target, 0) + _r2(v)

    # 特殊汇总项
    # 营业收入 = 主营业务收入 + 其他业务收入
    tb_vals["营业收入"] = _r2(tb.revenue_main) + _r2(tb.revenue_other)
    # 营业成本 = 主营业务成本 + 其他业务成本
    tb_vals["营业成本"] = _r2(tb.cost_main) + _r2(tb.cost_other)

    # 写入利润表
    for row_num, label in PL_ROW_MAP.items():
        # 特殊处理
        if label == "一、营业总收入":
            val = tb_vals.get("营业收入", 0)
            if val:
                _safe_write(ws, row_num, 4, val)
                _safe_write(ws, row_num, 5, val)
                filled.append(f"R{row_num}={label}: {val:,.2f}")
            # 也写营业收入行
            _safe_write(ws, row_num + 1, 4, val)
            _safe_write(ws, row_num + 1, 5, val)
            continue

        if label == "减：营业总成本":
            # 汇总：营业成本+税金+销售+管理+研发+财务+减值
            costs = [
                "营业成本",
                "税金及附加",
                "销售费用",
                "管理费用",
                "研发费用",
                "财务费用",
                "资产减值损失",
            ]
            total = sum(tb_vals.get(c, 0) for c in costs)
            if total:
                _safe_write(ws, row_num, 4, total)
                _safe_write(ws, row_num, 5, total)
                filled.append(f"R{row_num}={label}: {total:,.2f}")
            continue

        # 常规匹配
        val = tb_vals.get(label)
        if not val:
            # 尝试找子科目
            for k, v in tb.items.items():
                if v and abs(v) > 0.01 and (label in k or k in label):
                    val = _r2(v)
                    break

        if val:
            _safe_write(ws, row_num, 4, val)
            _safe_write(ws, row_num, 5, val)
            filled.append(f"R{row_num}={label}: {val:,.2f}")

    # 写入资产负债表
    for row_num, label in BS_ROW_MAP.items():
        val = tb_vals.get(label)
        if not val:
            # 尝试模糊匹配
            for k, v in tb.items.items():
                if v and abs(v) > 0.01 and (label in k or k in label):
                    val = _r2(v)
                    break

        if val:
            _safe_write(ws, row_num, 4, val)
            _safe_write(ws, row_num, 5, val)
            filled.append(f"R{row_num}={label}: {val:,.2f}")

    # 写入利润总额（特殊计算）
    profit = _r2(tb.accounting_profit)
    if profit:
        _safe_write(ws, 59, 4, profit)
        _safe_write(ws, 59, 5, profit)

    # 写入营业收入合计行
    rev_total = _r2(tb.revenue_total)
    if rev_total:
        _safe_write(ws, 11, 4, rev_total)
        _safe_write(ws, 11, 5, rev_total)

    wb.save(str(dst))
    return filled


if __name__ == "__main__":
    import sys

    sys.path.insert(0, "E:/Projects/my-workspace")
    from tax_audit_engine.main import build_sample_data
    from tax_audit_engine.calculator import TaxCalculator

    BASE = r"D:\Users\12844\Desktop\业务工作底稿模版\2026_07_04_1-1、中税网-2026年企业所得税纳税申报审核报告及底稿模板-适用于独立纳税企业V1\1-1、中税网-2026年企业所得税纳税申报审核报告及底稿模板-适用于独立纳税企业V1\1、中税网企业所得税汇缴鉴证报告、申报表及工作底稿模板-适用独立纳税企业-必做底稿2026"

    data = build_sample_data()
    data["tb"].items.update(
        {
            "货币资金": 12500000,
            "应收账款": 8500000,
            "预付款项": 1200000,
            "其他应收款": 800000,
            "存货": 18500000,
            "在建工程": 3000000,
            "短期借款": 5000000,
            "应付账款": 9600000,
            "预收款项": 1800000,
            "其他应付款": 1200000,
            "应付股利": 500000,
            "长期借款": 8000000,
            "长期应付款": 2000000,
            "实收资本": 30000000,
            "资本公积": 5000000,
            "盈余公积": 3500000,
            "未分配利润": 8200000,
        }
    )

    calc = TaxCalculator()
    result = calc.calculate(
        tb=data["tb"], enterprise=data["enterprise"], assets=data["assets"]
    )

    template = os.path.join(BASE, "税审工作底稿-1.xls")
    output = "D:/Users/12844/Desktop/税审底稿_2-00_测试.xlsx"

    print("填充 2-00 会计账簿:")
    # 先转为.xlsx再填充（openpyxl不支持.xls）
    import xlrd

    old_wb = xlrd.open_workbook(template)
    from openpyxl import Workbook

    new_wb = Workbook()
    for sn in old_wb.sheet_names():
        old_ws = old_wb.sheet_by_name(sn)
        new_ws = (
            new_wb.create_sheet(title=sn[:31])
            if new_wb.sheetnames != [new_wb.active.title]
            else new_wb.active
        )
        new_ws.title = sn[:31]
        for r in range(old_ws.nrows):
            for c in range(old_ws.ncols):
                cell = old_ws.cell(r, c)
                if cell.ctype == 2:  # number
                    new_ws.cell(row=r + 1, column=c + 1, value=cell.value)
                elif cell.ctype == 3:  # date
                    new_ws.cell(row=r + 1, column=c + 1, value=cell.value)
                else:
                    v = cell.value
                    if v:
                        new_ws.cell(row=r + 1, column=c + 1, value=str(v).strip())
    new_wb.save(output)

    filled = fill_trial_balance(output, output, result)
    wb = load_workbook(output, data_only=True)
    for sn in wb.sheetnames:
        if "2-00" in sn:
            ws = wb[sn]
            print(f"\n验证 [{sn}]:")
            for r in range(8, min(ws.max_row + 1, 65)):
                c1 = ws.cell(row=r, column=1).value or ""
                c4 = ws.cell(row=r, column=4).value
                c5 = ws.cell(row=r, column=5).value
                if c4 or c5:
                    print(
                        f"  R{r}: {str(c1)[:40]:40s} | C4={str(c4):>12s} | C5={str(c5):>12s}"
                    )

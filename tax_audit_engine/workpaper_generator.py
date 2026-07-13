"""
底稿生成器 — 将计算结果输出为结构化Excel工作底稿
不复制原模板格式，按审计底稿逻辑生成清晰表格

样式统一引用 excel_styles.py，消除内联样式重复
"""

from pathlib import Path
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import (
    CalculationResult,
    TaxAdjustment,
    AdjustmentCategory,
    AssetItem,
    EnterpriseInfo,
    TrialBalance,
)
from .excel_styles import (
    F,
    Fill,
    B,
    A,
    NF,
    set_cols,
    header_row,
    data_cell,
    write_row,
    write_row_ex,
    cover_info,
    merge_title,
    section_row,
    set_vcenter,
    freeze_header,
    auto_filter,
    page_setup,
    auto_width,
    highlight_cells,
)


def _r2(val) -> float:
    """统一round到2位小数，保证数值精度"""
    if val is None:
        return 0.0
    return round(float(val), 2)


def _set_row_height_for_multiline(ws, row, values, min_height=30):
    """根据多行内容自动调整行高（按换行符数量）"""
    max_lines = 1
    for v in values:
        if v and isinstance(v, str) and "\n" in v:
            lines = len(v.split("\n"))
            if lines > max_lines:
                max_lines = lines
    if max_lines > 1:
        ws.row_dimensions[row].height = max(min_height, max_lines * 15)


class WorkpaperGenerator:
    """
    税审工作底稿生成器
    基于计算结果，生成结构化Excel底稿
    所有数值写入党使用 _r2() 确保精度
    """

    def __init__(self, result: CalculationResult):
        self.result = result
        self.wb = Workbook()
        self._sheet_index = 0

    def generate(self, output_path: str):
        """生成完整工作底稿Excel"""
        # 1. 封面
        self._create_cover_sheet()

        # 2. 利润表审定表
        self._create_pl_audit_sheet()

        # 3-6. 纳税调整明细表（分类别）
        for cat, title, subtitle in [
            (
                AdjustmentCategory.INCOME,
                "3-01 收入类纳税调整明细表",
                "收入类纳税调整项目",
            ),
            (
                AdjustmentCategory.DEDUCTION,
                "3-02 扣除类纳税调整明细表",
                "扣除类纳税调整项目",
            ),
            (
                AdjustmentCategory.ASSET,
                "3-03 资产类纳税调整明细表",
                "资产类纳税调整项目",
            ),
            (AdjustmentCategory.SPECIAL, "3-04 特殊事项调整明细表", "特殊事项调整项目"),
            (AdjustmentCategory.OVERSEAS, "4 境外税收调整明细表", "境外税收调整项目"),
            (AdjustmentCategory.TAX_INCENTIVE, "5 税收优惠明细表", "税收优惠明细项目"),
            (AdjustmentCategory.PAYMENT, "6 缴纳情况明细表", "缴纳情况调整项目"),
        ]:
            if self._has_adjustments(cat):
                self._create_adjustment_sheet(title, cat, subtitle)

        # 7. 资产折旧测算表（如有资产数据）
        if (
            self.result.enterprise
            and hasattr(self.result, "_assets")
            and self.result._assets
        ):
            self._create_depreciation_sheet()

        # 8. 纳税调整汇总表
        self._create_summary_sheet()

        # 9. 应纳税所得额计算表
        self._create_tax_calculation_sheet()

        # 保存
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)

    # ============================================================
    # 封面
    # ============================================================

    def _create_cover_sheet(self):
        ws = self.wb.active
        ws.title = "封面"
        set_cols(ws, [3, 20, 30, 20, 3])
        ent = self.result.enterprise

        row = 2
        merge_title(ws, row, "企业所得税汇算清缴纳税申报审核工作底稿", 4)
        row = 4
        info_items = [
            ("被审核单位名称", ent.name),
            ("统一社会信用代码", ent.uscc),
            ("所属行业", ent.industry),
            ("纳税年度", f"{ent.tax_year}年度" if ent.tax_year else ""),
            ("法定代表人", ent.legal_rep),
            ("注册地址", ent.address),
            ("从业人数", f"{ent.employee_count}人"),
            ("资产总额", f"{_r2(ent.total_assets):,.2f}元" if ent.total_assets else ""),
            ("申报类型", "独立纳税企业"),
            ("编制日期", ""),
        ]
        row = cover_info(ws, row, info_items)

        row += 2
        section_row(ws, row, "计算结果摘要", 4)
        row += 1
        summary = self.result.summary()
        for label, value in summary.items():
            data_cell(ws, row, 2, label, bold=True)
            if isinstance(value, float):
                data_cell(ws, row, 3, _r2(value), fmt=NF.AMOUNT)
            else:
                data_cell(ws, row, 3, value)
            row += 1

    # ============================================================
    # 利润表审定表
    # ============================================================

    def _create_pl_audit_sheet(self):
        ws = self.wb.create_sheet("2-00 利润表审定表")
        set_cols(ws, [5, 25, 15, 15, 15, 15])

        tb = self.result.tb
        ent = self.result.enterprise

        # 标题
        merge_title(ws, 1, "利润表及纳税申报审核表", 6)
        if ent.name:
            ws.cell(row=2, column=1, value=f"被审核单位: {ent.name}").font = F.SMALL
        if ent.tax_year:
            ws.cell(
                row=2, column=4, value=f"所属年度: {ent.tax_year}年度"
            ).font = F.SMALL

        # 列头
        headers = ["行次", "项目", "账载金额", "税收金额", "调增金额", "调减金额"]
        row = 4
        header_row(ws, row, headers)

        # 利润表项目
        pl_items = [
            ("一、营业收入", tb.revenue_total),
            ("  主营业务收入", tb.revenue_main),
            ("  其他业务收入", tb.revenue_other),
            ("减：营业成本", _r2(tb.cost_main) + _r2(tb.cost_other)),
            ("  主营业务成本", tb.cost_main),
            ("  其他业务成本", tb.cost_other),
            ("  税金及附加", tb.tax_surcharge),
            ("  销售费用", tb.selling_expense),
            ("  管理费用", tb.admin_expense),
            ("  财务费用", tb.finance_expense),
            ("  资产减值损失", tb.asset_impairment),
            ("加：公允价值变动收益", tb.fair_value_change),
            ("  投资收益", tb.investment_income),
            ("  资产处置收益", tb.asset_disposal_income),
            ("  其他收益", tb.get("其他收益", 0)),
            ("  营业外收入", tb.non_operating_income),
            ("减：营业外支出", tb.get("营业外支出", 0)),
            ("二、利润总额", tb.accounting_profit),
        ]

        row = 5
        for i, (name, amount) in enumerate(pl_items, 1):
            is_total = name.startswith("一") or name.startswith("二")
            data_cell(ws, row, 1, i)
            data_cell(ws, row, 2, name, bold=is_total)
            data_cell(ws, row, 3, _r2(amount), fmt=NF.AMOUNT)
            for c in range(4, 7):
                data_cell(ws, row, c, None, fmt=NF.AMOUNT)
            if is_total:
                for c in range(1, 7):
                    ws.cell(row=row, column=c).fill = Fill.SECTION_BG
            row += 1

    # ============================================================
    # 纳税调整明细表（通用模板）
    # ============================================================

    def _create_adjustment_sheet(self, title, category, subtitle):
        ws = self.wb.create_sheet(title[:31])
        set_cols(ws, [5, 6, 28, 18, 18, 18, 18, 40])

        adjustments = self._get_adjustments(category)

        merge_title(ws, 1, title, 8)
        ws.cell(
            row=2, column=1, value=f"{subtitle} — 税法依据及计算过程"
        ).font = F.SMALL

        headers = [
            "行次",
            "类别",
            "项目",
            "账载金额",
            "计税基础",
            "纳税调增",
            "纳税调减",
            "税法依据/计算说明",
        ]
        row = 4
        header_row(ws, row, headers)

        row = 5
        for i, adj in enumerate(adjustments, 1):
            data_cell(ws, row, 1, i)
            data_cell(ws, row, 2, adj.category.value if adj.category else "")
            data_cell(ws, row, 3, adj.item_name, bold=True)

            vals = [
                _r2(adj.book_amount),
                _r2(adj.tax_base),
                _r2(adj.increase),
                _r2(adj.decrease),
            ]
            for c, v in zip([4, 5, 6, 7], vals):
                data_cell(ws, row, c, v, fmt=NF.AMOUNT)

            ref = adj.tax_law_ref or ""
            calc = adj.calculation or ""
            remark = adj.remark or ""
            note_parts = [p for p in [ref, calc, remark] if p]
            note_text = "\n".join(note_parts)
            data_cell(ws, row, 8, note_text, fmt=None)
            ws.cell(row=row, column=8).font = F.SMALL
            ws.cell(row=row, column=8).alignment = A.WRAP

            # 斑马纹
            fill = Fill.ROW_EVEN if i % 2 == 0 else Fill.ROW_ODD
            for c in range(1, 9):
                if ws.cell(row=row, column=c).fill == fill:
                    pass
                cell = ws.cell(row=row, column=c)
                if not cell.fill or cell.fill.start_color.index == "00000000":
                    cell.fill = fill
                cell.border = B.THIN

            # 自动行高（多行内容）
            _set_row_height_for_multiline(ws, row, [note_text])
            row += 1

        # 合计行
        total_book = _r2(sum(a.book_amount for a in adjustments))
        total_increase = _r2(sum(a.increase for a in adjustments))
        total_decrease = _r2(sum(a.decrease for a in adjustments))
        row_data = ["", "", "合计", total_book, 0, total_increase, total_decrease]
        for ci, val in enumerate(row_data, 1):
            data_cell(ws, row, ci, val, bold=True, fmt=NF.AMOUNT)
            ws.cell(row=row, column=ci).fill = Fill.SUMMARY
        row += 1

        freeze_header(ws, 4)

    # ============================================================
    # 资产折旧测算表
    # ============================================================

    def _create_depreciation_sheet(self):
        ws = self.wb.create_sheet("3-03-01 固定资产折旧测算")
        set_cols(ws, [5, 15, 12, 12, 12, 12, 12, 12, 12])

        assets = self.result._assets
        merge_title(ws, 1, "固定资产折旧及纳税调整审核表", 9)

        headers = [
            "行次",
            "资产类别",
            "资产原值",
            "会计折旧",
            "税收折旧",
            "差异(税收-会计)",
            "税法年限",
            "加速折旧",
        ]
        header_row(ws, 3, headers)

        row = 4
        total_orig = total_acct = total_tax = total_diff = 0.0
        for i, asset in enumerate(assets, 1):
            diff = _r2(asset.current_tax_depr) - _r2(asset.current_accounting_depr)
            orig = _r2(asset.original_value)
            acct = _r2(asset.current_accounting_depr)
            tax = _r2(asset.current_tax_depr)
            total_orig += orig
            total_acct += acct
            total_tax += tax
            total_diff += diff

            data_cell(ws, row, 1, i)
            data_cell(ws, row, 2, f"{asset.category}—{asset.name}")
            data_cell(ws, row, 3, orig, fmt=NF.AMOUNT)
            data_cell(ws, row, 4, acct, fmt=NF.AMOUNT)
            data_cell(ws, row, 5, tax, fmt=NF.AMOUNT)
            data_cell(ws, row, 6, diff, fmt=NF.AMOUNT)
            data_cell(
                ws,
                row,
                7,
                f"会计{asset.accounting_life_years}年/税法{asset.tax_life_years}年",
            )
            data_cell(ws, row, 8, "是" if asset.is_accelerated else "否")

            fill = Fill.ROW_EVEN if i % 2 == 0 else Fill.ROW_ODD
            for c in range(1, 9):
                if (
                    not ws.cell(row=row, column=c).fill
                    or ws.cell(row=row, column=c).fill.start_color.index == "00000000"
                ):
                    ws.cell(row=row, column=c).fill = fill
                ws.cell(row=row, column=c).border = B.THIN
            row += 1

        # 合计
        total_vals = [
            None,
            "合计",
            total_orig,
            total_acct,
            total_tax,
            total_diff,
            None,
            None,
        ]
        for ci, v in enumerate(total_vals, 1):
            data_cell(ws, row, ci, v, bold=True, fmt=NF.AMOUNT)
            ws.cell(row=row, column=ci).fill = Fill.SUMMARY
        row += 1

        freeze_header(ws, 3)

    # ============================================================
    # 纳税调整汇总表
    # ============================================================

    def _create_summary_sheet(self):
        ws = self.wb.create_sheet("纳税调整汇总表")
        set_cols(ws, [5, 30, 18, 18, 18])

        merge_title(ws, 1, "企业所得税纳税调整项目汇总表", 5)

        headers = ["行次", "调整类别", "纳税调增金额", "纳税调减金额", "调整项数"]
        header_row(ws, 3, headers)

        categories = [
            (AdjustmentCategory.INCOME, "一、收入类调整项目"),
            (AdjustmentCategory.DEDUCTION, "二、扣除类调整项目"),
            (AdjustmentCategory.ASSET, "三、资产类调整项目"),
            (AdjustmentCategory.SPECIAL, "四、特殊事项调整项目"),
            (AdjustmentCategory.OVERSEAS, "五、境外税收调整项目"),
            (AdjustmentCategory.TAX_INCENTIVE, "六、税收优惠调整项目"),
            (AdjustmentCategory.PAYMENT, "七、缴纳情况调整项目"),
        ]

        row = 4
        grand_inc = grand_dec = grand_count = 0.0
        for cat, label in categories:
            adjustments = self._get_adjustments(cat)
            inc = _r2(sum(a.increase for a in adjustments))
            dec = _r2(sum(a.decrease for a in adjustments))
            cnt = len(adjustments)
            grand_inc += inc
            grand_dec += dec
            grand_count += cnt

            data_cell(ws, row, 1, row - 3)
            data_cell(ws, row, 2, label, bold=True)
            data_cell(ws, row, 3, inc, fmt=NF.AMOUNT)
            data_cell(ws, row, 4, dec, fmt=NF.AMOUNT)
            data_cell(ws, row, 5, cnt)
            ws.cell(row=row, column=5).alignment = A.CENTER
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = B.THIN
            row += 1

        # 合计行
        data_cell(ws, row, 2, "合  计", bold=True)
        data_cell(ws, row, 3, grand_inc, bold=True, fmt=NF.AMOUNT)
        data_cell(ws, row, 4, grand_dec, bold=True, fmt=NF.AMOUNT)
        data_cell(ws, row, 5, int(grand_count), bold=True)
        ws.cell(row=row, column=5).alignment = A.CENTER
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = B.THIN
            ws.cell(row=row, column=c).fill = Fill.SUMMARY

        freeze_header(ws, 3)

    # ============================================================
    # 应纳税所得额计算表
    # ============================================================

    def _create_tax_calculation_sheet(self):
        ws = self.wb.create_sheet("应纳税所得额计算")
        set_cols(ws, [5, 35, 20, 20])

        merge_title(ws, 1, "应纳税所得额及应纳所得税额计算表", 4)
        r = self.result

        items = [
            ("一、会计利润总额", r.accounting_profit, None),
            ("加：纳税调增金额合计", r.total_increase, None),
            ("减：纳税调减金额合计", r.total_decrease, None),
            ("二、应纳税所得额", r.taxable_income, "会计利润+调增-调减"),
            ("三、适用税率", r.tax_rate, "详见税率说明"),
            ("四、应纳所得税额", r.tax_payable, "应纳税所得额×税率"),
            ("减：减免所得税额", r.deducted_tax, None),
            ("五、实际应纳所得税额", r.final_tax, None),
        ]

        headers = ["行次", "项目", "金额", "说明"]
        header_row(ws, 3, headers)

        row = 4
        for i, (label, amount, note) in enumerate(items, 1):
            is_result = label.startswith("二") or label.startswith("五")
            data_cell(ws, row, 1, i)
            data_cell(ws, row, 2, label, bold=is_result)
            if isinstance(amount, float):
                data_cell(ws, row, 3, _r2(amount), bold=is_result, fmt=NF.AMOUNT)
                ws.cell(row=row, column=3).alignment = A.RIGHT
            else:
                data_cell(ws, row, 3, amount)
                if isinstance(amount, float):
                    ws.cell(row=row, column=3).number_format = NF.PCT
            data_cell(ws, row, 4, note or "")
            ws.cell(row=row, column=4).font = F.SMALL
            if is_result:
                for c in range(1, 5):
                    ws.cell(row=row, column=c).fill = Fill.SUMMARY
            for c in range(1, 5):
                ws.cell(row=row, column=c).border = B.THIN
            row += 1

    # ============================================================
    # 辅助方法
    # ============================================================

    def _has_adjustments(self, category: AdjustmentCategory) -> bool:
        return len(self._get_adjustments(category)) > 0

    def _get_adjustments(self, category: AdjustmentCategory) -> List[TaxAdjustment]:
        return [a for a in self.result.adjustments if a.category == category]

    def set_assets(self, assets: List[AssetItem]):
        self.result._assets = assets

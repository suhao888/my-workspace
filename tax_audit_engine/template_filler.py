# -*- coding: utf-8 -*-
"""
template_filler.py — 完整版
按税法口径填充税审底稿模板
"""

import shutil, os, sys, tempfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from .models import CalculationResult, TaxAdjustment, AdjustmentCategory

sys.stdout.reconfigure(encoding="utf-8")


def _xls_to_xlsx_com(xls_path: str, xlsx_path: str):
    """
    用 Excel COM 将 xls 转换为 xlsx，完整保留格式
    FileFormat=51 → xlOpenXMLWorkbook
    """
    import win32com.client as win32

    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(os.path.abspath(xls_path))
    wb.SaveAs(os.path.abspath(xlsx_path), FileFormat=51)
    wb.Close()
    excel.Quit()


def _r2(val):
    return round(float(val) if val else 0.0, 2)


def _safe_write(ws, row, col, value):
    from openpyxl.cell.cell import MergedCell

    try:
        cell = ws.cell(row=row, column=col)
    except AttributeError:
        return
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                ws.cell(row=mr.min_row, column=mr.min_col).value = value
                return
    try:
        cell.value = value
    except AttributeError:
        pass


# ============================================================
# SH审定表 — 税法口径映射
# ============================================================

SH_TAX_MAP = {
    "SH-01货币资金审核表": {
        "tb_keys": ["货币资金"],
        "tax_note": "经核对银行对账单及余额调节表，期末余额与试算平衡表一致，计税基础与账面金额无差异",
        "col_book": 3,
        "col_audit": 5,
    },
    "SH-02存货审核表": {
        "tb_keys": ["存货"],
        "tax_note": "期末余额经盘点核对，已计提的存货跌价准备通过资产减值损失科目纳税调增，计税基础与账面原值一致",
        "col_book": 3,
        "col_audit": 5,
    },
    "SH-03应收账款审核表": {
        "tb_keys": ["应收账款"],
        "tax_note": "坏账准备已通过资产减值损失科目全额纳税调增（企业所得税法第十条），应收账款计税基础为账面余额",
        "col_book": 2,
        "col_audit": 4,
    },
    "SH-04预付帐款审核表": {
        "tb_keys": ["预付款项"],
        "tax_note": "经核查，预付款项均系正常经营所需，不存在已无对应商品/服务的预付账款，计税基础与账面金额一致",
        "col_book": 2,
        "col_audit": 4,
    },
    "SH-05其他应收款审核表": {
        "tb_keys": ["其他应收款"],
        "tax_note": "其他应收款余额经核对，不存在应确认为费用的挂账款项，计税基础与账面金额一致",
        "col_book": 2,
        "col_audit": 4,
    },
    "SH-06在建工程审核表": {
        "tb_keys": ["在建工程"],
        "tax_note": "在建工程计税基础与账面金额一致，关注是否涉及已完工转固未及时处理的情况",
        "col_book": 3,
        "col_audit": 5,
    },
    "SH-07预收账款审核表": {
        "tb_keys": ["预收款项"],
        "tax_note": "预收款项系正常经营预收，其中已符合税法收入确认条件的已结转收入，计税基础与账面金额一致",
        "col_book": 2,
        "col_audit": 4,
    },
    "SH-08应付账款审核表": {
        "tb_keys": ["应付账款"],
        "tax_note": "应付账款余额经核对无异常，不存在无需支付的应付款项（若存在需调增应税所得），计税基础与账面金额一致",
        "col_book": 3,
        "col_audit": 5,
    },
    "SH-09其他应付款审核表": {
        "tb_keys": ["其他应付款"],
        "tax_note": "其他应付款余额经核对，不存在应确认为收入的款项，计税基础与账面金额一致",
        "col_book": 3,
        "col_audit": 5,
    },
    "SH-10应付股利审核表": {
        "tb_keys": ["应付股利"],
        "tax_note": "应付股利为企业已宣告尚未发放的股息，居民企业股东收到的股息适用免税政策（企业所得税法第26条）",
        "col_book": 2,
        "col_audit": 5,
    },
    "SH-11短期借款审核表": {
        "tb_keys": ["短期借款"],
        "tax_note": "短期借款利息支出已通过利息支出科目核算，超同期同类利率部分已纳税调增（实施条例第38条）",
        "col_book": 2,
        "col_audit": 4,
    },
    "SH-12长期借款审核表": {
        "tb_keys": ["长期借款"],
        "tax_note": "长期借款利息支出经审核，资本化与费用化划分正确，超同期同类利率部分已纳税调增",
        "col_book": 2,
        "col_audit": 4,
    },
    "SH-13长期应付款审核表": {
        "tb_keys": ["长期应付款"],
        "tax_note": "长期应付款余额经核对，计税基础与账面金额一致",
        "col_book": 2,
        "col_audit": 4,
    },
    "SH-14实收资本审核表": {
        "tb_keys": ["实收资本"],
        "tax_note": "实收资本经核查与工商登记信息一致，计税基础与账面金额无差异",
        "col_book": 3,
        "col_audit": 5,
    },
    "SH-15资本公积审核表": {
        "tb_keys": ["资本公积"],
        "tax_note": "资本公积增减变动经核查，不存在应确认为应税收入的情况，计税基础与账面金额一致",
        "col_book": 2,
        "col_audit": 6,
    },
    "SH-16盈余公积审核表": {
        "tb_keys": ["盈余公积"],
        "tax_note": "盈余公积系按净利润10%提取，与公司法规定一致，计税基础与账面金额无差异",
        "col_book": 2,
        "col_audit": 6,
    },
    "SH-17未分配利润审核表": {
        "tb_keys": ["未分配利润"],
        "tax_note": "未分配利润经审核与利润分配表勾稽一致，已分配股息已在应付股利科目反映，计税基础与账面金额无差异",
        "col_book": 2,
        "col_audit": 6,
    },
}


# ============================================================
# 2-00 科目行映射
# ============================================================

PL_ROW_MAP = {
    11: "一、营业总收入",
    12: "营业收入",
    18: "减：营业总成本",
    19: "营业成本",
    33: "税金及附加",
    34: "销售费用",
    35: "管理费用",
    36: "研发费用",
    37: "财务费用",
    42: "加：其他收益",
    43: "投资收益",
    47: "公允价值变动收益",
    49: "资产减值损失",
    50: "资产处置收益",
    53: "二、营业利润",
    55: "加：营业外收入",
    57: "减：营业外支出",
    59: "三、利润总额",
}

BS_ROW_MAP = {
    94: "货币资金",
    100: "应收票据",
    101: "应收账款",
    103: "预付款项",
    110: "其他应收款",
    112: "存货",
    126: "长期股权投资",
    130: "固定资产",
    131: "在建工程",
    137: "无形资产",
    148: "短期借款",
    155: "应付票据",
    156: "应付账款",
    157: "预收款项",
    161: "应付职工薪酬",
    162: "应交税费",
    164: "应付股利",
    165: "其他应付款",
    176: "长期借款",
    181: "长期应付款",
    192: "实收资本",
    198: "资本公积",
    202: "盈余公积",
    204: "未分配利润",
}

TB_TO_200 = {
    "主营业务收入": "营业收入",
    "其他业务收入": "营业收入",
    "主营业务成本": "营业成本",
    "其他业务成本": "营业成本",
    "税金及附加": "税金及附加",
    "销售费用": "销售费用",
    "管理费用": "管理费用",
    "研发费用": "研发费用",
    "财务费用": "财务费用",
    "资产减值损失": "资产减值损失",
    "投资收益": "投资收益",
    "公允价值变动收益": "公允价值变动收益",
    "资产处置收益": "资产处置收益",
    "其他收益": "其他收益",
    "营业外收入": "营业外收入",
    "营业外支出": "营业外支出",
    "货币资金": "货币资金",
    "应收账款": "应收账款",
    "预付款项": "预付款项",
    "其他应收款": "其他应收款",
    "存货": "存货",
    "固定资产": "固定资产",
    "在建工程": "在建工程",
    "无形资产": "无形资产",
    "短期借款": "短期借款",
    "应付账款": "应付账款",
    "预收款项": "预收款项",
    "应付职工薪酬": "应付职工薪酬",
    "应交税费": "应交税费",
    "应付股利": "应付股利",
    "其他应付款": "其他应付款",
    "长期借款": "长期借款",
    "长期应付款": "长期应付款",
    "实收资本": "实收资本",
    "资本公积": "资本公积",
    "盈余公积": "盈余公积",
    "未分配利润": "未分配利润",
}


# ============================================================
# 3-03-01 折旧表 资产类别映射
# ============================================================

DEPR_CATEGORY_ROWS = {
    "房屋、建筑物": 10,
    "机器设备": 11,
    "生产器具、工具": 12,
    "运输工具": 13,
    "电子设备": 14,
    "其他设备": 15,
}


class TemplateFiller:
    """模板填充器 — 按税法口径填充税审底稿"""

    def __init__(self, result: CalculationResult):
        self.result = result
        self._enterprise = result.enterprise
        self._assets = []

    def set_assets(self, assets: list):
        """传入固定资产卡片列表，用于折旧表填充"""
        self._assets = assets

    # ============================================================
    # SH审定表填充
    # ============================================================

    def fill_SH_sheets(self, template_path: str, output_path: str) -> List[str]:
        """按税法口径填充SH审定表"""
        src = Path(template_path)
        dst = Path(output_path)
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(src), str(dst))

        import openpyxl

        wb = openpyxl.load_workbook(str(dst))
        tb = self.result.tb
        filled = []

        # 调整项税法口径速查
        adj_tax_map = {a.item_name: a.tax_base for a in self.result.adjustments}

        for sh_name, info in SH_TAX_MAP.items():
            if sh_name not in wb.sheetnames:
                continue

            ws = wb[sh_name]
            tb_keys = info["tb_keys"]
            col_book = info["col_book"]
            col_audit = info["col_audit"]

            # 从TB取值（去重）
            values = []
            for key in tb_keys:
                v = tb.get(key)
                if v and abs(v) > 0.01:
                    rv = _r2(v)
                    dup = False
                    for _, ev in values:
                        if abs(ev - rv) < 1.0:
                            dup = True
                            break
                    if not dup:
                        values.append((key, rv))

            if not values:
                continue

            total_book = sum(v for _, v in values)

            # ===== 数据行区域 =====
            sum_row = None
            for r in range(8, ws.max_row + 1):
                cv = ws.cell(row=r, column=1).value
                if cv and isinstance(cv, str) and "合" in cv and "计" in cv:
                    sum_row = r
                    break
            if not sum_row:
                sum_row = ws.max_row - 3
            data_start = 8

            # 写入行（仅写账载金额col_book，审核确认额列有公式自动计算）
            row = data_start
            for label, val in values:
                if row >= sum_row:
                    break
                _safe_write(ws, row, 1, row - data_start + 1)
                _safe_write(ws, row, 2, label)
                _safe_write(ws, row, col_book, val)
                row += 1

            # 合计行（仅写账载金额）
            if sum_row and sum_row <= ws.max_row:
                _safe_write(ws, sum_row, col_book, total_book)

            # 审核说明
            note_row = None
            for r in range(max(sum_row or row, row), ws.max_row + 1):
                cv = ws.cell(row=r, column=1).value
                if cv and isinstance(cv, str) and ("来源" in cv or "数据来源" in cv):
                    note_row = r
                    break
            if note_row:
                _safe_write(
                    ws,
                    note_row,
                    1,
                    f"审核说明：{info['tax_note']}。数据来源于客户提供的年度账表及已审计的年度财务报告，经审核余额可以确认。",
                )
            filled.append(sh_name)

        wb.save(str(dst))
        return filled

    # ============================================================
    # 2-00 会计账簿（试算平衡表）填充
    # ============================================================

    def fill_trial_balance(self, template_path: str, output_path: str) -> List[str]:
        """填充2-00会计账簿"""
        src = Path(template_path)
        dst = Path(output_path)
        if str(src) != str(dst):
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(str(src), str(dst))

        import openpyxl

        wb = openpyxl.load_workbook(str(dst))

        # 找2-00 sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "2-00" in name or "会计报表" in name:
                sheet_name = name
                break
        if not sheet_name:
            return []
        ws = wb[sheet_name]
        tb = self.result.tb
        filled = []

        # 构建科目映射
        tb_vals = {}
        for k, v in tb.items.items():
            if v and abs(v) > 0.01:
                target = TB_TO_200.get(k, k)
                tb_vals[target] = tb_vals.get(target, 0) + _r2(v)

        # 特殊汇总
        tb_vals["营业收入"] = _r2(tb.revenue_main) + _r2(tb.revenue_other)
        tb_vals["营业成本"] = _r2(tb.cost_main) + _r2(tb.cost_other)

        # 写入利润表
        for row_num, label in PL_ROW_MAP.items():
            if label == "一、营业总收入":
                val = tb_vals.get("营业收入", 0)
                if val:
                    _safe_write(ws, row_num, 4, val)
                    _safe_write(ws, row_num, 5, val)
                    _safe_write(ws, row_num + 1, 4, val)
                    _safe_write(ws, row_num + 1, 5, val)
                    filled.append(f"R{row_num}={val:,.0f}")
                continue
            if label == "减：营业总成本":
                costs = [
                    "营业成本",
                    "税金及附加",
                    "销售费用",
                    "管理费用",
                    "研发费用",
                    "财务费用",
                    "资产减值损失",
                ]
                total = sum(tb_vals.get(c, 0) for c in costs)
                if total:
                    _safe_write(ws, row_num, 4, total)
                    _safe_write(ws, row_num, 5, total)
                    filled.append(f"R{row_num}={total:,.0f}")
                continue
            val = tb_vals.get(label)
            if val:
                _safe_write(ws, row_num, 4, val)
                _safe_write(ws, row_num, 5, val)
                filled.append(f"R{row_num}={val:,.0f}")

        # 写入资产负债表
        for row_num, label in BS_ROW_MAP.items():
            val = tb_vals.get(label)
            if not val:
                for k, v in tb.items.items():
                    if v and abs(v) > 0.01 and (label in k or k in label):
                        val = _r2(v)
                        break
            if val:
                _safe_write(ws, row_num, 4, val)
                _safe_write(ws, row_num, 5, val)
                filled.append(f"R{row_num}={val:,.0f}")

        # 利润总额
        profit = _r2(tb.accounting_profit)
        if profit:
            _safe_write(ws, 59, 4, profit)
            _safe_write(ws, 59, 5, profit)

        wb.save(str(dst))
        return filled

    # ============================================================
    # 3-03-01 固定资产折旧审核表填充
    # ============================================================

    def fill_depreciation(self, template_path: str, output_path: str) -> List[str]:
        """填充3-03-01固定资产折旧及纳税调整审核表"""
        src = Path(template_path)
        dst = Path(output_path)
        if str(src) != str(dst):
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(str(src), str(dst))

        import openpyxl

        wb = openpyxl.load_workbook(str(dst))

        # 找3-03-01 sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "3-03-01" in name:
                sheet_name = name
                break
        if not sheet_name:
            return []
        ws = wb[sheet_name]

        assets = self._assets
        if not assets:
            return []

        filled = []
        # 按类别汇总
        cat_data = {}
        for a in assets:
            cat = a.category
            if cat not in cat_data:
                cat_data[cat] = {"orig": 0, "acct_depr": 0, "tax_depr": 0}
            cat_data[cat]["orig"] += _r2(a.original_value)
            cat_data[cat]["acct_depr"] += _r2(a.current_accounting_depr)
            cat_data[cat]["tax_depr"] += _r2(a.current_tax_depr)

        # 写入各资产类别行
        total_orig = total_acct = total_tax = 0
        for cat, row_num in DEPR_CATEGORY_ROWS.items():
            data = cat_data.get(cat)
            if not data:
                continue
            # C2=类别标签(已预置), C5=资产原值, C6=本年会计折旧, C9=税收折旧
            _safe_write(ws, row_num, 5, data["orig"])
            _safe_write(ws, row_num, 6, data["acct_depr"])
            _safe_write(ws, row_num, 9, data["tax_depr"])
            total_orig += data["orig"]
            total_acct += data["acct_depr"]
            total_tax += data["tax_depr"]
            filled.append(
                f"R{row_num} {cat}: {data['orig']:,.0f}/{data['acct_depr']:,.0f}/{data['tax_depr']:,.0f}"
            )

        # 合计行(R9: 固定资产合计)
        if total_orig:
            _safe_write(ws, 9, 5, total_orig)
            _safe_write(ws, 9, 6, total_acct)
            _safe_write(ws, 9, 9, total_tax)

        # 审核说明
        _safe_write(
            ws,
            29,
            1,
            f"一、审核项目说明及结论：本企业固定资产合计{total_orig:,.0f}元，会计折旧{total_acct:,.0f}元，税收折旧{total_tax:,.0f}元，纳税调整{_r2(total_tax - total_acct):,.0f}元。该数据来源于企业年度账表，经审核确认上述金额。",
        )

        wb.save(str(dst))
        return filled

    # ============================================================
    # 辅助底稿 — 汇总调整表填充
    # ============================================================

    def fill_fuzhu_digao(self, template_path: str, output_path: str) -> List[str]:
        """填充辅助底稿（汇总调整表37项调整条目）"""
        src = Path(template_path)
        dst = Path(output_path)
        if str(src) != str(dst):
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(str(src), str(dst))

        import openpyxl

        wb = openpyxl.load_workbook(str(dst))
        filled = []
        tb = self.result.tb
        adj_by_name = {a.item_name: a for a in self.result.adjustments}

        # === 汇总调整表 ===
        if "汇总调整表" in wb.sheetnames:
            ws = wb["汇总调整表"]
            # R3-R8: 收入类调整 (行1-5)
            # R9-R29: 扣除类调整 (行6-25)
            # R30-R35: 资产类调整 (行26-30)
            # R36-R38: 特殊事项 (行31-32)
            # R39-R43: 免税减计 (行33-36)
            # R44-R45: 其他 (行37)
            # R46: 合计, R47: 利润总额, R48: 纳税调整金额, R49: 调整后所得

            # 行->调整项映射
            row_map = {
                11: "工资薪金",
                12: "职工福利费",
                13: "职工教育经费",
                14: "工会经费",
                15: "基本社会保险",
                16: "住房公积金",
                17: "补充养老保险",
                18: "补充医疗保险",
                20: "业务招待费",
                21: "广告费和业务宣传费",
                22: "捐赠支出",
                23: "税收滞纳金",
                24: "罚金、罚款",
                25: "跨期费用",
                31: "资产折旧",
                32: "无形资产摊销",
                33: "资产减值损失",
                40: "研发费用",
                41: "残疾人工资",
            }
            for row, item_name in row_map.items():
                adj = adj_by_name.get(item_name)
                if not adj:
                    # 模糊匹配
                    for a in self.result.adjustments:
                        if item_name in a.item_name or a.item_name in item_name:
                            adj = a
                            break
                if adj:
                    _safe_write(ws, row, 5, _r2(adj.book_amount))  # 账载金额
                    _safe_write(ws, row, 7, _r2(adj.tax_base))  # 税收金额
                    if adj.increase > 0:
                        _safe_write(ws, row, 8, _r2(adj.increase))  # 调整金额
                    elif adj.decrease > 0:
                        _safe_write(ws, row, 8, _r2(-adj.decrease))
                    filled.append(f"R{row} {item_name}={adj.increase:.0f}")

            # R47: 利润总额
            profit = _r2(tb.accounting_profit)
            if profit:
                _safe_write(ws, 47, 7, profit)
                filled.append(f"R47 利润总额={profit}")
            # R48: 纳税调整金额
            net_adj = _r2(self.result.total_increase - self.result.total_decrease)
            _safe_write(ws, 48, 8, net_adj)
            # R49: 调整后所得
            tax_income = _r2(self.result.taxable_income)
            _safe_write(ws, 49, 8, tax_income)
            filled.append(f"R49 调整后所得={tax_income}")

        # === 交换意见表 ===
        if "交换意见表" in wb.sheetnames:
            ws = wb["交换意见表"]
            e = self._enterprise
            # R3: 被审核单位名称
            if e.name:
                _safe_write(ws, 3, 1, f"被审核单位名称（公章）：{e.name}")

        wb.save(str(dst))
        return filled

    # ============================================================
    # A类申报表 — 综合填充
    # ============================================================

    def fill_A_declaration(self, template_path: str, output_path: str) -> List[str]:
        """填充A类申报表（核心表A100000/A105000/A105050 + 利润表/资产负债表）"""
        src = Path(template_path)
        dst = Path(output_path)
        if str(src) != str(dst):
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(str(src), str(dst))

        import openpyxl

        wb = openpyxl.load_workbook(str(dst))
        filled = []
        tb = self.result.tb
        e = self._enterprise
        adj_by_name = {a.item_name: a for a in self.result.adjustments}
        assets = self._assets or []

        def _find(sheet_name):
            for sn in wb.sheetnames:
                if sn.strip() == sheet_name.strip():
                    return wb[sn]
            for sn in wb.sheetnames:
                if sheet_name.replace(" ", "") in sn.replace(" ", ""):
                    return wb[sn]
            return None

        # ===== 封面 =====
        ws = _find("封面")
        if ws:
            if e.name:
                _safe_write(ws, 3, 3, e.name)
            if e.tax_year:
                _safe_write(ws, 4, 3, f"{e.tax_year}年度")
            filled.append("封面")

        # ===== A000000 基础信息表 =====
        ws = _find("A000000")
        if ws:
            if e.name:
                _safe_write(ws, 4, 3, e.name)
            if e.uscc:
                _safe_write(ws, 5, 3, e.uscc)
            if e.industry:
                _safe_write(ws, 7, 3, e.industry)
            if e.employee_count:
                _safe_write(ws, 13, 3, e.employee_count)
            if e.total_assets:
                _safe_write(ws, 14, 3, e.total_assets)
            filled.append("A000000")

        # ===== 利润表 =====
        ws = _find("利润表")
        if ws and tb:
            pl_map = {
                4: tb.revenue_total,  # 营业总收入
                5: tb.revenue_total,  # 营业收入
                11: tb.cost_main + tb.cost_other,  # 营业成本
                24: tb.tax_surcharge,  # 税金及附加
                25: tb.selling_expense,  # 销售费用
                26: tb.admin_expense,  # 管理费用
                27: tb.r_and_d_expense,  # 研发费用
                28: tb.finance_expense,  # 财务费用
                33: tb["其他收益"],  # 其他收益
                34: tb.investment_income,  # 投资收益
                39: tb.fair_value_change,  # 公允价值变动
                41: tb.asset_impairment,  # 资产减值损失
                42: tb.asset_disposal_income,  # 资产处置收益
                44: tb.non_operating_income,  # 营业外收入
                46: tb["营业外支出"],  # 营业外支出
                47: tb.accounting_profit,  # 利润总额
            }
            for row, val in pl_map.items():
                if val and abs(val) > 0.01:
                    _safe_write(ws, row, 2, _r2(val))
            filled.append("利润表")

        # ===== 资产负债表 =====
        for bs_name in ["资产负债表", "资产负债表（续）"]:
            ws = _find(bs_name)
            if not ws:
                continue
            bs_map = {
                4: "货币资金",
                5: "应收账款",
                6: "预付款项",
                7: "其他应收款",
                8: "存货",
                15: "固定资产",
                16: "在建工程",
                17: "无形资产",
                27: "短期借款",
                29: "应付账款",
                30: "预收款项",
                31: "应付职工薪酬",
                32: "应交税费",
                33: "应付股利",
                34: "其他应付款",
                38: "长期借款",
                39: "长期应付款",
                42: "实收资本",
                43: "资本公积",
                44: "盈余公积",
                45: "未分配利润",
            }
            for row, key in bs_map.items():
                val = tb.get(key)
                if val:
                    _safe_write(ws, row, 2, _r2(val))
            filled.append(bs_name)

        # ===== A100000 纳税申报主表 =====
        ws = _find("A100000")
        if ws:
            a100_map = {
                6: tb.revenue_total,  # 营业收入
                7: tb.cost_main + tb.cost_other,  # 营业成本
                8: tb.tax_surcharge,  # 税金及附加
                9: tb.selling_expense,  # 销售费用
                10: tb.admin_expense,  # 管理费用
                11: tb.r_and_d_expense,  # 研发费用
                12: tb.finance_expense,  # 财务费用
                13: tb["其他收益"],  # 其他收益
                14: tb.investment_income,  # 投资收益
                16: tb.fair_value_change,  # 公允价值变动
                18: tb.asset_impairment,  # 资产减值损失
                19: tb.asset_disposal_income,  # 资产处置收益
                21: tb.non_operating_income,  # 营业外收入
                22: tb["营业外支出"],  # 营业外支出
                23: tb.accounting_profit,  # 利润总额
                25: self.result.total_increase,  # 纳税调增
                26: self.result.total_decrease,  # 纳税调减
                37: self.result.taxable_income,  # 应纳税所得额
                38: 0.25,  # 税率
                39: self.result.tax_payable,  # 应纳所得税额
                46: self.result.final_tax,  # 实际应纳
            }
            for row, val in a100_map.items():
                if (
                    val is not None
                    and abs(val if isinstance(val, (int, float)) else 0) > 0.01
                ):
                    _safe_write(ws, row, 4, _r2(val))
            filled.append("A100000")

        # ===== A105000 纳税调整项目明细表 =====
        ws = _find("A105000")
        if ws:
            # 调整项行映射: R20(职工薪酬), R21(业务招待费), R22(广告费),
            # R23(捐赠), R24(利息), R25(罚金), R26(滞纳金),
            # R27(赞助), R33(资产减值), R38(资产折旧)
            a105_row_map = {
                20: "职工薪酬",
                21: "业务招待费",
                22: "广告费和业务宣传费",
                23: "捐赠支出",
                24: "利息支出",
                25: "罚金、罚款",
                26: "税收滞纳金",
                27: "赞助支出",
                33: "资产减值损失",
                38: "资产折旧",
            }
            for row, name in a105_row_map.items():
                adj = adj_by_name.get(name)
                if not adj:
                    for a in self.result.adjustments:
                        if name in a.item_name or a.item_name in name:
                            adj = a
                            break
                if adj:
                    _safe_write(ws, row, 3, _r2(adj.book_amount))  # 账载金额
                    _safe_write(ws, row, 4, _r2(adj.tax_base))  # 税收金额
                    # C5=调增金额, C6=调减金额
                    if adj.increase > 0:
                        _safe_write(ws, row, 5, _r2(adj.increase))
                    elif adj.decrease > 0:
                        _safe_write(ws, row, 6, _r2(adj.decrease))
                    filled.append(f"A105000 R{row} {name}")

            # 合计行 R59
            _safe_write(ws, 59, 5, _r2(self.result.total_increase))
            _safe_write(ws, 59, 6, _r2(self.result.total_decrease))

        # ===== A105050 职工薪酬表 =====
        ws = _find("A105050")
        if ws:
            # R7: 工资薪金, R9: 福利费, R10: 教育经费,
            # R13: 工会经费, R14: 社保, R15: 公积金,
            # R16: 补充养老, R17: 补充医疗
            pay_map = {
                7: ("工资薪金", 8500000),
                9: ("职工福利费", 1300000),
                10: ("职工教育经费", 750000),
                13: ("工会经费", 180000),
                14: ("基本社会保险", 1785000),
                15: ("住房公积金", 680000),
                16: ("补充养老保险", 300000),
                17: ("补充医疗保险", 200000),
            }
            total_book = total_tax = total_adj = 0
            for row, (name, book_val) in pay_map.items():
                adj = adj_by_name.get(name)
                _safe_write(ws, row, 3, _r2(book_val))  # 账载金额
                _safe_write(ws, row, 4, _r2(book_val))  # 实际发生额
                if adj:
                    _safe_write(ws, row, 7, _r2(adj.tax_base))  # 税收金额
                    _safe_write(ws, row, 8, _r2(adj.increase))  # 纳税调整
                    total_book += book_val
                    total_tax += adj.tax_base
                    total_adj += adj.increase
                else:
                    _safe_write(ws, row, 7, _r2(book_val))
                    _safe_write(ws, row, 8, 0)
                    total_book += book_val
                    total_tax += book_val
                filled.append(f"A105050 R{row} {name}")

            # R19 合计
            _safe_write(ws, 19, 3, _r2(total_book))
            _safe_write(ws, 19, 4, _r2(total_book))
            _safe_write(ws, 19, 7, _r2(total_tax))
            _safe_write(ws, 19, 8, _r2(total_adj))

        # ===== A105080 资产折旧摊销表 =====
        ws = _find("A105080")
        if ws and assets:
            cat_data = {}
            for a in assets:
                cat = a.category
                if cat not in cat_data:
                    cat_data[cat] = {"orig": 0, "acct_depr": 0, "tax_depr": 0}
                cat_data[cat]["orig"] += _r2(a.original_value)
                cat_data[cat]["acct_depr"] += _r2(a.current_accounting_depr)
                cat_data[cat]["tax_depr"] += _r2(a.current_tax_depr)

            # R9-R14: 固定资产类别
            cat_rows = {
                "房屋、建筑物": 9,
                "机器设备": 10,
                "生产器具、工具": 11,
                "运输工具": 12,
                "电子设备": 13,
                "其他设备": 14,
            }
            total_orig = total_acct = total_tax = 0
            for cat, row in cat_rows.items():
                data = cat_data.get(cat)
                if not data:
                    continue
                _safe_write(ws, row, 5, data["orig"])
                _safe_write(ws, row, 6, data["acct_depr"])
                _safe_write(ws, row, 8, data["orig"])  # 计税基础=原值
                _safe_write(ws, row, 9, data["tax_depr"])
                total_orig += data["orig"]
                total_acct += data["acct_depr"]
                total_tax += data["tax_depr"]

            # R8: 固定资产合计
            if total_orig:
                _safe_write(ws, 8, 5, total_orig)
                _safe_write(ws, 8, 6, total_acct)
                _safe_write(ws, 8, 8, total_orig)
                _safe_write(ws, 8, 9, total_tax)

            # R41: 总计
            _safe_write(ws, 41, 5, total_orig)
            _safe_write(ws, 41, 8, total_orig)
            filled.append("A105080")

        wb.save(str(dst))
        return filled

    # ============================================================
    # 一键填充入口
    # ============================================================


def fill_all_templates(result, template_dir, output_dir=None, assets=None):
    """
    一键填充所有底稿模板

    Parameters
    ----------
    result : CalculationResult
        税审计算结果
    template_dir : str
        模板目录路径
    output_dir : str, optional
        输出目录，默认桌面
    assets : list[AssetItem], optional
        固定资产卡片列表

    Returns
    -------
    dict
        {文件类型: 输出路径}
    """
    outputs = {}
    base = output_dir or "D:/Users/12844/Desktop"
    filler = TemplateFiller(result)
    if assets:
        filler.set_assets(assets)

    # 模板路径
    fp0 = os.path.join(template_dir, "税审工作底稿-0.xlsx")
    fp1 = os.path.join(template_dir, "税审工作底稿-1.xls")
    fp2 = os.path.join(template_dir, "税审工作底稿-2.xls")
    fp_fuzhu = os.path.join(
        template_dir, "2026年企业所得税汇算清缴纳税调整汇总表-辅助底稿-选做.xlsx"
    )
    fp_a100 = os.path.join(template_dir, "企业所得税年度纳税申报表A类2017版.xlsx")

    # 1. SH审定表
    if os.path.exists(fp0):
        out0 = os.path.join(base, "税审底稿_SH审定表.xlsx")
        filler.fill_SH_sheets(fp0, out0)
        outputs["SH审定表"] = out0
        print(f"  ✅ SH审定表 → {out0}")

    # 2. 2-00会计账簿（Excel COM 转换，保留全部格式）
    if os.path.exists(fp1):
        temp_xlsx = os.path.join(base, "_temp_2-00.xlsx")
        print(f"  >> 转换 2-00: {fp1}")
        _xls_to_xlsx_com(fp1, temp_xlsx)
        out1 = os.path.join(base, "税审底稿_2-00会计账簿.xlsx")
        filler.fill_trial_balance(temp_xlsx, out1)
        outputs["2-00会计账簿"] = out1
        print(f"  ✅ 2-00会计账簿 → {out1}")
        if os.path.exists(temp_xlsx):
            os.remove(temp_xlsx)

    # 3. 折旧审核表（Excel COM 转换，保留全部格式）
    if os.path.exists(fp2):
        temp_xlsx2 = os.path.join(base, "_temp_3-03-01.xlsx")
        print(f"  >> 转换 3-03-01: {fp2}")
        _xls_to_xlsx_com(fp2, temp_xlsx2)
        out2 = os.path.join(base, "税审底稿_折旧审核表.xlsx")
        filler.fill_depreciation(temp_xlsx2, out2)
        outputs["折旧审核表"] = out2
        print(f"  ✅ 折旧审核表 → {out2}")
        if os.path.exists(temp_xlsx2):
            os.remove(temp_xlsx2)

    # 4. 辅助底稿（选做）
    if os.path.exists(fp_fuzhu):
        out4 = os.path.join(base, "税审底稿_辅助底稿.xlsx")
        filler.fill_fuzhu_digao(fp_fuzhu, out4)
        outputs["辅助底稿"] = out4
        print(f"  ✅ 辅助底稿 → {out4}")

    # 5. A类申报表
    if os.path.exists(fp_a100):
        out5 = os.path.join(base, "税审底稿_A类申报表.xlsx")
        filler.fill_A_declaration(fp_a100, out5)
        outputs["A类申报表"] = out5
        print(f"  ✅ A类申报表 → {out5}")

    return outputs

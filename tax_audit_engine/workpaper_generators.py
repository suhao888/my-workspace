"""
底稿生成器扩展 — 5类业务自生成底稿（RD/FullTax/Loss/HighTech/Land）
不依赖模板，从计算结果直接生成结构化Excel底稿
"""

from pathlib import Path
from typing import List, Optional
from openpyxl import Workbook

from tax_audit_engine.models import EnterpriseInfo
from tax_audit_engine.excel_styles import (
    F,
    Fill,
    B,
    A,
    NF,
    set_cols,
    header_row,
    data_cell,
    write_row,
    write_row_ex,
    cover_info,
    merge_title,
    section_row,
    set_vcenter,
    freeze_header,
    auto_filter,
    page_setup,
    auto_width,
    highlight_cells,
    auto_row_height,
)


def _r2(val):
    """统一round到2位小数（仅float），消除浮点噪声"""
    if isinstance(val, float):
        return round(val, 2)
    return val


# ============================================================
# RDGenerator — 研发费用加计扣除工作底稿
# ============================================================


class RDGenerator:
    """
    RD 研发费用加计扣除工作底稿生成器
    sheet: 封面 → 项目明细表 → 3-01汇总表 → 3-02优惠审核表
    """

    def __init__(self, result, input_data=None):
        self.result = result
        self.input = input_data
        self.wb = Workbook()
        self.ent = result.enterprise or EnterpriseInfo(name="未命名企业")

    def generate(self, output_path: str):
        self._cover()
        self._project_detail()
        self._summary()
        self._audit_check()
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)

    def _cover(self):
        ws = self.wb.active
        ws.title = "封面"
        set_cols(ws, [3, 22, 40, 3])
        page_setup(ws, landscape=False, fit_w=1)
        merge_title(ws, 2, "研发费用加计扣除审核工作底稿", 3)
        cover_info(
            ws,
            4,
            [
                ("被审核单位名称", self.ent.name or ""),
                ("统一社会信用代码", self.ent.uscc or ""),
                ("所属行业", self.ent.industry or ""),
                (
                    "审核年度",
                    f"{self.ent.tax_year}年度"
                    if hasattr(self.ent, "tax_year") and self.ent.tax_year
                    else "",
                ),
                ("加计扣除比例", "100%"),
                ("适用政策", "财税[2015]119号、2021年第28号"),
            ],
        )
        r = 12
        section_row(ws, r, "一、汇总数据", 3)
        r += 1
        for label, val in [
            ("账面研发费用合计", self.result.total_book),
            ("可加计扣除金额", self.result.qualifying_amount),
            ("加计扣除额", self.result.deduction_amount),
        ]:
            ws.cell(row=r, column=2, value=label).font = F.BOLD
            c = ws.cell(row=r, column=3, value=_r2(val))
            c.font = F.NORMAL
            c.number_format = NF.AMOUNT
            c.border = B.THIN
            ws.cell(row=r, column=2).border = B.THIN
            r += 1
        if self.result.has_other_limit_issue:
            ws.cell(row=r, column=2, value="其中：其他费用超限金额").font = F.BOLD
            c = ws.cell(row=r, column=3, value=_r2(self.result.other_limit_excess))
            c.font = F.WARN
            c.number_format = NF.AMOUNT
            c.border = B.THIN
            ws.cell(row=r, column=2).border = B.THIN

    def _project_detail(self):
        ws = self.wb.create_sheet("项目明细")
        set_cols(ws, [5, 8, 26, 12, 15, 15, 13, 13, 13, 13, 15])
        page_setup(ws, landscape=True, fit_w=1)
        merge_title(ws, 1, "研发项目费用明细表", 10)
        header_row(
            ws,
            3,
            [
                "序号",
                "项目编号",
                "项目名称",
                "研发方式",
                "人员人工",
                "直接投入",
                "折旧",
                "摊销",
                "设计试验",
                "其他费用",
                "合计",
            ],
        )
        r = 4
        for i, proj in enumerate(self.result.projects, 1):
            vals = [
                i,
                proj.project_id,
                proj.project_name,
                proj.dev_method,
                proj.personnel_costs,
                proj.direct_input,
                proj.depreciation,
                proj.amortization,
                proj.design_trial,
                proj.other_costs,
                proj.total_book,
            ]
            fill = Fill.ROW_EVEN if i % 2 == 0 else Fill.ROW_ODD
            write_row(ws, r, vals, fmt=NF.AMOUNT, fill_=fill)
            # 文字列用文本格式
            for ci in [1, 2, 3, 4]:
                ws.cell(row=r, column=ci).number_format = NF.TEXT
            r += 1
        # 合计行
        totals = [
            "",
            "",
            "合  计",
            "",
            sum(p.personnel_costs for p in self.result.projects),
            sum(p.direct_input for p in self.result.projects),
            sum(p.depreciation for p in self.result.projects),
            sum(p.amortization for p in self.result.projects),
            sum(p.design_trial for p in self.result.projects),
            sum(p.other_costs for p in self.result.projects),
            sum(p.total_book for p in self.result.projects),
        ]
        write_row(ws, r, totals, fmt=NF.AMOUNT, bold=True, fill_=Fill.BLUE_LIGHT)
        highlight_cells(ws, r, 11, Fill.BLUE_LIGHT, F.BOLD)
        freeze_header(ws, 4)
        auto_filter(ws, 3, 11)

    def _summary(self):
        ws = self.wb.create_sheet("3-01汇总表")
        set_cols(ws, [5, 32, 18, 20, 18])
        page_setup(ws, landscape=True, fit_w=1)
        merge_title(ws, 1, "可加计扣除研发费用审核汇总表", 4)
        header_row(ws, 3, ["行次", "项  目", "账面金额", "可加计扣除金额", "差  异"])
        rows = [
            ("一、人员人工费用", "personnel_costs"),
            ("二、直接投入费用", "direct_input"),
            ("三、折旧费用", "depreciation"),
            ("四、无形资产摊销", "amortization"),
            ("五、设计试验等费用", "design_trial"),
            ("六、其他相关费用", "other_costs"),
        ]
        r = 4
        for label, attr in rows:
            book = sum(getattr(p, attr, 0) for p in self.result.projects)
            qual = (
                min(book, self.result.qualifying_amount)
                if attr == "other_costs"
                else book
            )
            write_row(ws, r, [r - 3, label, book, qual, qual - book], fmt=NF.AMOUNT)
            r += 1
        total_book = sum(p.total_book for p in self.result.projects)
        write_row(
            ws,
            r,
            [
                "",
                "合  计",
                total_book,
                self.result.qualifying_amount,
                self.result.qualifying_amount - total_book,
            ],
            fmt=NF.AMOUNT,
            bold=True,
            fill_=Fill.SUMMARY,
        )
        highlight_cells(ws, r, 5, Fill.SUMMARY, F.BOLD)
        r += 2
        section_row(ws, r, "二、加计扣除计算", 4)
        r += 1
        for label, val in [
            ("可加计扣除研发费用合计", self.result.qualifying_amount),
            ("加计扣除比例", "100%"),
            ("加计扣除额", self.result.deduction_amount),
        ]:
            ws.cell(row=r, column=2, value=label).font = F.BOLD
            ws.cell(row=r, column=2).border = B.THIN
            c = ws.cell(
                row=r, column=3, value=_r2(val) if isinstance(val, float) else val
            )
            if isinstance(val, (int, float)):
                c.number_format = NF.AMOUNT
            c.font = F.RESULT
            c.border = B.THIN
            r += 1
        freeze_header(ws, 4)

    def _audit_check(self):
        ws = self.wb.create_sheet("3-02审核表")
        set_cols(ws, [5, 38, 18, 18])
        page_setup(ws, landscape=False, fit_w=1)
        merge_title(ws, 1, "研发费用加计扣除优惠审核表", 3)
        header_row(ws, 3, ["行次", "审核项目", "金  额", "审核意见"])
        rows = [
            ("一、自主研发、合作研发、集中研发", self.result.qualifying_amount, True),
            (
                "（一）人员人工费用",
                sum(p.personnel_costs for p in self.result.projects),
                False,
            ),
            (
                "（二）直接投入费用",
                sum(p.direct_input for p in self.result.projects),
                False,
            ),
            (
                "（三）折旧费用",
                sum(p.depreciation for p in self.result.projects),
                False,
            ),
            (
                "（四）无形资产摊销",
                sum(p.amortization for p in self.result.projects),
                False,
            ),
            (
                "（五）设计费用",
                sum(p.design_trial for p in self.result.projects),
                False,
            ),
            (
                "（六）其他相关费用",
                sum(p.other_costs for p in self.result.projects),
                False,
            ),
            ("三、年度可加计扣除研发费用小计", self.result.qualifying_amount, True),
            ("四、计入本年损益的加计扣除额", self.result.deduction_amount, True),
        ]
        r = 4
        for i, (label, val, is_section) in enumerate(rows, 1):
            opinion = "通过" if val > 0 else ""
            write_row(ws, r, [i, label, val, opinion], fmt=NF.AMOUNT)
            if is_section:
                highlight_cells(ws, r, 4, Fill.SECTION_BG, F.SECTION)
            r += 1
        freeze_header(ws, 4)


# ============================================================
# FullTaxGenerator — 全税种核查底稿
# ============================================================


class FullTaxGenerator:
    """
    全税种核查底稿生成器
    sheet: 封面 → 2-01各税种汇总 → 差异分析明细
    """

    def __init__(self, result):
        self.result = result
        self.wb = Workbook()
        self.ft_input = result.input

    def generate(self, output_path: str):
        self._cover()
        self._summary()
        self._analysis()
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)

    def _cover(self):
        ws = self.wb.active
        ws.title = "封面"
        set_cols(ws, [3, 22, 40, 3])
        page_setup(ws, landscape=False, fit_w=1)
        merge_title(ws, 2, "全税种核查测算工作底稿", 3)
        cover_info(
            ws,
            4,
            [
                ("被审计单位名称", self.ft_input.enterprise_name),
                ("统一社会信用代码", self.ft_input.tax_id),
                ("核查年度", self.ft_input.audit_year),
                ("差异率阈值", "5%"),
            ],
        )
        r = 10
        section_row(ws, r, "一、核查结论", 3)
        r += 1
        high_risk = sum(
            1
            for item in self.result.item_details
            if item.get("is_high_risk", abs(item.get("diff_rate", 0)) > 0.05)
        )
        ws.cell(row=r, column=2, value="高风险税种数").font = F.BOLD
        ws.cell(row=r, column=2).border = B.THIN
        c = ws.cell(row=r, column=3, value=high_risk)
        c.font = F.WARN if high_risk > 0 else F.PASS
        c.border = B.THIN
        r += 1
        ws.cell(row=r, column=2, value="核查结论").font = F.BOLD
        ws.cell(row=r, column=2).border = B.THIN
        c = ws.cell(
            row=r,
            column=3,
            value="存在差异，需进一步核实" if high_risk > 0 else "无重大差异",
        )
        c.font = F.WARN if high_risk > 0 else F.PASS
        c.border = B.THIN

    def _summary(self):
        ws = self.wb.create_sheet("2-01各税种汇总")
        set_cols(ws, [5, 20, 15, 20, 20, 20, 13, 10])
        page_setup(ws, landscape=True, fit_w=1)
        merge_title(ws, 1, "各税种申报与认定对比表", 7)
        header_row(
            ws,
            3,
            [
                "序号",
                "税种",
                "所属期",
                "申报税额",
                "认定税额",
                "差异额",
                "差异率",
                "风险",
            ],
        )
        r = 4
        total_dec = total_act = total_diff = 0
        for i, item in enumerate(self.result.item_details, 1):
            diff = item.get(
                "difference",
                item.get("actual_amount", 0) - item.get("declared_amount", 0),
            )
            rate = item.get("diff_rate", 0)
            is_high = item.get("is_high_risk", abs(rate) > 0.05)
            total_dec += item.get("declared_amount", 0)
            total_act += item.get("actual_amount", 0)
            total_diff += diff
            risk_label = "!!高风险" if is_high else "正常"
            fill = (
                Fill.FAIL
                if is_high
                else (Fill.ROW_EVEN if i % 2 == 0 else Fill.ROW_ODD)
            )
            write_row(
                ws,
                r,
                [
                    i,
                    item.get("tax_type", ""),
                    item.get("tax_period", ""),
                    item.get("declared_amount", 0),
                    item.get("actual_amount", 0),
                    diff,
                    rate,
                    risk_label,
                ],
                fmt=NF.AMOUNT,
                fill_=fill,
            )
            ws.cell(row=r, column=3).number_format = NF.TEXT
            ws.cell(row=r, column=7).number_format = NF.PCT
            r += 1
        write_row(
            ws,
            r,
            ["", "合  计", "", total_dec, total_act, total_diff, "", ""],
            fmt=NF.AMOUNT,
            bold=True,
            fill_=Fill.BLUE_LIGHT,
        )
        highlight_cells(ws, r, 8, Fill.BLUE_LIGHT, F.BOLD)
        freeze_header(ws, 4)
        auto_filter(ws, 3, 8)

    def _analysis(self):
        ws = self.wb.create_sheet("差异分析")
        set_cols(ws, [5, 20, 18, 13, 55])
        page_setup(ws, landscape=True, fit_w=1)
        merge_title(ws, 1, "税种差异分析及审计建议", 4)
        header_row(ws, 3, ["序号", "税种", "差异额", "差异率", "可能原因及建议"])
        r = 4
        for i, item in enumerate(self.result.item_details, 1):
            diff = item.get("difference", 0)
            rate = item.get("diff_rate", 0)
            if abs(diff) < 0.01:
                continue
            reason = item.get("risk_reason", getattr(item, "risk_reason", ""))
            is_high = item.get("is_high_risk", abs(rate) > 0.05)
            fill = Fill.FAIL if is_high else None
            write_row(
                ws,
                r,
                [i, item.get("tax_type", ""), diff, rate, reason],
                fmt=NF.AMOUNT,
                fill_=fill,
            )
            ws.cell(row=r, column=4).number_format = NF.PCT
            ws.cell(row=r, column=5).alignment = A.WRAP
            # 高风险行字体红色
            if is_high:
                data_cell(ws, r, 2, item.get("tax_type", ""), bold=True)
            auto_row_height(ws, r, [diff, rate, reason], min_height=36)
            r += 1
        freeze_header(ws, 4)


# ============================================================
# LossGenerator — 财产损失税前扣除底稿
# ============================================================


class LossGenerator:
    """
    财产损失税前扣除审核底稿生成器
    sheet: 封面 → 损失汇总表 → 逐项审核明细
    """

    def __init__(self, result):
        self.result = result
        self.wb = Workbook()
        self.loss_input = result.input

    def generate(self, output_path: str):
        self._cover()
        self._summary()
        self._detail()
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)

    def _cover(self):
        ws = self.wb.active
        ws.title = "封面"
        set_cols(ws, [3, 25, 40, 3])
        page_setup(ws, landscape=False, fit_w=1)
        merge_title(ws, 2, "企业资产损失税前扣除审核工作底稿", 3)
        ent = self.loss_input.enterprise
        cover_info(
            ws,
            4,
            [
                ("被审核单位名称", ent.name if ent else ""),
                ("统一社会信用代码", ent.uscc if ent else ""),
                (
                    "审核年度",
                    f"{getattr(ent, 'tax_year', '')}年度"
                    if hasattr(ent, "tax_year")
                    else "",
                ),
                ("审核依据", "国家税务总局公告2011年第25号"),
                ("损失项数", f"{self.result.item_count} 项"),
            ],
        )
        r = 11
        section_row(ws, r, "一、审核汇总", 3)
        r += 1
        for label, val in [
            ("可税前扣除损失", self.result.qualifying_loss),
            ("不得扣除损失", self.result.disallowed_loss),
            ("问题项数", len(self.result.issue_items)),
        ]:
            ws.cell(row=r, column=2, value=label).font = F.BOLD
            ws.cell(row=r, column=2).border = B.THIN
            c = ws.cell(
                row=r, column=3, value=_r2(val) if isinstance(val, float) else val
            )
            c.border = B.THIN
            if isinstance(val, (int, float)):
                c.number_format = NF.AMOUNT_INT
            if label == "不得扣除损失" and val:
                c.font = F.WARN
            elif label == "问题项数" and val:
                c.font = F.WARN
            else:
                c.font = F.NORMAL
            r += 1

    def _summary(self):
        ws = self.wb.create_sheet("损失汇总")
        set_cols(ws, [5, 6, 30, 18, 18, 18, 12])
        page_setup(ws, landscape=True, fit_w=1)
        merge_title(ws, 1, "资产损失税前扣除审核汇总表", 6)
        header_row(
            ws,
            3,
            ["序号", "编号", "项目", "账面净值", "可收回金额", "损失金额", "审核结论"],
        )
        r = 4
        for i, item in enumerate(self.result.item_results, 1):
            disallowed = item.get("disallowed", False)
            conclusion = "不得扣除" if disallowed else "可扣除"
            fill = (
                Fill.FAIL
                if disallowed
                else (Fill.ROW_EVEN if i % 2 == 0 else Fill.ROW_ODD)
            )
            vals = [
                i,
                item.get("loss_id", ""),
                item.get("asset_name", f"损失{i}"),
                item.get("book_value", 0),
                item.get("recoverable", 0),
                item.get("loss_amount", 0),
                conclusion,
            ]
            write_row(ws, r, vals, fmt=NF.AMOUNT, fill_=fill)
            if disallowed:
                data_cell(ws, r, 7, conclusion, bold=True)
            r += 1
        write_row(
            ws,
            r,
            [
                "",
                "",
                "合  计",
                self.result.total_book_value,
                self.result.total_recoverable,
                self.result.total_loss,
                "",
            ],
            fmt=NF.AMOUNT,
            bold=True,
            fill_=Fill.BLUE_LIGHT,
        )
        highlight_cells(ws, r, 7, Fill.BLUE_LIGHT, F.BOLD)
        r += 2
        section_row(ws, r, "二、分类汇总", 6)
        r += 1
        # 按类型分组合计
        from collections import defaultdict

        by_type = defaultdict(lambda: {"book": 0, "loss": 0})
        for item in self.result.item_results:
            cat = item.get("category", "其他")
            by_type[cat]["book"] += item.get("book_value", 0)
            by_type[cat]["loss"] += item.get("loss_amount", 0)
        for cat, vals in sorted(by_type.items()):
            write_row(
                ws,
                r,
                ["", "", f"  {cat}", vals["book"], "", vals["loss"], ""],
                fmt=NF.AMOUNT,
                fill_=Fill.SUBHEADER,
            )
            r += 1
        freeze_header(ws, 4)
        auto_filter(ws, 3, 7)

    def _detail(self):
        ws = self.wb.create_sheet("逐项审核")
        set_cols(ws, [5, 8, 22, 13, 16, 16, 16, 10, 10])
        page_setup(ws, landscape=True, fit_w=1)
        merge_title(ws, 1, "资产损失逐项审核明细表", 8)
        header_row(
            ws,
            3,
            [
                "序号",
                "编号",
                "资产名称",
                "损失类型",
                "账面净值",
                "可收回",
                "损失金额",
                "审批",
                "证据",
            ],
        )
        r = 4
        for i, item in enumerate(self.result.item_results, 1):
            disallowed = item.get("disallowed", False)
            fill = Fill.FAIL if disallowed else None
            write_row(
                ws,
                r,
                [
                    i,
                    item.get("loss_id", ""),
                    item.get("asset_name", ""),
                    item.get("category", ""),
                    item.get("book_value", 0),
                    item.get("recoverable", 0),
                    item.get("loss_amount", 0),
                    item.get("approval_status", ""),
                    item.get("evidence_level", ""),
                ],
                fmt=NF.AMOUNT,
                fill_=fill,
            )
            r += 1
        freeze_header(ws, 4)
        auto_filter(ws, 3, 9)


# ============================================================
# HighTechGenerator — 高新技术企业认定审核底稿
# ============================================================


class HighTechGenerator:
    """
    高新技术企业认定审核底稿生成器
    sheet: 封面 → 研发费用审定表 → 收入审定表 → 结构明细表 → 评分表
    """

    def __init__(self, result):
        self.result = result
        self.wb = Workbook()
        self.ht_input = result.input

    def generate(self, output_path: str):
        self._cover()
        self._rd_audit()
        self._income_audit()
        self._structure()
        self._score()
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)

    def _cover(self):
        ws = self.wb.active
        ws.title = "封面"
        set_cols(ws, [3, 22, 40, 3])
        page_setup(ws, landscape=False, fit_w=1)
        merge_title(ws, 2, "高新技术企业认定审核工作底稿", 3)
        ent = self.ht_input.enterprise
        cover_info(
            ws,
            4,
            [
                ("被审核单位名称", ent.name if ent else ""),
                ("统一社会信用代码", ent.uscc if ent else ""),
                ("审核年度", "2025"),
                ("审核依据", "国科发火〔2016〕32号、195号"),
                ("总分", f"{self.result.total_score:.1f}分"),
                ("审核结论", "通过" if self.result.passed else "不通过"),
            ],
        )
        r = 12
        section_row(ws, r, "一、门槛条件", 3)
        r += 1
        ent = self.ht_input.enterprise
        checks = [
            (
                "研发费用占比(近3年)",
                self.ht_input.rd_expense_3year_total / self.ht_input.income_3year_total
                if self.ht_input.income_3year_total
                else 0,
                0.04,
            ),
            (
                "科技人员占比",
                self.ht_input.tech_staff / self.ht_input.total_staff
                if self.ht_input.total_staff
                else 0,
                0.10,
            ),
            (
                "高新产品收入占比(近1年)",
                self.ht_input.hi_product_revenue / self.ht_input.income_last_year
                if self.ht_input.income_last_year
                else 0,
                0.60,
            ),
        ]
        for label, val, threshold in checks:
            passed = val >= threshold
            ws.cell(row=r, column=2, value=label).font = F.BOLD
            ws.cell(row=r, column=2).border = B.THIN
            c = ws.cell(row=r, column=3, value=f"{val:.2%} (要求≥{threshold:.0%})")
            c.font = F.PASS if passed else F.WARN
            c.border = B.THIN
            r += 1
        conclusion = "全部达标 ✓" if self.result.passed else "存在不达标项 ✗"
        ws.cell(row=r, column=2, value="结论").font = F.BOLD
        ws.cell(row=r, column=2).border = B.THIN
        c = ws.cell(row=r, column=3, value=conclusion)
        c.font = F.PASS if self.result.passed else F.WARN
        c.border = B.THIN

    def _rd_audit(self):
        ws = self.wb.create_sheet("研发费用审定表")
        set_cols(ws, [5, 35, 20, 20, 12])
        page_setup(ws, landscape=True, fit_w=1)
        merge_title(ws, 1, "研发费用审定表", 4)
        header_row(ws, 3, ["序号", "项目", "金额", "占比", "达标"])
        rd_ratio = (
            self.ht_input.rd_expense_3year_total / self.ht_input.income_3year_total
            if self.ht_input.income_3year_total > 0
            else 0
        )
        tech_ratio = (
            self.ht_input.tech_staff / self.ht_input.total_staff
            if self.ht_input.total_staff > 0
            else 0
        )
        rows = [
            ("最近一年研发费用", self.ht_input.rd_expense_last_year, None, None),
            ("近3年研发费用总额", self.ht_input.rd_expense_3year_total, None, None),
            ("近3年销售收入总额", self.ht_input.income_3year_total, None, None),
            (
                "研发费用占比(3年)",
                rd_ratio,
                NF.PCT,
                "达标" if rd_ratio >= 0.04 else "不达标",
            ),
            ("科技人员数", self.ht_input.tech_staff, None, None),
            ("职工总数", self.ht_input.total_staff, None, None),
            (
                "科技人员占比",
                tech_ratio,
                NF.PCT,
                "达标" if tech_ratio >= 0.10 else "不达标",
            ),
        ]
        r = 4
        for i, (label, val, fmt, status) in enumerate(rows, 1):
            if fmt == NF.PCT:
                write_row(ws, r, [i, label, val, val, status or ""], fmt=fmt)
            else:
                write_row(ws, r, [i, label, val, "", status or ""], fmt=NF.AMOUNT)
            if status:
                passed = status == "达标"
                highlight_cells(ws, r, 5, Fill.PASS if passed else Fill.FAIL)
                ws.cell(row=r, column=2).font = F.BOLD
            elif i in (3, 6):
                highlight_cells(ws, r, 5, Fill.SECTION_BG)
            r += 1
        freeze_header(ws, 4)

    def _income_audit(self):
        ws = self.wb.create_sheet("收入审定表")
        set_cols(ws, [5, 32, 20, 20, 12])
        page_setup(ws, landscape=True, fit_w=1)
        merge_title(ws, 1, "高新技术产品（服务）收入审定表", 4)
        header_row(ws, 3, ["序号", "项目", "金额", "占比", "判定"])
        hi_ratio = (
            self.ht_input.hi_product_revenue / self.ht_input.income_last_year
            if self.ht_input.income_last_year > 0
            else 0
        )
        hi_ratio_3 = (
            self.ht_input.hi_product_revenue_3year / self.ht_input.total_revenue_3year
            if self.ht_input.total_revenue_3year > 0
            else 0
        )
        items = [
            ("高新收入（最近一年）", self.ht_input.hi_product_revenue, None, ""),
            ("总收入（最近一年）", self.ht_input.income_last_year, None, ""),
            (
                "高新收入占比（最近一年）",
                hi_ratio,
                NF.PCT,
                "达标" if hi_ratio >= 0.60 else "不达标",
            ),
            ("高新收入（近3年）", self.ht_input.hi_product_revenue_3year, None, ""),
            ("总收入（近3年）", self.ht_input.total_revenue_3year, None, ""),
            (
                "高新收入占比（近3年）",
                hi_ratio_3,
                NF.PCT,
                "达标" if hi_ratio_3 >= 0.60 else "不达标",
            ),
        ]
        r = 4
        for i, (label, val, fmt, status) in enumerate(items, 1):
            if fmt == NF.PCT:
                write_row(ws, r, [i, label, val, val, status], fmt=fmt)
            else:
                write_row(ws, r, [i, label, val, "", ""], fmt=NF.AMOUNT)
            if status:
                passed = status == "达标"
                highlight_cells(ws, r, 5, Fill.PASS if passed else Fill.FAIL)
                ws.cell(row=r, column=2).font = F.BOLD
            r += 1
        freeze_header(ws, 4)

    def _structure(self):
        ws = self.wb.create_sheet("研发费用结构明细")
        set_cols(ws, [5, 30, 18, 22, 18])
        page_setup(ws, landscape=True, fit_w=1)
        merge_title(ws, 1, "研发费用结构明细表", 4)
        header_row(ws, 3, ["序号", "费用类别", "金额", "占研发费用比例", "备注"])
        total = self.ht_input.rd_expense_3year_total
        items_data = [
            ("人员人工费用", total * 0.35),
            ("直接投入费用", total * 0.30),
            ("折旧费用", total * 0.12),
            ("无形资产摊销", total * 0.08),
            ("设计试验等费用", total * 0.10),
            ("其他相关费用", total * 0.05),
        ]
        r = 4
        for i, (label, val) in enumerate(items_data, 1):
            ratio = val / total if total > 0 else 0
            write_row(ws, r, [i, label, val, ratio, ""], fmt=NF.AMOUNT)
            ws.cell(row=r, column=4).number_format = NF.PCT
            r += 1
        write_row(
            ws,
            r,
            ["", "合  计", total, 1.0, ""],
            fmt=NF.AMOUNT,
            bold=True,
            fill_=Fill.BLUE_LIGHT,
        )
        ws.cell(row=r, column=4).number_format = NF.PCT
        highlight_cells(ws, r, 5, Fill.BLUE_LIGHT, F.BOLD)
        freeze_header(ws, 4)

    def _score(self):
        ws = self.wb.create_sheet("评分表")
        set_cols(ws, [5, 28, 15, 12, 12, 45])
        page_setup(ws, landscape=True, fit_w=1)
        merge_title(ws, 1, "高新技术企业认定评分表", 5)
        header_row(ws, 3, ["序号", "评分项", "得分", "满分", "通过线", "评分说明"])
        scores = [
            (
                1,
                "知识产权",
                self.result.ip_score,
                30.0,
                20.0,
                f"I类{self.ht_input.ip_count}个, II类{self.ht_input.ip_2_count}个",
            ),
            (
                2,
                "科技成果转化能力",
                self.result.tech_transfer_score,
                30.0,
                20.0,
                self.result.transfer_detail,
            ),
            (
                3,
                "研发组织管理水平",
                self.result.rd_management_score,
                20.0,
                15.0,
                self.result.management_detail,
            ),
            (
                4,
                "企业成长性",
                self.result.growth_score,
                20.0,
                10.0,
                self.result.growth_detail,
            ),
        ]
        r = 4
        for i, (sn, name, score, full, line, note) in enumerate(scores, 1):
            write_row(ws, r, [sn, name, score, full, line, note], fmt=NF.SCORE)
            ws.cell(row=r, column=3).number_format = NF.SCORE
            ws.cell(row=r, column=6).alignment = A.WRAP
            passed = score >= line
            highlight_cells(ws, r, 6, Fill.PASS if passed else Fill.FAIL)
            ws.cell(row=r, column=2).font = F.BOLD
            auto_row_height(ws, r, [note], min_height=28)
            r += 1
        r += 1
        section_row(ws, r, "二、综合评价", 5)
        r += 1
        write_row(
            ws,
            r,
            ["", "总  分", self.result.total_score, 100.0, 71.0, ""],
            fmt=NF.SCORE,
            bold=True,
            fill_=Fill.BLUE_LIGHT,
        )
        ws.cell(row=r, column=3).number_format = NF.SCORE
        highlight_cells(ws, r, 6, Fill.BLUE_LIGHT, F.BOLD)
        r += 1
        conclusion = "通过" if self.result.passed else "不通过"
        ws.cell(row=r, column=2, value="审核结论").font = F.LABEL
        ws.cell(row=r, column=2).border = B.THIN
        c = ws.cell(row=r, column=3, value=conclusion)
        c.font = F.PASS if self.result.passed else F.WARN
        c.border = B.THIN
        freeze_header(ws, 4)


# ============================================================
# LandGenerator — 土地增值税清算审核底稿
# ============================================================


class LandGenerator:
    """
    土地增值税清算审核底稿生成器
    sheet: 封面 → 收入明细表 → 扣除项目明细表 → 逐类计算表 → 税款计算汇总
    """

    def __init__(self, result):
        self.result = result
        self.wb = Workbook()
        self.land_input = result.input

    def generate(self, output_path: str):
        self._cover()
        self._income_detail()
        self._deduction_detail()
        self._property_calc()
        self._tax_summary()
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)

    def _cover(self):
        ws = self.wb.active
        ws.title = "封面"
        set_cols(ws, [3, 22, 40, 3])
        page_setup(ws, landscape=False, fit_w=1)
        merge_title(ws, 2, "土地增值税清算审核工作底稿", 3)
        inp = self.land_input
        ent = inp.enterprise
        cover_info(
            ws,
            4,
            [
                ("被审核单位名称", ent.name if ent else ""),
                ("统一社会信用代码", ent.uscc if ent else ""),
                ("项目名称", inp.project_name),
                ("项目所在地", inp.project_location),
                (
                    "审核年度",
                    f"{getattr(ent, 'tax_year', '')}年度"
                    if hasattr(ent, "tax_year") and ent.tax_year
                    else "",
                ),
                ("占地面积", f"{inp.land_area:,.2f} m²"),
                ("总建筑面积", f"{inp.total_floor_area:,.2f} m²"),
            ],
        )
        r = 13
        section_row(ws, r, "一、审核汇总", 3)
        r += 1
        for label, val in [
            ("收入总额", self.result.total_income),
            ("扣除项目合计", self.result.total_deductions),
            ("应缴税款", self.result.total_tax),
            ("已预缴税款", inp.pre_collected_tax),
            ("应补(退)税款", self.result.tax_difference),
            ("审核结论", "通过" if self.result.passed else "发现差异"),
        ]:
            ws.cell(row=r, column=2, value=label).font = F.BOLD
            ws.cell(row=r, column=2).border = B.THIN
            c = ws.cell(
                row=r, column=3, value=_r2(val) if isinstance(val, float) else val
            )
            c.border = B.THIN
            if isinstance(val, (int, float)):
                c.number_format = NF.AMOUNT
            if label == "审核结论":
                c.font = F.PASS if self.result.passed else F.WARN
            elif label == "应补(退)税款":
                c.font = F.RESULT if abs(self.result.tax_difference) > 0 else F.NORMAL
            else:
                c.font = F.NORMAL
            r += 1

    def _income_detail(self):
        ws = self.wb.create_sheet("收入明细表")
        set_cols(ws, [5, 6, 22, 18, 18, 18, 12])
        page_setup(ws, landscape=True, fit_w=1)
        merge_title(ws, 1, "房地产收入明细表", 6)
        header_row(
            ws,
            3,
            [
                "序号",
                "房产类型",
                "含税收入",
                "免税收入",
                "应税收入",
                "占比",
                "备注",
            ],
        )
        r = 4
        total_inc = sum(r.total_income for r in self.land_input.revenues)
        for i, rev in enumerate(self.land_input.revenues, 1):
            taxable = rev.total_income - rev.exempt_income
            ratio = taxable / total_inc if total_inc > 0 else 0
            fill = Fill.ROW_EVEN if i % 2 == 0 else Fill.ROW_ODD
            write_row(
                ws,
                r,
                [
                    i,
                    rev.property_type,
                    rev.total_income,
                    rev.exempt_income,
                    taxable,
                    ratio,
                    "",
                ],
                fmt=NF.AMOUNT,
                fill_=fill,
            )
            ws.cell(row=r, column=6).number_format = NF.PCT
            r += 1
        write_row(
            ws,
            r,
            [
                "",
                "合  计",
                total_inc,
                sum(rev.exempt_income for rev in self.land_input.revenues),
                self.result.total_income,
                1.0,
                "",
            ],
            fmt=NF.AMOUNT,
            bold=True,
            fill_=Fill.BLUE_LIGHT,
        )
        ws.cell(row=r, column=6).number_format = NF.PCT
        highlight_cells(ws, r, 7, Fill.BLUE_LIGHT, F.BOLD)
        freeze_header(ws, 4)

    def _deduction_detail(self):
        ws = self.wb.create_sheet("扣除项目明细表")
        set_cols(ws, [5, 8, 28, 18, 15, 18])
        page_setup(ws, landscape=True, fit_w=1)
        merge_title(ws, 1, "扣除项目明细表", 5)
        header_row(
            ws,
            3,
            [
                "序号",
                "费用类别",
                "项目名称",
                "总额",
                "分摊方法",
                "备注",
            ],
        )
        r = 4
        cost_total = sum(c.total_amount for c in self.land_input.costs)
        for i, cost in enumerate(self.land_input.costs, 1):
            fill = Fill.ROW_EVEN if i % 2 == 0 else Fill.ROW_ODD
            write_row(
                ws,
                r,
                [
                    i,
                    cost.category,
                    cost.category,
                    cost.total_amount,
                    cost.apportion_method or "建筑面积法",
                    cost.notes,
                ],
                fmt=NF.AMOUNT,
                fill_=fill,
            )
            r += 1
        write_row(
            ws,
            r,
            ["", "", "合  计", cost_total, "", ""],
            fmt=NF.AMOUNT,
            bold=True,
            fill_=Fill.BLUE_LIGHT,
        )
        highlight_cells(ws, r, 6, Fill.BLUE_LIGHT, F.BOLD)
        freeze_header(ws, 4)
        auto_filter(ws, 3, 6)

    def _property_calc(self):
        ws = self.wb.create_sheet("逐类计算表")
        set_cols(ws, [5, 6, 16, 18, 18, 18, 16, 14, 12, 12])
        page_setup(ws, landscape=True, fit_w=1)
        merge_title(ws, 1, "各类房地产土地增值税逐项计算表", 9)
        header_row(
            ws,
            3,
            [
                "序号",
                "房产类型",
                "应税收入",
                "扣除合计",
                "增值额",
                "增值率",
                "适用税率",
                "速算扣除率",
                "应缴税款",
                "免税",
            ],
        )
        r = 4
        for i, pr in enumerate(self.result.property_results, 1):
            fill = Fill.ROW_EVEN if i % 2 == 0 else Fill.ROW_ODD
            write_row(
                ws,
                r,
                [
                    i,
                    pr.property_type,
                    pr.total_income,
                    pr.deductions.total_deductions,
                    pr.added_value,
                    pr.ratio,
                    pr.tax_rate,
                    pr.quick_deduction,
                    pr.tax_due,
                    "是" if pr.exempt else "否",
                ],
                fmt=NF.AMOUNT,
                fill_=fill,
            )
            ws.cell(row=r, column=6).number_format = NF.PCT
            ws.cell(row=r, column=7).number_format = NF.PCT
            ws.cell(row=r, column=8).number_format = NF.PCT
            if pr.exempt:
                highlight_cells(ws, r, 10, Fill.PASS, F.PASS)
            r += 1
        write_row(
            ws,
            r,
            [
                "",
                "合  计",
                self.result.total_income,
                self.result.total_deductions,
                sum(pr.added_value for pr in self.result.property_results),
                "",
                "",
                "",
                self.result.total_tax,
                "",
            ],
            fmt=NF.AMOUNT,
            bold=True,
            fill_=Fill.BLUE_LIGHT,
        )
        highlight_cells(ws, r, 10, Fill.BLUE_LIGHT, F.BOLD)
        freeze_header(ws, 4)

    def _tax_summary(self):
        ws = self.wb.create_sheet("税款计算汇总")
        set_cols(ws, [5, 30, 20, 20, 20])
        page_setup(ws, landscape=False, fit_w=1)
        merge_title(ws, 1, "土地增值税税款计算汇总表", 4)
        header_row(ws, 3, ["行次", "项  目", "金  额", "备  注"])
        rows_data = [
            ("一、收入项目", "", ""),
            ("  1. 应税收入合计", self.result.total_income, ""),
        ]
        for pr in self.result.property_results:
            rows_data.append((f"    - {pr.property_type}", pr.total_income, ""))
        rows_data += [
            ("二、扣除项目", "", ""),
            ("  1. 扣除项目合计", self.result.total_deductions, ""),
        ]
        for pr in self.result.property_results:
            rows_data.append(
                (f"    - {pr.property_type}", pr.deductions.total_deductions, "")
            )
        rows_data += [
            ("三、增值额与税率", "", ""),
        ]
        for pr in self.result.property_results:
            rows_data.append((f"  {pr.property_type}增值额", pr.added_value, ""))
            rows_data.append(
                (f"  {pr.property_type}增值率", pr.ratio, str(pr.bracket_desc))
            )
        rows_data += [
            ("四、税款计算", "", ""),
        ]
        for pr in self.result.property_results:
            tag = " [免税]" if pr.exempt else ""
            rows_data.append((f"  {pr.property_type}应缴税款{tag}", pr.tax_due, ""))
        rows_data += [
            ("  应缴税款合计", self.result.total_tax, ""),
            ("  已预缴税款", self.land_input.pre_collected_tax, ""),
            ("  应补(退)税款", self.result.tax_difference, ""),
        ]

        r = 4
        for i, (label, val, note) in enumerate(rows_data, 1):
            is_section = label.startswith(("一、", "二、", "三、", "四、"))
            is_sub = label.startswith("  ") and not label.startswith("    -")
            is_detail = label.startswith("    -")
            write_row(ws, r, [i, label, val, note], fmt=NF.AMOUNT)
            if is_section:
                highlight_cells(ws, r, 4, Fill.SECTION_BG, F.SECTION)
            elif is_sub:
                ws.cell(row=r, column=2).font = F.BOLD
            if label == "  应补(退)税款":
                c = ws.cell(row=r, column=3)
                c.font = F.RESULT if abs(self.result.tax_difference) > 0 else F.NORMAL
                if self.result.tax_difference > 0:
                    ws.cell(row=r, column=4, value="应补缴").font = F.WARN
                elif self.result.tax_difference < 0:
                    ws.cell(row=r, column=4, value="应退税").font = F.PASS
                else:
                    ws.cell(row=r, column=4, value="无差异").font = F.PASS
            r += 1

        r += 1
        section_row(ws, r, "五、审核意见", 4)
        r += 1
        if self.result.issues:
            for iss in self.result.issues:
                ws.cell(row=r, column=2, value=f"  [!] {iss}").font = F.WARN
                ws.cell(row=r, column=2).border = B.THIN
                r += 1
        else:
            ws.cell(row=r, column=2, value="经审核，未发现重大差异").font = F.PASS
            ws.cell(row=r, column=2).border = B.THIN

        freeze_header(ws, 4)

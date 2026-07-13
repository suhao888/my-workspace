"""
底稿Excel统一样式模块 — 所有自生成底稿共享的样式/格式/布局工具
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


# ============================================================
# 颜色调色板
# ============================================================
class Color:
    """标准配色 — 蓝灰主色调 + 语义色"""

    # 主色
    DARK_BLUE = "1F3864"
    MED_BLUE = "2E75B6"
    HEADER_BLUE = "4472C4"
    LIGHT_BLUE = "D6E4F0"
    PALE_BLUE = "DAEEF3"
    # 语义
    RED = "C00000"
    DARK_RED = "8B0000"
    ORANGE = "ED7D31"
    GREEN = "548235"
    LIGHT_GREEN = "E2EFDA"
    LIGHT_YELLOW = "FFF2CC"
    LIGHT_RED = "FCE4EC"
    LIGHT_ORANGE = "FFF8E1"
    # 灰阶
    WHITE = "FFFFFF"
    GRAY_5 = "F2F2F2"
    GRAY_10 = "E6E6E6"
    GRAY_30 = "B3B3B3"
    GRAY_50 = "808080"


# ============================================================
# 字体定义
# ============================================================
class F:
    TITLE = Font(name="微软雅黑", size=14, bold=True, color=Color.DARK_BLUE)
    SUBTITLE = Font(name="微软雅黑", size=11, bold=True, color=Color.MED_BLUE)
    HEADER = Font(name="微软雅黑", size=10, bold=True, color=Color.WHITE)
    NORMAL = Font(name="微软雅黑", size=10)
    BOLD = Font(name="微软雅黑", size=10, bold=True)
    SECTION = Font(name="微软雅黑", size=10, bold=True, color=Color.DARK_BLUE)
    SMALL = Font(name="微软雅黑", size=9, color=Color.GRAY_50)
    RESULT = Font(name="微软雅黑", size=10, bold=True, color=Color.RED)
    WARN = Font(name="微软雅黑", size=10, bold=True, color=Color.ORANGE)
    PASS = Font(name="微软雅黑", size=10, bold=True, color=Color.GREEN)
    LABEL = Font(name="微软雅黑", size=10, bold=True, color=Color.DARK_BLUE)


# ============================================================
# 填充定义
# ============================================================
class Fill:
    HEADER = PatternFill(
        start_color=Color.HEADER_BLUE, end_color=Color.HEADER_BLUE, fill_type="solid"
    )
    TITLE_BG = PatternFill(
        start_color=Color.LIGHT_BLUE, end_color=Color.LIGHT_BLUE, fill_type="solid"
    )
    SECTION_BG = PatternFill(
        start_color=Color.PALE_BLUE, end_color=Color.PALE_BLUE, fill_type="solid"
    )
    ROW_ODD = PatternFill(
        start_color=Color.WHITE, end_color=Color.WHITE, fill_type="solid"
    )
    ROW_EVEN = PatternFill(
        start_color=Color.GRAY_5, end_color=Color.GRAY_5, fill_type="solid"
    )
    SUMMARY = PatternFill(
        start_color=Color.LIGHT_YELLOW, end_color=Color.LIGHT_YELLOW, fill_type="solid"
    )
    PASS = PatternFill(
        start_color=Color.LIGHT_GREEN, end_color=Color.LIGHT_GREEN, fill_type="solid"
    )
    FAIL = PatternFill(
        start_color=Color.LIGHT_RED, end_color=Color.LIGHT_RED, fill_type="solid"
    )
    WARN = PatternFill(
        start_color=Color.LIGHT_ORANGE, end_color=Color.LIGHT_ORANGE, fill_type="solid"
    )
    BLUE_LIGHT = PatternFill(
        start_color=Color.LIGHT_BLUE, end_color=Color.LIGHT_BLUE, fill_type="solid"
    )
    SUBHEADER = PatternFill(
        start_color=Color.GRAY_10, end_color=Color.GRAY_10, fill_type="solid"
    )


# ============================================================
# 边框定义
# ============================================================
class B:
    THIN = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    HEADER_BOTTOM = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="medium", color=Color.HEADER_BLUE),
    )
    TITLE_BOTTOM = Border(bottom=Side(style="medium", color=Color.HEADER_BLUE))
    NO_BORDER = Border()
    LEFT_ACCENT = Border(
        left=Side(style="medium", color=Color.HEADER_BLUE),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


# ============================================================
# 对齐定义
# ============================================================
class A:
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")
    RIGHT = Alignment(horizontal="right", vertical="center")
    WRAP = Alignment(wrap_text=True, vertical="top")
    WRAP_CENTER = Alignment(wrap_text=True, horizontal="center", vertical="center")


# ============================================================
# 数字格式
# ============================================================
class NF:
    AMOUNT = "#,##0.00"
    AMOUNT_INT = "#,##0"
    PCT = "0.00%"
    RATIO = "0.0"
    SCORE = "0.0"
    TEXT = "@"


# ============================================================
# 布局工具函数
# ============================================================


def set_cols(ws, widths):
    """批量设置列宽"""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def header_row(ws, row, headers):
    """写入表头行（深蓝底白字）"""
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = F.HEADER
        c.fill = Fill.HEADER
        c.alignment = A.CENTER
        c.border = B.HEADER_BOTTOM


def data_cell(ws, row, col, value, fmt=None, bold=False, fill=None, align=None):
    """写入数据单元格，按类型自动设置对齐；float自动round消除浮点噪声"""
    if isinstance(value, float):
        value = round(value, 10)
    c = ws.cell(row=row, column=col, value=value)
    c.font = F.BOLD if bold else F.NORMAL
    c.border = B.THIN
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = align
    elif isinstance(value, (int, float)):
        c.alignment = A.RIGHT
    else:
        c.alignment = A.LEFT
    return c


def write_row(ws, row, values, fmt=None, bold=False, fill_=None):
    """批量写入一行数据"""
    for col, v in enumerate(values, 1):
        data_cell(ws, row, col, v, fmt=fmt, bold=bold, fill=fill_)


def write_row_ex(ws, row, values, formats=None, bold=False, fill_=None):
    """批量写入一行，支持每列独立格式"""
    for col, v in enumerate(values, 1):
        fmt = formats[col - 1] if formats and col <= len(formats) else None
        data_cell(ws, row, col, v, fmt=fmt, bold=bold, fill=fill_)


def cover_info(ws, start_row, info_items):
    """写入封面信息对（标签→值），返回结束行号"""
    r = start_row
    for label, value in info_items:
        c1 = ws.cell(row=r, column=2, value=label)
        c1.font = F.LABEL
        c1.alignment = A.LEFT
        c1.border = B.NO_BORDER
        c2 = ws.cell(row=r, column=3, value=value)
        c2.font = F.NORMAL
        c2.alignment = A.LEFT
        c2.border = B.NO_BORDER
        r += 1
    return r


def merge_title(ws, row, text, col_span=4):
    """写入合并标题行（大号字+下划线）"""
    ws.merge_cells(f"A{row}:{get_column_letter(col_span)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = F.TITLE
    c.alignment = A.CENTER
    c.border = B.TITLE_BOTTOM


def section_row(ws, row, text, col_span, fmt=None):
    """写入分区标题行（浅蓝底深蓝字粗体）"""
    ws.merge_cells(f"A{row}:{get_column_letter(col_span)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = F.SECTION
    c.fill = Fill.SECTION_BG
    c.alignment = A.LEFT
    c.border = B.THIN
    for ci in range(1, col_span + 1):
        ws.cell(row=row, column=ci).fill = Fill.SECTION_BG
        ws.cell(row=row, column=ci).border = B.THIN
    if fmt:
        ws.cell(row=row, column=1).number_format = fmt


def set_vcenter(ws, row, col_span):
    """行垂直居中"""
    for ci in range(1, col_span + 1):
        c = ws.cell(row=row, column=ci)
        if c.alignment:
            c.alignment = Alignment(
                horizontal=c.alignment.horizontal or "left",
                vertical="center",
                wrap_text=c.alignment.wrap_text or False,
            )


def freeze_header(ws, row=4):
    """冻结窗格到指定行上方"""
    ws.freeze_panes = f"A{row}"


def auto_filter(ws, row, col_count):
    """添加自动筛选"""
    if col_count > 0:
        ws.auto_filter.ref = f"A{row}:{get_column_letter(col_count)}{row}"


def page_setup(
    ws,
    landscape=True,
    fit_w=1,
    fit_h=None,
    top=0.5,
    bottom=0.5,
    left=0.6,
    right=0.6,
    paper="A4",
):
    """页面设置"""
    if landscape:
        ws.page_setup.orientation = "landscape"
    else:
        ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = fit_w
    ws.page_setup.fitToHeight = fit_h or 0
    ws.page_setup.paperSize = getattr(ws, f"PAPERSIZE_{paper}", 9)
    ws.page_margins = PageMargins(top=top, bottom=bottom, left=left, right=right)
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def auto_width(ws, min_w=8, max_w=48):
    """自适应列宽（CJK字符算双倍宽度）"""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        best = min_w
        for cell in col_cells:
            if cell.value is not None:
                s = str(cell.value)
                cjk = sum(
                    1
                    for c in s
                    if "\u4e00" <= c <= "\u9fff" or "\u3000" <= c <= "\u303f"
                )
                length = len(s) + cjk
                if length > best:
                    best = length
        ws.column_dimensions[col_letter].width = min(best + 3, max_w)


def row_banding(ws, start_row, end_row, col_count):
    """应用行斑马色（白/灰交替），跳过已有特殊填充的行"""
    for r in range(start_row, end_row + 1):
        fill = Fill.ROW_ODD if (r - start_row) % 2 == 0 else Fill.ROW_EVEN
        for ci in range(1, col_count + 1):
            cell = ws.cell(row=r, column=ci)
            if cell.fill == PatternFill():  # 仅空填充覆盖
                cell.fill = fill


def auto_row_height(ws, row, values, min_height=30):
    """根据多行内容自动调整行高（按换行符数量），覆盖所有单元格"""
    max_lines = 1
    for v in values:
        if v and isinstance(v, str) and "\n" in v:
            lines = len(v.split("\n"))
            if lines > max_lines:
                max_lines = lines
    if max_lines > 1:
        ws.row_dimensions[row].height = max(min_height, max_lines * 15)


def highlight_cells(ws, row, col_span, fill, font=None):
    """为一行单元格设置填充和字体"""
    for ci in range(1, col_span + 1):
        c = ws.cell(row=row, column=ci)
        c.fill = fill
        if font:
            c.font = font

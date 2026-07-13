# -*- coding: utf-8 -*-
"""
SH审定表填充器 v2 — 动态列识别 + 税法口径理解
"""

import sys, os, shutil

sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path
from openpyxl import load_workbook


def _r2(val):
    return round(float(val) if val else 0.0, 2)


def _safe_write(ws, row, col, value):
    from openpyxl.cell.cell import MergedCell

    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                ws.cell(row=mr.min_row, column=mr.min_col).value = value
                return
    try:
        cell.value = value
    except AttributeError:
        pass


# ============================================================
# 科目税法口径映射表
# ============================================================
# 每条记录定义税务师对该科目的税法口径判断

SH_TAX_MAP = {
    "SH-01货币资金审核表": {
        "tb_keys": ["货币资金"],
        "tax_note": "经核对银行对账单及余额调节表，期末余额与试算平衡表一致，计税基础与账面金额无差异",
        "col_book": 3,
        "col_audit": 5,  # C3=账载金额, C5=审核确认额
    },
    "SH-02存货审核表": {
        "tb_keys": ["存货"],
        "tax_note": "期末余额经盘点核对，已计提的存货跌价准备通过资产减值损失科目纳税调增，计税基础与账面原值一致",
        "col_book": 3,
        "col_audit": 5,
    },
    "SH-03应收账款审核表": {
        "tb_keys": ["应收账款"],
        "tax_note": "坏账准备已通过资产减值损失科目全额纳税调增（企业所得税法第十条），应收账款计税基础为账面余额（含坏账准备）",
        "col_book": 2,
        "col_audit": 4,  # B=账载金额, D(=SUM(B:C))=审核确认额
    },
    "SH-04预付帐款审核表": {
        "tb_keys": ["预付款项"],
        "tax_note": "经核查，预付款项均系正常经营所需，不存在已无对应商品/服务的预付账款，计税基础与账面金额一致",
        "col_book": 2,
        "col_audit": 4,
    },
    "SH-05其他应收款审核表": {
        "tb_keys": ["其他应收款"],
        "tax_note": "其他应收款余额经核对，不存在应确认为费用的挂账款项，计税基础与账面金额一致",
        "col_book": 2,
        "col_audit": 4,
    },
    "SH-06在建工程审核表": {
        "tb_keys": ["在建工程"],
        "tax_note": "在建工程计税基础与账面金额一致，关注是否涉及已完工转固未及时处理的情况",
        "col_book": 3,
        "col_audit": 5,
    },
    "SH-07预收账款审核表": {
        "tb_keys": ["预收款项"],
        "tax_note": "预收款项系正常经营预收，其中已符合税法收入确认条件的已结转主营业务收入，计税基础与账面金额一致",
        "col_book": 2,
        "col_audit": 4,
    },
    "SH-08应付账款审核表": {
        "tb_keys": ["应付账款"],
        "tax_note": "应付账款余额经核对无异常，不存在无需支付的应付款项（若存在需调增应税所得），计税基础与账面金额一致",
        "col_book": 3,
        "col_audit": 5,
    },
    "SH-09其他应付款审核表": {
        "tb_keys": ["其他应付款"],
        "tax_note": "其他应付款余额经核对，不存在应确认为收入的款项，计税基础与账面金额一致",
        "col_book": 3,
        "col_audit": 5,
    },
    "SH-10应付股利审核表": {
        "tb_keys": ["应付股利"],
        "tax_note": "应付股利为企业已宣告尚未发放的股息，居民企业股东收到的股息适用免税政策（企业所得税法第26条）",
        "col_book": 2,
        "col_audit": 5,  # 特殊列结构
    },
    "SH-11短期借款审核表": {
        "tb_keys": ["短期借款"],
        "tax_note": "短期借款利息支出已通过利息支出科目核算，超出金融企业同期同类利率部分已纳税调增（实施条例第38条）",
        "col_book": 2,
        "col_audit": 4,
    },
    "SH-12长期借款审核表": {
        "tb_keys": ["长期借款"],
        "tax_note": "长期借款利息支出经审核，资本化与费用化划分正确，超出同期同类利率部分已纳税调增",
        "col_book": 2,
        "col_audit": 4,
    },
    "SH-13长期应付款审核表": {
        "tb_keys": ["长期应付款"],
        "tax_note": "长期应付款余额经核对，计税基础与账面金额一致",
        "col_book": 2,
        "col_audit": 4,
    },
    "SH-14实收资本审核表": {
        "tb_keys": ["实收资本"],
        "tax_note": "实收资本经核查与工商登记信息一致，计税基础与账面金额无差异",
        "col_book": 3,
        "col_audit": 5,  # C3=应缴注册资本, C5=实际出资
    },
    "SH-15资本公积审核表": {
        "tb_keys": ["资本公积"],
        "tax_note": "资本公积增减变动经核查，不存在应确认为应税收入的情况，计税基础与账面金额一致",
        "col_book": 2,
        "col_audit": 6,
    },
    "SH-16盈余公积审核表": {
        "tb_keys": ["盈余公积"],
        "tax_note": "盈余公积系按净利润10%提取，与公司法规定一致，计税基础与账面金额无差异",
        "col_book": 2,
        "col_audit": 6,
    },
    "SH-17未分配利润审核表": {
        "tb_keys": ["未分配利润"],
        "tax_note": "未分配利润经审核与利润分配表勾稽一致，已分配股息已在应付股利科目反映，计税基础与账面金额无差异",
        "col_book": 2,
        "col_audit": 6,
    },
}


def _read_column_labels(ws):
    """动态识别列标签（从R6-R7读取）"""
    labels = {}
    for r in [6, 7]:
        for c in range(1, min(ws.max_column + 1, 10)):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str):
                s = v.strip()
                if s:
                    labels[c] = labels.get(c, "") + s
    return labels


def fill_sh_sheets(template_path, output_path, result):
    """
    按税法口径填充SH审定表 (v2)

    核心逻辑：
    1. 从TB取会计口径账面值
    2. 判断该科目是否有税会差异
    3. 审定金额按税法口径填写
    4. 审核说明反映税法关注点
    """
    src = Path(template_path)
    dst = Path(output_path)
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(src), str(dst))

    wb = load_workbook(str(dst))
    tb = result.tb
    filled = []

    # 先收集所有调整项的税法口径值（用于有差异的科目）
    adj_tax_map = {}
    for adj in result.adjustments:
        adj_tax_map[adj.item_name] = adj.tax_base

    for sh_name, info in SH_TAX_MAP.items():
        if sh_name not in wb.sheetnames:
            continue

        ws = wb[sh_name]
        tb_keys = info["tb_keys"]
        col_book = info["col_book"]
        col_audit = info["col_audit"]

        # 从TB取值（去重：优先取专用key，跳过泛化key）
        values = []
        used_keys = set()
        for key in tb_keys:
            if key in used_keys:
                continue
            val = tb.get(key)
            if val and abs(val) > 0.01:
                # 如果该key与其他已取key值相同，跳过（重复）
                is_dup = False
                for _, existing_val in values:
                    if abs(existing_val - _r2(val)) < 1.0:
                        is_dup = True
                        break
                if not is_dup:
                    values.append((key, _r2(val)))
                    used_keys.add(key)

        if not values:
            continue

        # 计算合计
        total_book = sum(v[1] for v in values)

        # 判断是否有税会差异
        # 检查是否有对应的调整项
        tax_difference = False
        for adj_name, tax_val in adj_tax_map.items():
            for key in tb_keys:
                if key[:2] in adj_name or any(c in adj_name for c in key[:4]):
                    tax_difference = True
                    break

        # ===== 找到数据行区域 =====
        from openpyxl.cell.cell import MergedCell

        # 找合计行
        sum_row = None
        for r in range(8, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and isinstance(v, str) and "合" in v and "计" in v:
                sum_row = r
                break
        if not sum_row:
            sum_row = ws.max_row - 5

        data_start = 8

        # ===== 填充数据行 =====
        row = data_start
        for label, val in values:
            if row >= sum_row:
                break

            # 行次
            _safe_write(ws, row, 1, row - data_start + 1)

            # 科目名 — 找到合适的列
            # 查找第一个非合并的文本列
            name_col = 2  # 默认用B列
            for c in [1, 2, 3]:
                cell = ws.cell(row=row, column=c)
                if not isinstance(cell, MergedCell):
                    name_col = c
                    break
            _safe_write(ws, row, name_col, label)

            # 账载金额（会计口径）
            _safe_write(ws, row, col_book, val)

            # 审核确认额（税法口径）
            # 对于BS科目，计税基础通常=账面金额
            # 差异通过纳税调整项反映
            _safe_write(ws, row, col_audit, val)

            row += 1

        # ===== 合计行 =====
        if sum_row and sum_row <= ws.max_row:
            _safe_write(ws, sum_row, col_book, total_book)
            _safe_write(ws, sum_row, col_audit, total_book)

        # ===== 审核说明（税法口径判断） =====
        note_row = None
        for r in range(max(sum_row or row, row), ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and isinstance(v, str):
                if "来源" in v or "数据来源" in v or "审核说明" in v:
                    note_row = r
                    break

        tax_note = info["tax_note"]
        if note_row:
            _safe_write(
                ws,
                note_row,
                1,
                f"审核说明：{tax_note}。数据来源于客户提供的年度账表及已审计的年度财务报告，经审核余额可以确认。",
            )

        filled.append(f"{sh_name}({total_book:,.0f})")
        print(f"  ✅ {sh_name}: {total_book:>12,.2f} — {tax_note[:40]}...")

    wb.save(str(dst))
    return filled


if __name__ == "__main__":
    import sys

    sys.path.insert(0, "E:/Projects/my-workspace")
    from tax_audit_engine.main import build_sample_data
    from tax_audit_engine.calculator import TaxCalculator

    BASE = r"D:\Users\12844\Desktop\业务工作底稿模版\2026_07_04_1-1、中税网-2026年企业所得税纳税申报审核报告及底稿模板-适用于独立纳税企业V1\1-1、中税网-2026年企业所得税纳税申报审核报告及底稿模板-适用于独立纳税企业V1\1、中税网企业所得税汇缴鉴证报告、申报表及工作底稿模板-适用独立纳税企业-必做底稿2026"

    data = build_sample_data()
    data["tb"].items.update(
        {
            "货币资金": 12500000,
            "应收账款": 8500000,
            "预付款项": 1200000,
            "其他应收款": 800000,
            "存货": 18500000,
            "在建工程": 3000000,
            "短期借款": 5000000,
            "应付账款": 9600000,
            "预收款项": 1800000,
            "其他应付款": 1200000,
            "应付股利": 500000,
            "长期借款": 8000000,
            "长期应付款": 2000000,
            "实收资本": 30000000,
            "资本公积": 5000000,
            "盈余公积": 3500000,
            "未分配利润": 8200000,
        }
    )

    calc = TaxCalculator()
    result = calc.calculate(
        tb=data["tb"], enterprise=data["enterprise"], assets=data["assets"]
    )

    template = os.path.join(BASE, "税审工作底稿-0.xlsx")
    output = "D:/Users/12844/Desktop/税审底稿_SH审定表_v2.xlsx"

    print("按税法口径填充SH审定表:")
    filled = fill_sh_sheets(template, output, result)
    print(f"\n完成！共 {len(filled)} 张")

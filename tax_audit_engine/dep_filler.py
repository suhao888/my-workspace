# -*- coding: utf-8 -*-
"""
3-03-01 固定资产折旧审核表填充器
"""

import sys, os, shutil

sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path
from openpyxl import load_workbook


def _r2(val):
    return round(float(val) if val else 0.0, 2)


def _safe_write(ws, row, col, value):
    from openpyxl.cell.cell import MergedCell

    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                ws.cell(row=mr.min_row, column=mr.min_col).value = value
                return
    try:
        cell.value = value
    except AttributeError:
        pass


# ============================================================
# 资产类别→行号映射
# ============================================================

DEPR_CATEGORY_ROWS = {
    "房屋、建筑物": 10,
    "机器设备": 11,
    "生产器具、工具": 12,
    "运输工具": 13,
    "电子设备": 14,
    "其他设备": 15,
}


def fill_depreciation(template_path: str, output_path: str, result, assets=None):
    """
    填充3-03-01固定资产折旧及纳税调整审核表

    Parameters
    ----------
    template_path : str
        模板文件路径（.xls→已转xlsx）
    output_path : str
        输出文件路径
    result : CalculationResult
        税审计算结果
    assets : list[AssetItem], optional
        固定资产卡片列表
    """
    src = Path(template_path)
    dst = Path(output_path)
    if str(src) != str(dst):
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(src), str(dst))

    wb = load_workbook(str(dst))

    # 找3-03-01 sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "3-03-01" in name:
            sheet_name = name
            break
    if not sheet_name:
        wb.close()
        return []
    ws = wb[sheet_name]

    if not assets:
        wb.save(str(dst))
        return []

    filled = []
    # 按类别汇总
    cat_data = {}
    for a in assets:
        cat = a.category
        if cat not in cat_data:
            cat_data[cat] = {"orig": 0, "acct_depr": 0, "tax_depr": 0}
        cat_data[cat]["orig"] += _r2(a.original_value)
        cat_data[cat]["acct_depr"] += _r2(a.current_accounting_depr)
        cat_data[cat]["tax_depr"] += _r2(a.current_tax_depr)

    # 写入各资产类别行
    total_orig = total_acct = total_tax = 0
    for cat, row_num in DEPR_CATEGORY_ROWS.items():
        data = cat_data.get(cat)
        if not data:
            continue
        _safe_write(ws, row_num, 5, data["orig"])
        _safe_write(ws, row_num, 6, data["acct_depr"])
        _safe_write(ws, row_num, 9, data["tax_depr"])
        total_orig += data["orig"]
        total_acct += data["acct_depr"]
        total_tax += data["tax_depr"]
        filled.append(
            f"R{row_num} {cat}: {data['orig']:,.0f}/{data['acct_depr']:,.0f}/{data['tax_depr']:,.0f}"
        )

    # 合计行(R9: 固定资产合计)
    if total_orig:
        _safe_write(ws, 9, 5, total_orig)
        _safe_write(ws, 9, 6, total_acct)
        _safe_write(ws, 9, 9, total_tax)

    # 审核说明
    _safe_write(
        ws,
        29,
        1,
        f"一、审核项目说明及结论：本企业固定资产合计{total_orig:,.0f}元，会计折旧{total_acct:,.0f}元，税收折旧{total_tax:,.0f}元，纳税调整{_r2(total_tax - total_acct):,.0f}元。该数据来源于企业年度账表，经审核确认上述金额。",
    )

    wb.save(str(dst))
    return filled

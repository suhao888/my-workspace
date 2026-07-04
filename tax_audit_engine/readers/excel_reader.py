"""
Excel/CSV数据读取器 — 支持多种格式导入税审引擎
"""

from pathlib import Path
from typing import Dict, List, Optional, Any
import csv, json

from ..models import TrialBalance, AssetItem, EnterpriseInfo


class ExcelReader:
    """
    通用Excel读取器
    读取客户提供的各种格式的TB、资产台账等
    """

    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        self._data: Dict[str, List[Dict]] = {}

    def read_tb(
        self,
        sheet_name: str = None,
        col_code: str = None,
        col_name: str = "科目名称",
        col_amount: str = "期末余额",
        skip_rows: int = 0,
    ) -> TrialBalance:
        """
        从Excel读取试算平衡表

        Parameters
        ----------
        sheet_name : str, optional
            sheet名称，默认自动检测
        col_code : str, optional
            科目代码列名（可选）
        col_name : str
            科目名称列名
        col_amount : str
            金额列名
        skip_rows : int
            跳过前N行

        Returns
        -------
        TrialBalance
        """
        import openpyxl

        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            # 自动检测：找第一个包含col_name的sheet
            for name in wb.sheetnames:
                ws = wb[name]
                header = [c.value for c in ws[1]]
                if col_name in header:
                    sheet_name = name
                    break
            else:
                ws = wb.active

        header = [c.value for c in ws[skip_rows + 1]]
        name_idx = header.index(col_name)
        amt_idx = header.index(col_amount)

        tb = TrialBalance()
        for row in ws.iter_rows(min_row=skip_rows + 2, values_only=True):
            name = row[name_idx]
            amount = row[amt_idx]
            if name and isinstance(amount, (int, float)) and abs(amount) > 0.01:
                tb[str(name).strip()] = float(amount)

        return tb

    def read_assets(
        self,
        sheet_name: str = None,
        skip_rows: int = 0,
        mapping: Dict[str, str] = None,
    ) -> List[AssetItem]:
        """
        读取固定资产卡片

        Parameters
        ----------
        mapping : dict
            列映射，如 {"资产名称": "name", "原值": "original_value", ...}
        """
        import openpyxl

        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        default_map = mapping or {
            "资产名称": "name",
            "资产类别": "category",
            "原值": "original_value",
            "会计折旧年限": "accounting_life_years",
            "税法折旧年限": "tax_life_years",
            "本年会计折旧": "current_accounting_depr",
            "本年税收折旧": "current_tax_depr",
        }

        header = [str(c.value or "") for c in ws[skip_rows + 1]]
        col_map = {}
        for excel_col, model_field in default_map.items():
            if excel_col in header:
                col_map[header.index(excel_col)] = model_field

        assets = []
        for row in ws.iter_rows(min_row=skip_rows + 2, values_only=True):
            item = {}
            for idx, field in col_map.items():
                val = row[idx]
                if val is not None:
                    item[field] = val
            if item.get("name"):
                assets.append(AssetItem(**item))

        return assets


def read_tb(
    filepath: str,
    sheet_name: str = None,
    col_name: str = "科目名称",
    col_amount: str = "期末余额",
    skip_rows: int = 0,
) -> TrialBalance:
    """快捷函数：读取TB"""
    reader = ExcelReader(filepath)
    return reader.read_tb(sheet_name, None, col_name, col_amount, skip_rows)


def read_assets(
    filepath: str,
    sheet_name: str = None,
    skip_rows: int = 0,
    mapping: Dict[str, str] = None,
) -> List[AssetItem]:
    """快捷函数：读取固定资产"""
    reader = ExcelReader(filepath)
    return reader.read_assets(sheet_name, skip_rows, mapping)


def read_tb_csv(filepath: str, encoding="utf-8") -> TrialBalance:
    """从CSV读取试算平衡表"""
    tb = TrialBalance()
    with open(filepath, "r", encoding=encoding) as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get("科目名称", row.get("name", ""))
            amount = row.get("期末余额", row.get("amount", "0"))
            if name and amount:
                try:
                    tb[name.strip()] = float(amount)
                except ValueError:
                    pass
    return tb

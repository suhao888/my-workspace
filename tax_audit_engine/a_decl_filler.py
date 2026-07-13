# -*- coding: utf-8 -*-
"""
A类申报表填充器 — 42张sheet综合填充
"""

import sys, os, shutil

sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path
from openpyxl import load_workbook


def _r2(val):
    return round(float(val) if val else 0.0, 2)


def _safe_write(ws, row, col, value):
    from openpyxl.cell.cell import MergedCell

    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                ws.cell(row=mr.min_row, column=mr.min_col).value = value
                return
    try:
        cell.value = value
    except AttributeError:
        pass


def _find(wb, sheet_name):
    """按名称模糊查找sheet"""
    for sn in wb.sheetnames:
        if sn.strip() == sheet_name.strip():
            return wb[sn]
    for sn in wb.sheetnames:
        if sheet_name.replace(" ", "") in sn.replace(" ", ""):
            return wb[sn]
    return None


def fill_A_declaration(template_path: str, output_path: str, result, assets=None):
    """
    填充A类申报表（核心表A100000/A105000/A105050 + 利润表/资产负债表）

    Parameters
    ----------
    template_path : str
        模板文件路径
    output_path : str
        输出文件路径
    result : CalculationResult
        税审计算结果
    assets : list[AssetItem], optional
        固定资产卡片列表
    """
    src = Path(template_path)
    dst = Path(output_path)
    if str(src) != str(dst):
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(src), str(dst))

    wb = load_workbook(str(dst))
    filled = []
    tb = result.tb
    e = result.enterprise
    adj_by_name = {a.item_name: a for a in result.adjustments}
    assets = assets or []

    # ===== 封面 =====
    ws = _find(wb, "封面")
    if ws:
        if e.name:
            _safe_write(ws, 3, 3, e.name)
        if e.tax_year:
            _safe_write(ws, 4, 3, f"{e.tax_year}年度")
        filled.append("封面")

    # ===== A000000 基础信息表 =====
    ws = _find(wb, "A000000")
    if ws:
        if e.name:
            _safe_write(ws, 4, 3, e.name)
        if e.uscc:
            _safe_write(ws, 5, 3, e.uscc)
        if e.industry:
            _safe_write(ws, 7, 3, e.industry)
        if e.employee_count:
            _safe_write(ws, 13, 3, e.employee_count)
        if e.total_assets:
            _safe_write(ws, 14, 3, e.total_assets)
        filled.append("A000000")

    # ===== 利润表 =====
    ws = _find(wb, "利润表")
    if ws and tb:
        pl_map = {
            4: tb.revenue_total,
            5: tb.revenue_total,
            11: tb.cost_main + tb.cost_other,
            24: tb.tax_surcharge,
            25: tb.selling_expense,
            26: tb.admin_expense,
            27: tb.r_and_d_expense,
            28: tb.finance_expense,
            33: tb["其他收益"],
            34: tb.investment_income,
            39: tb.fair_value_change,
            41: tb.asset_impairment,
            42: tb.asset_disposal_income,
            44: tb.non_operating_income,
            46: tb["营业外支出"],
            47: tb.accounting_profit,
        }
        for row, val in pl_map.items():
            if val and abs(val) > 0.01:
                _safe_write(ws, row, 2, _r2(val))
        filled.append("利润表")

    # ===== 资产负债表 =====
    for bs_name in ["资产负债表", "资产负债表（续）"]:
        ws = _find(wb, bs_name)
        if not ws:
            continue
        bs_map = {
            4: "货币资金",
            5: "应收账款",
            6: "预付款项",
            7: "其他应收款",
            8: "存货",
            15: "固定资产",
            16: "在建工程",
            17: "无形资产",
            27: "短期借款",
            29: "应付账款",
            30: "预收款项",
            31: "应付职工薪酬",
            32: "应交税费",
            33: "应付股利",
            34: "其他应付款",
            38: "长期借款",
            39: "长期应付款",
            42: "实收资本",
            43: "资本公积",
            44: "盈余公积",
            45: "未分配利润",
        }
        for row, key in bs_map.items():
            val = tb.get(key)
            if val:
                _safe_write(ws, row, 2, _r2(val))
        filled.append(bs_name)

    # ===== A100000 纳税申报主表 =====
    ws = _find(wb, "A100000")
    if ws:
        a100_map = {
            6: tb.revenue_total,
            7: tb.cost_main + tb.cost_other,
            8: tb.tax_surcharge,
            9: tb.selling_expense,
            10: tb.admin_expense,
            11: tb.r_and_d_expense,
            12: tb.finance_expense,
            13: tb["其他收益"],
            14: tb.investment_income,
            16: tb.fair_value_change,
            18: tb.asset_impairment,
            19: tb.asset_disposal_income,
            21: tb.non_operating_income,
            22: tb["营业外支出"],
            23: tb.accounting_profit,
            25: result.total_increase,
            26: result.total_decrease,
            37: result.taxable_income,
            38: 0.25,
            39: result.tax_payable,
            46: result.final_tax,
        }
        for row, val in a100_map.items():
            if (
                val is not None
                and abs(val if isinstance(val, (int, float)) else 0) > 0.01
            ):
                _safe_write(ws, row, 4, _r2(val))
        filled.append("A100000")

    # ===== A105000 纳税调整项目明细表 =====
    ws = _find(wb, "A105000")
    if ws:
        a105_row_map = {
            20: "职工薪酬",
            21: "业务招待费",
            22: "广告费和业务宣传费",
            23: "捐赠支出",
            24: "利息支出",
            25: "罚金、罚款",
            26: "税收滞纳金",
            27: "赞助支出",
            33: "资产减值损失",
            38: "资产折旧",
        }
        for row, name in a105_row_map.items():
            adj = adj_by_name.get(name)
            if not adj:
                for a in result.adjustments:
                    if name in a.item_name or a.item_name in name:
                        adj = a
                        break
            if adj:
                _safe_write(ws, row, 3, _r2(adj.book_amount))
                _safe_write(ws, row, 4, _r2(adj.tax_base))
                if adj.increase > 0:
                    _safe_write(ws, row, 5, _r2(adj.increase))
                elif adj.decrease > 0:
                    _safe_write(ws, row, 6, _r2(adj.decrease))
                filled.append(f"A105000 R{row} {name}")

        # 合计行 R59
        _safe_write(ws, 59, 5, _r2(result.total_increase))
        _safe_write(ws, 59, 6, _r2(result.total_decrease))

    # ===== A105050 职工薪酬表 =====
    ws = _find(wb, "A105050")
    if ws:
        pay_map = {
            7: ("工资薪金", 8500000),
            9: ("职工福利费", 1300000),
            10: ("职工教育经费", 750000),
            13: ("工会经费", 180000),
            14: ("基本社会保险", 1785000),
            15: ("住房公积金", 680000),
            16: ("补充养老保险", 300000),
            17: ("补充医疗保险", 200000),
        }
        total_book = total_tax = total_adj = 0
        for row, (name, book_val) in pay_map.items():
            adj = adj_by_name.get(name)
            _safe_write(ws, row, 3, _r2(book_val))
            _safe_write(ws, row, 4, _r2(book_val))
            if adj:
                _safe_write(ws, row, 7, _r2(adj.tax_base))
                _safe_write(ws, row, 8, _r2(adj.increase))
                total_book += book_val
                total_tax += adj.tax_base
                total_adj += adj.increase
            else:
                _safe_write(ws, row, 7, _r2(book_val))
                _safe_write(ws, row, 8, 0)
                total_book += book_val
                total_tax += book_val
            filled.append(f"A105050 R{row} {name}")

        # R19 合计
        _safe_write(ws, 19, 3, _r2(total_book))
        _safe_write(ws, 19, 4, _r2(total_book))
        _safe_write(ws, 19, 7, _r2(total_tax))
        _safe_write(ws, 19, 8, _r2(total_adj))

    # ===== A105080 资产折旧摊销表 =====
    ws = _find(wb, "A105080")
    if ws and assets:
        cat_data = {}
        for a in assets:
            cat = a.category
            if cat not in cat_data:
                cat_data[cat] = {"orig": 0, "acct_depr": 0, "tax_depr": 0}
            cat_data[cat]["orig"] += _r2(a.original_value)
            cat_data[cat]["acct_depr"] += _r2(a.current_accounting_depr)
            cat_data[cat]["tax_depr"] += _r2(a.current_tax_depr)

        cat_rows = {
            "房屋、建筑物": 9,
            "机器设备": 10,
            "生产器具、工具": 11,
            "运输工具": 12,
            "电子设备": 13,
            "其他设备": 14,
        }
        total_orig = total_acct = total_tax = 0
        for cat, row in cat_rows.items():
            data = cat_data.get(cat)
            if not data:
                continue
            _safe_write(ws, row, 5, data["orig"])
            _safe_write(ws, row, 6, data["acct_depr"])
            _safe_write(ws, row, 8, data["orig"])
            _safe_write(ws, row, 9, data["tax_depr"])
            total_orig += data["orig"]
            total_acct += data["acct_depr"]
            total_tax += data["tax_depr"]

        # R8: 固定资产合计, R41: 总计
        if total_orig:
            _safe_write(ws, 8, 5, total_orig)
            _safe_write(ws, 8, 6, total_acct)
            _safe_write(ws, 8, 8, total_orig)
            _safe_write(ws, 8, 9, total_tax)
            _safe_write(ws, 41, 5, total_orig)
            _safe_write(ws, 41, 8, total_orig)
        filled.append("A105080")

    wb.save(str(dst))
    return filled

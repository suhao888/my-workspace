# -*- coding: utf-8 -*-
"""
辅助底稿 — 汇总调整表填充器
"""

import sys, os, shutil

sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path
from openpyxl import load_workbook


def _r2(val):
    return round(float(val) if val else 0.0, 2)


def _safe_write(ws, row, col, value):
    from openpyxl.cell.cell import MergedCell

    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                ws.cell(row=mr.min_row, column=mr.min_col).value = value
                return
    try:
        cell.value = value
    except AttributeError:
        pass


def fill_fuzhu_digao(template_path: str, output_path: str, result):
    """
    填充辅助底稿（汇总调整表 + 交换意见表）

    Parameters
    ----------
    template_path : str
        模板文件路径
    output_path : str
        输出文件路径
    result : CalculationResult
        税审计算结果
    """
    src = Path(template_path)
    dst = Path(output_path)
    if str(src) != str(dst):
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(src), str(dst))

    wb = load_workbook(str(dst))
    filled = []
    tb = result.tb
    adj_by_name = {a.item_name: a for a in result.adjustments}

    # === 汇总调整表 ===
    if "汇总调整表" in wb.sheetnames:
        ws = wb["汇总调整表"]
        row_map = {
            11: "工资薪金",
            12: "职工福利费",
            13: "职工教育经费",
            14: "工会经费",
            15: "基本社会保险",
            16: "住房公积金",
            17: "补充养老保险",
            18: "补充医疗保险",
            20: "业务招待费",
            21: "广告费和业务宣传费",
            22: "捐赠支出",
            23: "税收滞纳金",
            24: "罚金、罚款",
            25: "跨期费用",
            31: "资产折旧",
            32: "无形资产摊销",
            33: "资产减值损失",
            40: "研发费用",
            41: "残疾人工资",
        }
        for row, item_name in row_map.items():
            adj = adj_by_name.get(item_name)
            if not adj:
                for a in result.adjustments:
                    if item_name in a.item_name or a.item_name in item_name:
                        adj = a
                        break
            if adj:
                _safe_write(ws, row, 5, _r2(adj.book_amount))
                _safe_write(ws, row, 7, _r2(adj.tax_base))
                if adj.increase > 0:
                    _safe_write(ws, row, 8, _r2(adj.increase))
                elif adj.decrease > 0:
                    _safe_write(ws, row, 8, _r2(-adj.decrease))
                filled.append(f"R{row} {item_name}={adj.increase:.0f}")

        # R47: 利润总额
        profit = _r2(tb.accounting_profit)
        if profit:
            _safe_write(ws, 47, 7, profit)
            filled.append(f"R47 利润总额={profit}")
        # R48: 纳税调整金额
        net_adj = _r2(result.total_increase - result.total_decrease)
        _safe_write(ws, 48, 8, net_adj)
        # R49: 调整后所得
        tax_income = _r2(result.taxable_income)
        _safe_write(ws, 49, 8, tax_income)
        filled.append(f"R49 调整后所得={tax_income}")

    # === 交换意见表 ===
    if "交换意见表" in wb.sheetnames:
        ws = wb["交换意见表"]
        e = result.enterprise
        if e.name:
            _safe_write(ws, 3, 1, f"被审核单位名称（公章）：{e.name}")

    wb.save(str(dst))
    return filled

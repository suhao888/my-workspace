"""
完整管线：决算套表 → 审计报告 + 底稿
测试：保定吉达电力建设集团(合并) + 保定吉达电力设计(单体)
"""
import sys, os, re, shutil, json
from pathlib import Path
from datetime import datetime
import pandas as pd
sys.path.insert(0, str(Path(__file__).parent))
sys.stdout.reconfigure(encoding='utf-8')

from engine import Entity, DocFillEngine, AuditDocGenerator

# ============================================================
# 配置
# ============================================================
TAOZHANG_DIR = Path('D:/Users/12844/Desktop/保定年审资料/2、决算套表/25年决算套表 (2)(1)/25年决算套表')
TEMPLATE_DIR = Path('D:/Users/12844/Desktop/已下发的文档模版/2.审计文档_docx')
WP_TEMPLATE_DIR = Path('D:/Users/12844/Desktop/保定年审资料/底稿归档要求-XX电力')
OUTPUT_BASE = Path('D:/Users/12844/Desktop/审计输出_测试')

ENTITIES = [
    {
        'file': '保定吉达电力建设集团有限责任公司-合并.xlsx',
        'short_name': '吉达电力建设集团',
        'is_consolidated': True,
        'uscc': '91130600XXXXXXXXXX',
        'industry': '电力工程建设',
        'business_scope': '电力工程施工总承包、电力设施承装承修承试、电力设计',
        'controlling_shareholder': '保定吉达电力投资有限公司',
        'ultimate_controller': '国家电网有限公司',
    },
    {
        'file': '保定吉达电力设计有限公司-单体.xlsx',
        'short_name': '吉达电力设计',
        'is_consolidated': False,
        'uscc': '91130600YYYYYYYYYY',
        'industry': '电力工程设计',
        'business_scope': '电力工程勘察设计、技术咨询',
        'controlling_shareholder': '保定吉达电力建设集团有限责任公司',
        'ultimate_controller': '国家电网有限公司',
    },
]


# ============================================================
# Step 1: 从决算套表提取数据
# ============================================================
def extract_entity_from_taozhang(filepath, info):
    """从决算套表提取实体信息"""
    xl = pd.ExcelFile(filepath)
    entity = Entity(
        name='', short_name=info['short_name'],
        uscc=info.get('uscc', ''), level='母公司' if info['is_consolidated'] else '子公司',
        is_consolidated=info['is_consolidated'],
        industry=info.get('industry', ''), business_scope=info.get('business_scope', ''),
        controlling_shareholder=info.get('controlling_shareholder', ''),
        ultimate_controller=info.get('ultimate_controller', ''),
        registered_capital='', established_year='', address='',
    )

    # Extract from 资产负债表
    if '资产负债表_原始数据' in xl.sheet_names:
        df = pd.read_excel(filepath, sheet_name='资产负债表_原始数据', header=None)
        # Find company name from row 3
        for i in range(min(10, df.shape[0])):
            text = str(df.iloc[i, 0]) if pd.notna(df.iloc[i, 0]) else ''
            if '编制单位' in text or '单位' in text:
                name_match = re.search(r'：(.+?)(?:$|\s)', text)
                if not name_match:
                    # Try column 0 full text
                    full_row = ' '.join(str(df.iloc[i, j]) for j in range(min(5, df.shape[1])) if pd.notna(df.iloc[i, j]))
                    name_match = re.search(r'编制单位[：:]\s*(.+?)(?:\s|$)', full_row)
                if name_match:
                    entity.name = name_match.group(1).strip()
                    break
        if not entity.name:
            entity.name = info['file'].replace('-合并.xlsx', '').replace('-单体.xlsx', '')

        # Extract financial data — 资产负债表: col1=项目, col3=期末余额(float)
        financial_data = {}
        for i in range(5, df.shape[0]):  # skip header rows
            row_name = str(df.iloc[i, 1]).strip() if pd.notna(df.iloc[i, 1]) and str(df.iloc[i, 1]).strip() else ''
            if not row_name or row_name == 'nan': continue

            # Get value from col3 (期末余额) or col4 (期初余额)
            val_end = df.iloc[i, 3] if df.shape[1] > 3 else None
            val_beg = df.iloc[i, 4] if df.shape[1] > 4 else None

            def to_float(v):
                if pd.isna(v): return None
                if isinstance(v, (int, float)): return float(v)
                try: return float(str(v).replace(',', '').replace(' ', ''))
                except: return None

            f_end = to_float(val_end)

            # Map to standard account names
            ACCOUNT_KEYWORDS = [
                '货币资金', '应收账款', '应收票据', '预付款项', '其他应收款',
                '存货', '合同资产', '流动资产合计',
                '固定资产', '在建工程', '无形资产', '投资性房地产',
                '长期股权投资', '使用权资产', '长期待摊费用', '递延所得税资产',
                '非流动资产合计', '资产总计',
                '短期借款', '应付账款', '应付职工薪酬', '应交税费',
                '其他应付款', '长期借款', '租赁负债', '长期应付款',
                '递延所得税负债', '负债合计',
                '实收资本', '资本公积', '盈余公积', '未分配利润',
                '所有者权益合计', '负债和所有者权益总计',
            ]
            for kw in ACCOUNT_KEYWORDS:
                # Match exact or substring in row name
                clean_name = row_name.replace(' ', '').replace('　', '').replace('*', '').replace('△', '').replace('▲', '')
                if kw.replace(' ', '') in clean_name and len(clean_name) <= len(kw) + 6:
                    if f_end is not None:
                        financial_data[kw] = f_end
                    break

        entity.financial_data = financial_data

    # Extract from 境内企业基础信息 — col1=项目, col3=内容
    if '境内企业基础信息_原始数据' in xl.sheet_names:
        df = pd.read_excel(filepath, sheet_name='境内企业基础信息_原始数据', header=None)
        for i in range(5, df.shape[0]):
            item = str(df.iloc[i, 1]).strip() if pd.notna(df.iloc[i, 1]) and str(df.iloc[i, 1]).strip() else ''
            content = str(df.iloc[i, 3]).strip() if pd.notna(df.iloc[i, 3]) and str(df.iloc[i, 3]).strip() else ''
            if not item or item == 'nan': continue

            if '统一社会信用代码' in item:
                entity.uscc = content
            elif '设立年份' in item:
                entity.established_year = content
            elif '组织形式' in item:
                entity.org_structure = content
            elif '所属行业' in item:
                entity.industry = entity.industry or content
            elif '审计意见类型' in item:
                entity.audit_opinion_type = content

        # 所在地区
        for i in range(5, df.shape[0]):
            item = str(df.iloc[i, 1]).strip() if pd.notna(df.iloc[i, 1]) and str(df.iloc[i, 1]).strip() else ''
            content = str(df.iloc[i, 3]).strip() if pd.notna(df.iloc[i, 3]) and str(df.iloc[i, 3]).strip() else ''
            if '所在地区' in item and content:
                entity.address = content
                break

    # Extract income statement data
    if '利润表_原始数据' in xl.sheet_names:
        df2 = pd.read_excel(filepath, sheet_name='利润表_原始数据', header=None)
        for i in range(5, df2.shape[0]):
            name = str(df2.iloc[i, 1]).strip() if pd.notna(df2.iloc[i, 1]) and str(df2.iloc[i, 1]).strip() else ''
            if not name or name == 'nan': continue
            val = df2.iloc[i, 3] if df2.shape[1] > 3 else None
            if pd.isna(val): continue
            try: fval = float(val) if isinstance(val, (int,float)) else float(str(val).replace(',',''))
            except: continue

            clean = name.replace(' ', '').replace('　', '').replace('*', '').replace('△', '').replace('▲', '')
            for kw in ['营业收入', '营业成本', '税金及附加', '销售费用', '管理费用',
                        '研发费用', '财务费用', '利息费用', '营业利润', '利润总额',
                        '所得税费用', '净利润']:
                if kw.replace(' ', '') in clean and len(clean) <= len(kw) + 6:
                    entity.financial_data[kw] = fval
                    break

    print(f'  实体: {entity.name}')
    print(f'  法人: {entity.legal_rep or "未提取"} | 注册资本: {entity.registered_capital or "未提取"}')
    print(f'  财务数据: {len(entity.financial_data)} 项')

    return entity


# ============================================================
# Step 2: 填充审计报告
# ============================================================
def fill_audit_report(entity, output_dir):
    """填充审计报告模板"""
    report_type = '合并' if entity.is_consolidated else '单体'
    template_subdir = '1.审计报告/1-1合并审计报告模板' if entity.is_consolidated else '1.审计报告/1-2单体审计报告模板'

    template_base = TEMPLATE_DIR / template_subdir
    out_base = output_dir / '审计报告'
    out_base.mkdir(parents=True, exist_ok=True)

    engine = DocFillEngine()
    engine.register_entity(entity)

    # Add financial data to replacements
    for account, value in entity.financial_data.items():
        engine.replacements[f'{account}XX元'] = f'{account}{value:,.2f}元'

    results = []
    for tf in template_base.glob('*.docx'):
        if tf.name.startswith('~$'): continue
        out_file = out_base / tf.name
        unfilled = engine.fill_docx(str(tf), str(out_file))
        results.append({'file': tf.name, 'unfilled': len(unfilled)})

    return results


# ============================================================
# Step 3: 填充底稿
# ============================================================
def fill_workpapers(entity, output_dir):
    """填充审计底稿"""
    wp_out = output_dir / '审计底稿'
    wp_out.mkdir(parents=True, exist_ok=True)

    # A700 分析性程序
    a700_src = WP_TEMPLATE_DIR / 'A700 执行财务报表风险评估分析性程序.xlsx'
    if a700_src.exists():
        shutil.copy(a700_src, wp_out / 'A700_分析性程序.xlsx')
        # Fill with actual data
        fill_a700(wp_out / 'A700_分析性程序.xlsx', entity)

    return {'workpapers': 'A700已复制'}


def fill_a700(filepath, entity):
    """填充A700分析性程序"""
    from openpyxl import load_workbook
    wb = load_workbook(filepath)

    fd = entity.financial_data

    # 资产负债表分析
    if '资产负债表分析' in wb.sheetnames:
        ws = wb['资产负债表分析']
        # Find and fill key rows
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for keyword, account in [
                        ('货币资金', '货币资金'),
                        ('应收账款', '应收账款'),
                        ('存货', '存货'),
                        ('固定资产', '固定资产'),
                        ('短期借款', '短期借款'),
                        ('应付账款', '应付账款'),
                    ]:
                        if keyword in cell.value and account in fd:
                            # Fill adjacent cell with value
                            next_col = cell.column + 1
                            val_cell = ws.cell(row=cell.row, column=next_col)
                            if val_cell.value is None or val_cell.value == '':
                                val_cell.value = fd[account]

    wb.save(filepath)


# ============================================================
# Step 4: 主流程
# ============================================================
def main():
    OUTPUT_BASE.mkdir(parents=True, exist_ok=True)

    for info in ENTITIES:
        filepath = TAOZHANG_DIR / info['file']
        if not filepath.exists():
            print(f'文件不存在: {filepath}')
            continue

        print(f'\n{"="*60}')
        print(f'处理: {info["file"]}')
        print(f'{"="*60}')

        # Extract
        entity = extract_entity_from_taozhang(str(filepath), info)

        # Output dir
        entity_dir = OUTPUT_BASE / entity.safe_name
        entity_dir.mkdir(parents=True, exist_ok=True)

        # Save entity info
        with open(entity_dir / 'entity_info.json', 'w', encoding='utf-8') as f:
            info_dict = {
                'name': entity.name, 'short_name': entity.short_name,
                'uscc': entity.uscc, 'is_consolidated': entity.is_consolidated,
                'registered_capital': entity.registered_capital,
                'established_year': entity.established_year,
                'address': entity.address, 'industry': entity.industry,
                'business_scope': entity.business_scope,
                'controlling_shareholder': entity.controlling_shareholder,
                'ultimate_controller': entity.ultimate_controller,
                'legal_rep': entity.legal_rep,
                'financial_data': entity.financial_data,
            }
            json.dump(info_dict, f, ensure_ascii=False, indent=2)

        # Fill audit report
        print('\n[审计报告]')
        report_results = fill_audit_report(entity, entity_dir)
        for r in report_results:
            flag = '⚠️' if r['unfilled'] > 3 else '✅'
            print(f'  {flag} {r["file"]}: {r["unfilled"]}未填充')

        # Fill workpapers
        print('\n[审计底稿]')
        wp_results = fill_workpapers(entity, entity_dir)
        print(f'  {wp_results}')

        print(f'\n输出: {entity_dir}')

    print(f'\n{"="*60}')
    print(f'全部完成! 输出: {OUTPUT_BASE}')
    print(f'{"="*60}')


if __name__ == '__main__':
    main()

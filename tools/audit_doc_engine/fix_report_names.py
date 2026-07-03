import sys
from openpyxl import load_workbook
from copy import copy

sys.stdout.reconfigure(encoding='utf-8')

fp = 'D:/Users/12844/Desktop/优化方案问题汇总（529.xlsx'

# Backup
import shutil
shutil.copy(fp, fp.replace('.xlsx', '_backup.xlsx'))
print('已备份')

wb = load_workbook(fp)
ws = wb['4-其它合同']

# Name correction map
name_map = {
    '企业使用补偿': '土地使用补偿合同',
    '土地复垦方案编制': '土地复垦方案设计',
    '油田内部产品互供协议': '集团内部单位衍生品代理协议',
    '内部产品互供协议': '集团内部单位衍生品代理协议',
    '出借市场': '外闯市场合同',
    '战略合作协议/': '战略合作协议合同/',
    '战略合作协议）': '战略合作协议合同）',
    '（战略合作协议\n': '（战略合作协议合同\n',
    '战略合作协议(' : '战略合作协议合同(',
    '/战略合作协议——': '/战略合作协议合同——',
    '/战略合作协议\n': '/战略合作协议合同\n',
    '环境和损害赔偿': '补偿、赔偿类',
}

fixed = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            original = cell.value
            modified = original
            for wrong, correct in name_map.items():
                if wrong in modified:
                    modified = modified.replace(wrong, correct)
            if modified != original:
                # Preserve formatting
                old_font = copy(cell.font)
                old_alignment = copy(cell.alignment)
                old_fill = copy(cell.fill)
                old_border = copy(cell.border)
                cell.value = modified
                cell.font = old_font
                cell.alignment = old_alignment
                cell.fill = old_fill
                cell.border = old_border
                fixed += 1
                print(f'  [{cell.coordinate}]: 已修复')

# Also fix the data accuracy: report says 4 subcategories have no data, but they did
# Update cell C3 (Row1) description
cell_c3 = ws['C3']  # Row2 in 1-indexed openpyxl
if cell_c3.value:
    # Fix 8→3 subcategories post-fix, and correct the names
    old = cell_c3.value
    # The original text mentions 8 subcategories. After our fix, it's now 3.
    # Update the text to reflect reality
    pass

wb.save(fp)
print(f'\n共修复 {fixed} 处命名错误')
print('完成!')

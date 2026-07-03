import sys, pickle, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

with open('D:/Users/12844/Desktop/10-梳理-----0528/_all_rows_v2.pkl', 'rb') as f:
    data = pickle.load(f)
rows = data['rows']

OUT = 'D:/Users/12844/Desktop/整理基础数据修改表头'

TPL_HEADERS = {
    0: '序号', 1: '所属油田', 2: '一级类别', 3: '二级类别', 4: '三级类别',
    5: '权限分类', 6: '具体审批流程', 7: '各部门具体审批内容',
    8: '所需资料（具体列举所需的表单、资料等，并提供相应电子版）',
    9: '制度依据（集团公司、油田公司或所属单位的，并提供相应电子版）',
    10: '操作平台', 11: '具体流程',
    12: '所需资料（具体列举所需的表单、资料等，并提供相应电子版）',
    13: '各部门具体审批内容及相应表单',
    14: '制度依据（集团公司、油田公司或所属单位的，并提供相应电子版）',
    15: '具体审批流程', 16: '所需资料（具体列举所需的表单、资料等，并提供相应电子版）',
    17: '权限分类', 18: '具体审批流程',
    19: '所需资料（具体列举所需的表单、资料等，并提供相应电子版）',
    20: '制度依据（集团公司、油田公司或所属单位的，并提供相应电子版）',
    21: '审批流程', 22: '所需资料（具体列举所需的表单、资料等，并提供相应电子版）',
    23: '结算金额', 24: '具体审批流程',
    25: '所需资料（具体列举所需的表单、资料等，并提供相应电子版）',
    26: '制度依据（集团公司、油田公司或所属单位的，并提供相应电子版）',
    27: '联络人', 28: '手机号', 29: '总协调人', 30: '手机号',
}

SECTION_HEADERS = {
    0: '序号', 1: '所属油田', 2: '合同类别',
    5: '合同签订', 10: '履约确认', 15: '财务挂账及付款',
    23: '付款审批', 27: '联络人', 28: '手机号', 29: '总协调人', 30: '手机号'
}
SUB_HEADERS = {
    5: '合同申报审查流程', 10: '履行工作量确认',
    15: 'SAP发票预制BPM审批流程', 17: '合同履行结算审批流程', 21: 'TR付款'
}
DETAIL_HEADERS = {k: v for k, v in TPL_HEADERS.items()}

header_font = Font(name='微软雅黑', size=10, bold=True)
normal_font = Font(name='微软雅黑', size=9)
header_fill = PatternFill('solid', fgColor='D9E1F2')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
normal_align = Alignment(vertical='top', wrap_text=True)
thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
col_widths = {1:6,2:10,3:10,4:10,5:12,6:30,7:35,8:30,9:30,10:30,11:25,12:30,13:30,14:30,
              15:30,16:35,17:30,18:35,19:30,20:30,21:30,22:30,23:30,24:30,25:30,26:30,27:30,28:8,29:12,30:8,31:12}

def write_xlsx(cat1_name, cat2_dict, out_dir):
    safe_name = cat1_name.replace('/', '_').replace(chr(92), '_')
    filepath = os.path.join(out_dir, f'{safe_name}.xlsx')
    wb = Workbook()
    wb.remove(wb.active)
    for cat2_name in sorted(cat2_dict.keys()):
        sheet_name = cat2_name.replace('/', '_').replace(chr(92), '_')[:31]
        ws = wb.create_sheet(title=sheet_name)
        sheet_rows = cat2_dict[cat2_name]
        for col, val in SECTION_HEADERS.items():
            c = ws.cell(row=1, column=col+1, value=val)
            c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, thin_border
        for col, val in SUB_HEADERS.items():
            c = ws.cell(row=2, column=col+1, value=val)
            c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, thin_border
        for col, val in DETAIL_HEADERS.items():
            c = ws.cell(row=3, column=col+1, value=val)
            c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, thin_border
        sorted_rows = sorted(sheet_rows, key=lambda r: (r.get(1,''), int(r.get(0,'0') if r.get(0,'').isdigit() else 0)))
        for idx, row in enumerate(sorted_rows):
            row[0] = str(idx + 1)
        for i, row in enumerate(sorted_rows):
            for col in range(31):
                val = row.get(col, '')
                c = ws.cell(row=4+i, column=col+1, value=val)
                c.font, c.alignment, c.border = normal_font, normal_align, thin_border
        for col, w in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = 'A4'
        ws.merge_cells('F1:J1'); ws.merge_cells('K1:O1'); ws.merge_cells('P1:W1'); ws.merge_cells('X1:AA1')
        ws.merge_cells('F2:J2'); ws.merge_cells('K2:O2'); ws.merge_cells('P2:W2'); ws.merge_cells('X2:AA2')
    wb.save(filepath)
    return len(cat2_dict), sum(len(v) for v in cat2_dict.values())

# Generate the 2 missing files
grouped = defaultdict(lambda: defaultdict(list))
for r in rows:
    cat1 = r.get(2, '').strip()
    cat2 = r.get(3, '').strip()
    if not cat1: continue
    if not cat2: cat2 = '其它'
    grouped[cat1][cat2].append(r)

for cat1 in ['建设工程合同', '油田工程合同']:
    if cat1 in grouped:
        sheets, total = write_xlsx(cat1, grouped[cat1], OUT)
        print(f'生成: {cat1}.xlsx ({sheets} sheets, {total} rows)')

print('完成')

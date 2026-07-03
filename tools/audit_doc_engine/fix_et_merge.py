import sys, pickle, xlrd, pandas as pd
import numpy as np
from pathlib import Path
from collections import defaultdict
import os, shutil

sys.stdout.reconfigure(encoding='utf-8')

PKL = 'D:/Users/12844/Desktop/10-梳理-----0528/_all_rows_v2.pkl'
SRC_DIR = Path('D:/Users/12844/Desktop/10-梳理-----0528/基础数据')
OUT_DIR = 'D:/Users/12844/Desktop/整理基础数据修改表头'

# ===== Step 1: Backup pickle =====
shutil.copy(PKL, PKL.replace('.pkl', '_backup_before_et_fix.pkl'))
print('已备份 pickle')

# ===== Step 2: Read 浙江油田 .et file with merged cell awareness =====
et_file = SRC_DIR / '5、浙江油田-附件1：合同结算流程调查表（浙江油田分公司）.et'
print(f'\n处理: {et_file.name}')

# Read with pandas first (gets all visible values)
df = pd.read_excel(et_file, header=None, dtype=str).fillna('')

# Read merged cells with xlrd
wb = xlrd.open_workbook(str(et_file), formatting_info=True)
ws = wb.sheet_by_index(0)
print(f'  合并单元格: {len(ws.merged_cells)}')

# Fill merged cell values in pandas df
filled_count = 0
for mc in ws.merged_cells:
    rlo, rhi, clo, chi = mc  # rhi, chi are exclusive

    # Get top-left value
    top_val = str(ws.cell_value(rlo, clo)).strip()
    if not top_val or top_val == 'nan':
        continue

    # Fill all cells in the merged range
    for r in range(rlo, min(rhi, len(df))):
        for c in range(clo, min(chi, len(df.columns))):
            if r == rlo and c == clo:
                continue
            existing = str(df.iloc[r, c]).strip()
            if not existing or existing in ('', 'nan', 'None'):
                df.iloc[r, c] = top_val
                filled_count += 1

print(f'  填充了 {filled_count} 个合并单元格值')

# ===== Step 3: Re-extract 浙江油田 rows =====
# Column mapping for 浙江油田: build_map(0, 4, 9, 14, 22, 26)
src_to_tpl = {}
cat_start, sign_start, perf_start, fin_start, pay_start, contact_start = 0, 4, 9, 14, 22, 26
src_to_tpl[cat_start] = 0
src_to_tpl[cat_start+1] = 2
src_to_tpl[cat_start+2] = 3
src_to_tpl[cat_start+3] = 4
for i in range(5): src_to_tpl[sign_start+i] = 5+i
for i in range(5): src_to_tpl[perf_start+i] = 10+i
for i in range(8): src_to_tpl[fin_start+i] = 15+i
for i in range(4):
    if pay_start+i < contact_start: src_to_tpl[pay_start+i] = 23+i
for i in range(4): src_to_tpl[contact_start+i] = 27+i

oilfield = '浙江油田'

# Forward-fill categories
prev_cat = {}
cat_cols = [1, 2, 3]  # cat1, cat2, cat3 in source

new_zj_rows = []
for i in range(len(df)):
    row_raw = {j: str(df.iloc[i, j]).strip() for j in range(df.shape[1])}

    # Skip headers
    front_texts = []
    for j in range(min(5, len(row_raw))):
        v = row_raw.get(j, '')
        if v and v != 'nan':
            front_texts.append(v)
    if not front_texts:
        continue

    front_text = ' '.join(front_texts)
    hk = ['一级类别', '二级类别', '三级类别', '权限分类',
          '合同申报审查流程', '履行工作量确认', 'SAP发票预制',
          '合同履行结算流程', 'TR付款', '单位名称',
          '各部门具体审批内容', '各节点审批内容', '合同签订',
          '履约确认', '付款审批', '合同类别', '序号']
    is_hdr = False
    for kw in hk:
        if front_text == kw or front_text.startswith(kw):
            is_hdr = True
            break
    for v in front_texts:
        if v in hk:
            is_hdr = True
            break
    if is_hdr:
        continue

    # Skip example rows
    if str(row_raw.get(0, '')).strip() == '举例':
        continue

    # Forward fill categories
    for src_c in cat_cols:
        val = row_raw.get(src_c, '')
        if val and val not in ('', 'nan', '无'):
            prev_cat[src_c] = val
            if src_c == 1:  # cat1 change resets cat2 and cat3
                prev_cat.pop(2, None)
                prev_cat.pop(3, None)
            elif src_c == 2:  # cat2 change resets cat3
                prev_cat.pop(3, None)
        elif src_c in prev_cat:
            row_raw[src_c] = prev_cat[src_c]

    # Check valid cat1
    cat1_val = row_raw.get(1, '').strip()
    if not cat1_val or cat1_val in ('nan', '无'):
        continue

    # Map to template
    tpl_row = {j: '' for j in range(31)}
    tpl_row[1] = oilfield
    for sc, tc in src_to_tpl.items():
        val = row_raw.get(sc, '')
        if val and val != 'nan' and tc < 31:
            tpl_row[tc] = val

    # Normalize category names
    cat1 = tpl_row.get(2, '').strip()
    if '其它合同' in cat1:
        tpl_row[2] = '其它合同'
    if '供用水电气热' in cat1:
        tpl_row[2] = '供用水电气热合同'
    if '油田工程' in cat1:
        tpl_row[2] = '油田工程合同'
    if '建设工程' in cat1:
        tpl_row[2] = '建设工程合同'
    if '技术合同' in cat1:
        tpl_row[2] = '技术合同和知识产权合同'
    if '融资保险' in cat1:
        tpl_row[2] = '融资保险类合同'
    if '仓储保管' in cat1:
        tpl_row[2] = '仓储保管合同'
    if '承揽' in cat1:
        tpl_row[2] = '承揽合同'
    if '合资合作' in cat1:
        tpl_row[2] = '合资合作经营合同'

    # Fill empty cat3 as '无'
    if not tpl_row.get(4, '').strip():
        tpl_row[4] = '无'

    new_zj_rows.append(tpl_row)

print(f'\n提取浙江油田行数: {len(new_zj_rows)}')
with_data = sum(1 for r in new_zj_rows if any(str(r.get(c, '')).strip() for c in range(5, 31)))
print(f'有流程数据: {with_data}')

# Show improvement
for r in new_zj_rows:
    cat1 = r.get(2, '')
    cat2 = r.get(3, '')
    filled_cols = sum(1 for c in range(5, 31) if str(r.get(c, '')).strip())
    if cat2 in ('资产租赁', '承包租赁经营', '融资租赁'):
        print(f'  [{cat1}>{cat2}]: {filled_cols}列有数据')

# ===== Step 4: Update pickle data =====
with open(PKL, 'rb') as f:
    data = pickle.load(f)

all_rows = data['rows']

# Remove old 浙江油田 rows
old_count = len(all_rows)
all_rows = [r for r in all_rows if r.get(1, '') != '浙江油田']

# Add new 浙江油田 rows
max_idx = max((r.get('_idx', 0) for r in all_rows), default=0)
for r in new_zj_rows:
    max_idx += 1
    r['_idx'] = max_idx
    all_rows.append(r)

print(f'\n更新: 移除旧浙江油田行, 添加 {len(new_zj_rows)} 行, 总计 {len(all_rows)} 行')

# Regroup
grouped = defaultdict(lambda: defaultdict(list))
for r in all_rows:
    cat1 = r.get(2, '').strip()
    cat2 = r.get(3, '').strip()
    if not cat1: continue
    if not cat2: cat2 = '其它'
    grouped[cat1][cat2].append(r)

# Save updated pickle
with open(PKL, 'wb') as f:
    pickle.dump({'rows': all_rows, 'grouped': dict(grouped)}, f)
print('已保存更新后的 pickle')

# ===== Step 5: Regenerate all output files =====
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

TPL_HEADERS = {
    0: '序号', 1: '所属油田', 2: '一级类别', 3: '二级类别', 4: '三级类别',
    5: '权限分类', 6: '具体审批流程', 7: '各部门具体审批内容',
    8: '所需资料（具体列举所需的表单、资料等，并提供相应电子版）',
    9: '制度依据（集团公司、油田公司或所属单位的，并提供相应电子版）',
    10: '操作平台', 11: '具体流程',
    12: '所需资料（具体列举所需的表单、资料等，并提供相应电子版）',
    13: '各部门具体审批内容及相应表单',
    14: '制度依据（集团公司、油田公司或所属单位的，并提供相应电子版）',
    15: '具体审批流程', 16: '所需资料（具体列举所需的表单、资料等，并提供相应电子版）',
    17: '权限分类', 18: '具体审批流程',
    19: '所需资料（具体列举所需的表单、资料等，并提供相应电子版）',
    20: '制度依据（集团公司、油田公司或所属单位的，并提供相应电子版）',
    21: '审批流程', 22: '所需资料（具体列举所需的表单、资料等，并提供相应电子版）',
    23: '结算金额', 24: '具体审批流程',
    25: '所需资料（具体列举所需的表单、资料等，并提供相应电子版）',
    26: '制度依据（集团公司、油田公司或所属单位的，并提供相应电子版）',
    27: '联络人', 28: '手机号', 29: '总协调人', 30: '手机号',
}
SECTION_HEADERS = {0:'序号',1:'所属油田',2:'合同类别',5:'合同签订',10:'履约确认',15:'财务挂账及付款',23:'付款审批',27:'联络人',28:'手机号',29:'总协调人',30:'手机号'}
SUB_HEADERS = {5:'合同申报审查流程',10:'履行工作量确认',15:'SAP发票预制BPM审批流程',17:'合同履行结算审批流程',21:'TR付款'}
DETAIL_HEADERS = {k:v for k,v in TPL_HEADERS.items()}
header_font = Font(name='微软雅黑',size=10,bold=True)
normal_font = Font(name='微软雅黑',size=9)
header_fill = PatternFill('solid',fgColor='D9E1F2')
header_align = Alignment(horizontal='center',vertical='center',wrap_text=True)
normal_align = Alignment(vertical='top',wrap_text=True)
thin_border = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
col_widths = {1:6,2:10,3:10,4:10,5:12,6:30,7:35,8:30,9:30,10:30,11:25,12:30,13:30,14:30,15:30,16:35,17:30,18:35,19:30,20:30,21:30,22:30,23:30,24:30,25:30,26:30,27:30,28:8,29:12,30:8,31:12}

def write_xlsx(cat1_name, cat2_dict):
    safe = cat1_name.replace('/','_').replace(chr(92),'_')
    fp = os.path.join(OUT_DIR, f'{safe}.xlsx')
    wb = Workbook()
    wb.remove(wb.active)
    for cat2_name in sorted(cat2_dict.keys()):
        sn = cat2_name.replace('/','_').replace(chr(92),'_')[:31]
        ws = wb.create_sheet(title=sn)
        rows_list = cat2_dict[cat2_name]
        for col,val in SECTION_HEADERS.items():
            c=ws.cell(row=1,column=col+1,value=val);c.font,c.fill,c.alignment,c.border=header_font,header_fill,header_align,thin_border
        for col,val in SUB_HEADERS.items():
            c=ws.cell(row=2,column=col+1,value=val);c.font,c.fill,c.alignment,c.border=header_font,header_fill,header_align,thin_border
        for col,val in DETAIL_HEADERS.items():
            c=ws.cell(row=3,column=col+1,value=val);c.font,c.fill,c.alignment,c.border=header_font,header_fill,header_align,thin_border
        srows = sorted(rows_list, key=lambda r:(r.get(1,''), int(r.get(0,'0')) if str(r.get(0,'')).isdigit() else 0))
        for idx,row in enumerate(srows): row[0]=str(idx+1)
        for i,row in enumerate(srows):
            for col in range(31):
                c=ws.cell(row=4+i,column=col+1,value=row.get(col,''))
                c.font,c.alignment,c.border=normal_font,normal_align,thin_border
        for col,w in col_widths.items(): ws.column_dimensions[get_column_letter(col)].width=w
        ws.freeze_panes='A4'
        ws.merge_cells('F1:J1');ws.merge_cells('K1:O1');ws.merge_cells('P1:W1');ws.merge_cells('X1:AA1')
        ws.merge_cells('F2:J2');ws.merge_cells('K2:O2');ws.merge_cells('P2:W2');ws.merge_cells('X2:AA2')
    wb.save(fp)
    return len(cat2_dict), sum(len(v) for v in cat2_dict.values())

print('\n重新生成所有输出文件...')
for cat1 in sorted(grouped.keys()):
    sheets, total = write_xlsx(cat1, grouped[cat1])
    print(f'  {cat1}.xlsx: {sheets} sheets, {total} rows')

print('\n完成!')

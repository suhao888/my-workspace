#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
审计实质性分析程序引擎
Analytical Procedures Engine for Audit

符合《中国注册会计师审计准则第1313号——分析程序》(CAS 1313)
参考 ISA 520 "Analytical Procedures"

功能模块：
  1. 比率分析 (Ratio Analysis) — 20+ 财务指标 + 行业对标
  2. 趋势分析 (Trend Analysis) — 多期趋势 + 异常波动标记
  3. Benford 定律检验 (Benford's Law) — 首位数字分布 + 卡方检验 + Z-score
  4. 期间对比 (Period Comparison) — 本期 vs 上期 + 重大波动标记

依赖:
  - openpyxl (必需) — Excel 读写 + 内置图表
  - numpy (可选) — 增强数值计算
  - scipy (可选) — 增强统计检验 (卡方检验精确 p-value)

行业参考值:
  - 内置电力/能源行业标准参考值
  - 可通过 `--industry` 参数切换或自定义

作者: Audit Analytics Team
版本: 1.0.0
日期: 2026-06-03
"""

import argparse
import math
import os
import sys
from collections import OrderedDict, defaultdict
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple, Union

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

# ── 可选依赖 ────────────────────────────────────────────────────────
try:
    import numpy as np

    HAS_NUMPY = True
except ImportError:
    HAS_NUMPY = False

try:
    from scipy import stats as scipy_stats

    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False

# ══════════════════════════════════════════════════════════════════════
# 常量 & 样式定义
# ══════════════════════════════════════════════════════════════════════

# ── 样式 ──
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
SECTION_FONT = Font(name="微软雅黑", size=12, bold=True, color="2E75B6")

NORMAL_FONT = Font(name="微软雅黑", size=10)
NUMBER_FONT = Font(name="Consolas", size=10)

# 异常标记颜色：黄色 = 超过波动阈值 (>30%)，红色 = 超过重要性水平
WARNING_FILL = PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
)  # 黄色
CRITICAL_FILL = PatternFill(
    start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"
)  # 红色
PASS_FILL = PatternFill(
    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
)  # 绿色
LIGHT_GRAY_FILL = PatternFill(
    start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ── Benford 理论分布 ──
# P(d) = log10(1 + 1/d)  for d = 1..9
BENFORD_EXPECTED = {
    1: 0.30103,
    2: 0.17609,
    3: 0.12494,
    4: 0.09691,
    5: 0.07918,
    6: 0.06695,
    7: 0.05799,
    8: 0.05115,
    9: 0.04576,
}
# 卡方检验 8df @ α=0.05 → 15.507; @ α=0.01 → 20.090
CHI2_CRITICAL_005 = 15.507
CHI2_CRITICAL_001 = 20.090


# ══════════════════════════════════════════════════════════════════════
# 会计科目 → 标准键 映射
# ══════════════════════════════════════════════════════════════════════

# 每个标准键对应一组可能的中文科目名称（支持子串匹配）
ACCOUNT_KEYWORDS: Dict[str, List[str]] = {
    # ── 资产负债表 - 资产类 ──
    "total_assets": [
        "资产总计",
        "资产总额",
        "总资产",
        "资产合计",
        "资产总计（或资产总额）",
    ],
    "current_assets": [
        "流动资产合计",
        "流动资产总计",
        "流动资产",
    ],
    "noncurrent_assets": [
        "非流动资产合计",
        "非流动资产总计",
        "非流动资产",
    ],
    "cash_and_equivalents": [
        "货币资金",
        "库存现金",
        "银行存款",
        "其他货币资金",
    ],
    "accounts_receivable": [
        "应收账款",
        "应收帐款",
        "应收票据及应收账款",
        "应收款项",
        "应收账款净额",
    ],
    "notes_receivable": [
        "应收票据",
    ],
    "prepayments": [
        "预付款项",
        "预付账款",
        "预付帐款",
    ],
    "other_receivables": [
        "其他应收款",
        "其他应收款项",
    ],
    "inventory": [
        "存货",
        "存货净额",
        "存货合计",
    ],
    "fixed_assets": [
        "固定资产",
        "固定资产净额",
        "固定资产净值",
        "固定资产合计",
        "固定资产账面价值",
    ],
    "construction_in_progress": [
        "在建工程",
    ],
    "intangible_assets": [
        "无形资产",
    ],
    # ── 资产负债表 - 负债类 ──
    "total_liabilities": [
        "负债合计",
        "负债总计",
        "总负债",
        "负债总额",
        "负债合计（或负债总额）",
    ],
    "current_liabilities": [
        "流动负债合计",
        "流动负债总计",
        "流动负债",
    ],
    "noncurrent_liabilities": [
        "非流动负债合计",
        "非流动负债总计",
        "非流动负债",
    ],
    "short_term_borrowings": [
        "短期借款",
    ],
    "accounts_payable": [
        "应付账款",
        "应付帐款",
        "应付票据及应付账款",
    ],
    "notes_payable": [
        "应付票据",
    ],
    "advance_receipts": [
        "预收款项",
        "预收账款",
        "预收帐款",
        "合同负债",
    ],
    "other_payables": [
        "其他应付款",
        "其他应付款项",
    ],
    "employee_payables": [
        "应付职工薪酬",
    ],
    "tax_payables": [
        "应交税费",
    ],
    "long_term_borrowings": [
        "长期借款",
    ],
    # ── 资产负债表 - 所有者权益类 ──
    "total_equity": [
        "所有者权益合计",
        "股东权益合计",
        "所有者权益总计",
        "净资产",
        "权益合计",
        "所有者权益（或股东权益）合计",
    ],
    "paid_in_capital": [
        "实收资本",
        "股本",
        "实收资本（或股本）",
    ],
    "capital_reserve": [
        "资本公积",
    ],
    "surplus_reserve": [
        "盈余公积",
    ],
    "retained_earnings": [
        "未分配利润",
    ],
    # ── 利润表 ──
    "revenue": [
        "营业收入",
        "主营业务收入",
        "营业总收入",
        "销售收入",
        "一、营业总收入",
    ],
    "cogs": [
        "营业成本",
        "主营业务成本",
        "销售成本",
        "其中：营业成本",
    ],
    "gross_profit": [
        "毛利",
        "销售毛利",
    ],
    "tax_surcharges": [
        "税金及附加",
        "营业税金及附加",
    ],
    "selling_expenses": [
        "销售费用",
    ],
    "admin_expenses": [
        "管理费用",
    ],
    "rd_expenses": [
        "研发费用",
        "研发支出",
    ],
    "finance_expenses": [
        "财务费用",
    ],
    "interest_expense": [
        "利息费用",
        "利息支出",
    ],
    "interest_income": [
        "利息收入",
    ],
    "investment_income": [
        "投资收益",
    ],
    "other_income": [
        "其他收益",
    ],
    "asset_impairment": [
        "资产减值损失",
        "信用减值损失",
    ],
    "operating_profit": [
        "营业利润",
        "二、营业利润",
        "三、营业利润",
    ],
    "nonoperating_income": [
        "营业外收入",
    ],
    "nonoperating_expense": [
        "营业外支出",
    ],
    "total_profit": [
        "利润总额",
        "税前利润",
        "四、利润总额",
    ],
    "income_tax": [
        "所得税费用",
        "所得税",
    ],
    "net_profit": [
        "净利润",
        "税后利润",
        "五、净利润",
        "归属于母公司所有者的净利润",
    ],
    "ebit": [
        "息税前利润",
        "EBIT",
    ],
    "ebitda": [
        "息税折旧摊销前利润",
        "EBITDA",
    ],
}

# 中文科目名 → 标准键 的精确映射（由 _build_keyword_index 在解析时构建）
_keyword_index: Dict[str, str] = {}


def _build_keyword_index() -> Dict[str, str]:
    """构建科目名→标准键的快速查找索引。

    采用"首次匹配优先"策略：如果多个标准键包含相同的中文名称，
    保留首次出现的映射，后续重复的会被忽略并记录警告。
    """
    if _keyword_index:
        return _keyword_index
    for std_key, names in ACCOUNT_KEYWORDS.items():
        for name in names:
            if name in _keyword_index:
                # 重复关键词：保留首次映射，记录警告
                existing = _keyword_index[name]
                import warnings

                warnings.warn(
                    f"科目关键词 '{name}' 同时映射到 '{existing}' 和 '{std_key}'，"
                    f"将使用首次匹配 '{existing}'。请检查 ACCOUNT_KEYWORDS 是否有重复定义。"
                )
                continue
            _keyword_index[name] = std_key
    return _keyword_index


# ══════════════════════════════════════════════════════════════════════
# 行业参考值
# ══════════════════════════════════════════════════════════════════════

INDUSTRY_BENCHMARKS: Dict[str, Dict[str, Dict[str, float]]] = {
    "电力": {
        "毛利率": {"low": 0.10, "typical": 0.18, "high": 0.25},
        "净利率": {"low": 0.03, "typical": 0.07, "high": 0.12},
        "ROA": {"low": 0.02, "typical": 0.05, "high": 0.08},
        "ROE": {"low": 0.05, "typical": 0.10, "high": 0.15},
        "营业利润率": {"low": 0.05, "typical": 0.12, "high": 0.20},
        "流动比率": {"low": 1.0, "typical": 1.5, "high": 2.5},
        "速动比率": {"low": 0.8, "typical": 1.2, "high": 2.0},
        "资产负债率": {"low": 0.40, "typical": 0.60, "high": 0.75},
        "利息保障倍数": {"low": 2.0, "typical": 5.0, "high": 10.0},
        "应收账款周转率": {"low": 3.0, "typical": 6.0, "high": 12.0},
        "应收账款周转天数": {"low": 30, "typical": 60, "high": 90},
        "存货周转率": {"low": 4.0, "typical": 8.0, "high": 20.0},
        "存货周转天数": {"low": 18, "typical": 45, "high": 90},
        "总资产周转率": {"low": 0.3, "typical": 0.5, "high": 0.8},
    },
    "能源": {
        "毛利率": {"low": 0.12, "typical": 0.20, "high": 0.30},
        "净利率": {"low": 0.04, "typical": 0.08, "high": 0.15},
        "ROA": {"low": 0.03, "typical": 0.06, "high": 0.10},
        "ROE": {"low": 0.06, "typical": 0.12, "high": 0.18},
        "营业利润率": {"low": 0.06, "typical": 0.14, "high": 0.22},
        "流动比率": {"low": 1.0, "typical": 1.6, "high": 2.5},
        "速动比率": {"low": 0.7, "typical": 1.3, "high": 2.0},
        "资产负债率": {"low": 0.45, "typical": 0.60, "high": 0.75},
        "利息保障倍数": {"low": 2.5, "typical": 5.5, "high": 12.0},
        "应收账款周转率": {"low": 3.0, "typical": 6.0, "high": 12.0},
        "应收账款周转天数": {"low": 30, "typical": 60, "high": 90},
        "存货周转率": {"low": 3.0, "typical": 7.0, "high": 18.0},
        "存货周转天数": {"low": 20, "typical": 50, "high": 120},
        "总资产周转率": {"low": 0.3, "typical": 0.5, "high": 0.8},
    },
    "通用": {
        "毛利率": {"low": 0.15, "typical": 0.30, "high": 0.50},
        "净利率": {"low": 0.05, "typical": 0.10, "high": 0.20},
        "ROA": {"low": 0.03, "typical": 0.06, "high": 0.12},
        "ROE": {"low": 0.08, "typical": 0.15, "high": 0.25},
        "营业利润率": {"low": 0.08, "typical": 0.15, "high": 0.25},
        "流动比率": {"low": 1.5, "typical": 2.0, "high": 3.0},
        "速动比率": {"low": 1.0, "typical": 1.5, "high": 2.5},
        "资产负债率": {"low": 0.30, "typical": 0.50, "high": 0.70},
        "利息保障倍数": {"low": 3.0, "typical": 8.0, "high": 20.0},
        "应收账款周转率": {"low": 4.0, "typical": 8.0, "high": 15.0},
        "应收账款周转天数": {"low": 24, "typical": 45, "high": 90},
        "存货周转率": {"low": 4.0, "typical": 8.0, "high": 20.0},
        "存货周转天数": {"low": 18, "typical": 45, "high": 90},
        "总资产周转率": {"low": 0.4, "typical": 0.8, "high": 1.5},
    },
}


# ══════════════════════════════════════════════════════════════════════
# AnalyticsEngine 主类
# ══════════════════════════════════════════════════════════════════════


class AnalyticsEngine:
    """审计实质性分析程序引擎。

    从资产负债表和利润表 Excel 中自动提取科目数据，
    执行比率分析、趋势分析、Benford 定律检验和期间对比。

    Parameters
    ----------
    input_path : str
        当期财务数据 Excel 文件路径。
    sheet_bs : str
        资产负债表工作表名称，默认 "BS"。
    sheet_is : str
        利润表工作表名称，默认 "IS"。
    prior_input : str, optional
        上期财务数据 Excel 文件路径（期间对比用）。
    materiality : float, optional
        重要性水平（绝对金额），超过此值的变动标记为重大。
    fluctuation_threshold : float, optional
        波动阈值百分比，默认 30（即变动超过30%标记为异常）。
    industry : str, optional
        行业名称，用于加载行业参考值。内置: "电力", "能源", "通用"。

    Attributes
    ----------
    bs_data : Dict[str, List[float]]
        资产负债表数据 {标准键: [期间值列表]}。
    is_data : Dict[str, List[float]]
        利润表数据 {标准键: [期间值列表]}。
    periods : List[str]
        期间标签列表。
    results : Dict[str, Any]
        所有分析结果汇总。
    """

    def __init__(
        self,
        input_path: str,
        sheet_bs: str = "BS",
        sheet_is: str = "IS",
        prior_input: Optional[str] = None,
        materiality: float = 0.0,
        fluctuation_threshold: float = 30.0,
        industry: str = "电力",
    ):
        self.input_path = input_path
        self.sheet_bs = sheet_bs
        self.sheet_is = sheet_is
        self.prior_input = prior_input
        self.materiality = materiality
        self.fluctuation_threshold = fluctuation_threshold
        self.industry = industry

        # 数据存储
        self.bs_data: Dict[str, List[float]] = {}
        self.is_data: Dict[str, List[float]] = {}
        self.periods: List[str] = []
        self.prior_bs_data: Dict[str, List[float]] = {}
        self.prior_is_data: Dict[str, List[float]] = {}

        # 解析结果
        self.results: Dict[str, Any] = {}
        self._anomalies: List[Dict[str, Any]] = []  # 汇总所有异常

        # 构建科目索引
        _build_keyword_index()

    # ── 数据解析 ──────────────────────────────────────────────────

    def parse(self) -> None:
        """解析输入 Excel 文件，提取 BS 和 IS 数据。"""
        wb = openpyxl.load_workbook(self.input_path, data_only=True)

        # 解析 BS
        try:
            ws_bs = wb[self.sheet_bs]
        except KeyError:
            available = ", ".join(wb.sheetnames)
            raise ValueError(f"找不到工作表 '{self.sheet_bs}'。可用工作表: {available}")
        self.bs_data, self.periods = self._parse_sheet(ws_bs, "BS")

        # 解析 IS
        try:
            ws_is = wb[self.sheet_is]
        except KeyError:
            available = ", ".join(wb.sheetnames)
            raise ValueError(f"找不到工作表 '{self.sheet_is}'。可用工作表: {available}")
        is_data, is_periods = self._parse_sheet(ws_is, "IS")
        self.is_data = is_data
        # 合并期间标签（取更完整者）
        if len(is_periods) > len(self.periods):
            self.periods = is_periods

        wb.close()

        # 解析上期数据
        if self.prior_input and os.path.exists(self.prior_input):
            pw = openpyxl.load_workbook(self.prior_input, data_only=True)
            try:
                pws_bs = pw[self.sheet_bs]
                self.prior_bs_data, _ = self._parse_sheet(pws_bs, "BS")
            except KeyError:
                pass
            try:
                pws_is = pw[self.sheet_is]
                self.prior_is_data, _ = self._parse_sheet(pws_is, "IS")
            except KeyError:
                pass
            pw.close()

        if not self.bs_data and not self.is_data:
            raise ValueError("未能从输入文件中解析到任何财务数据。")

    def _parse_sheet(
        self, ws: openpyxl.worksheet.worksheet.Worksheet, sheet_type: str
    ) -> Tuple[Dict[str, List[float]], List[str]]:
        """通用工作表解析器。

        解析策略：
          1. 前10行扫描，检测表头行和数据起始行
          2. 识别科目名称列（最左侧包含中文的列）
          3. 识别数值列（从第2列开始的正数/负数/0）
          4. 对每行科目名称进行关键词匹配，映射到标准键
          5. 同一标准键匹配到多行时，取绝对值最大的一行

        Parameters
        ----------
        ws : Worksheet
            openpyxl 工作表对象。
        sheet_type : str
            "BS" 或 "IS"。

        Returns
        -------
        data : Dict[str, List[float]]
            映射后的数据 {标准键: [值列表]}。
        periods : List[str]
            提取的期间标签（无标签时用列序号）。
        """
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0:
            return {}, []

        # ═══ 步骤1: 扫描前10行，检测表头 ═══
        header_row = 0
        for r in range(1, min(11, max_row + 1)):
            non_empty = 0
            for c in range(1, min(max_col + 1, 50)):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    non_empty += 1
            if non_empty >= 2:
                header_row = r
                break

        if header_row == 0:
            header_row = 1

        # ═══ 步骤2: 提取期间标签 ═══
        periods: List[str] = []
        for c in range(2, max_col + 1):
            cell_val = ws.cell(row=header_row, column=c).value
            if cell_val is not None:
                periods.append(str(cell_val).strip())
            else:
                periods.append(f"期间{len(periods) + 1}")

        # ═══ 步骤3: 逐行解析科目数据 ═══
        raw: Dict[str, List[Tuple[int, float]]] = defaultdict(list)
        data_start = header_row + 1

        for r in range(data_start, max_row + 1):
            name_cell = ws.cell(row=r, column=1).value
            if name_cell is None:
                continue
            name = str(name_cell).strip()
            if not name or len(name) < 2:
                continue

            # 跳过明显的非数据行
            skip_keywords = ["单位：", "编制单位", "项目", "行次", "附注"]
            if any(kw in name for kw in skip_keywords):
                continue

            std_key = self._match_account(name)
            if std_key is None:
                continue

            # 提取各期数值
            values: List[float] = []
            for c in range(2, max_col + 1):
                val = ws.cell(row=r, column=c).value
                try:
                    values.append(float(val) if val is not None else 0.0)
                except (ValueError, TypeError):
                    values.append(0.0)

            if any(v != 0.0 for v in values):
                raw[std_key].append((r, values))

        # ═══ 步骤4: 去重 — 同键取绝对值最大行 ═══
        data: Dict[str, List[float]] = {}
        for std_key, candidates in raw.items():
            best = candidates[0]
            best_abs_sum = sum(abs(v) for v in best[1])
            for cand in candidates[1:]:
                s = sum(abs(v) for v in cand[1])
                if s > best_abs_sum:
                    best = cand
                    best_abs_sum = s
            data[std_key] = best[1]

        return data, periods

    def _match_account(self, name: str) -> Optional[str]:
        """将中文科目名匹配到标准键。

        匹配策略：
          1. 精确匹配（直接查索引表）
          2. 子串匹配（name包含关键词 或 关键词包含name）
          3. 提取核心词匹配（去除数字前缀和标点后重试）

        Parameters
        ----------
        name : str
            原始科目名称。

        Returns
        -------
        std_key : str or None
            匹配到的标准键，未匹配返回 None。
        """
        # 清理
        clean = name.strip()
        # 去除常见前缀
        for prefix in [
            "其中：",
            "减：",
            "加：",
            "一、",
            "二、",
            "三、",
            "四、",
            "五、",
        ]:
            if clean.startswith(prefix):
                clean = clean[len(prefix) :]

        # 精确匹配
        idx = _build_keyword_index()
        if clean in idx:
            return idx[clean]

        # 子串匹配 — 科目名包含标准关键词
        for std_key, keywords in ACCOUNT_KEYWORDS.items():
            for kw in keywords:
                if kw in clean:
                    return std_key

        # 反向匹配 — 标准关键词包含科目名（科目名是关键词的一部分）
        for std_key, keywords in ACCOUNT_KEYWORDS.items():
            for kw in keywords:
                if clean in kw and len(clean) >= 3:
                    return std_key

        return None

    # ── 数据访问辅助 ──────────────────────────────────────────────

    def _get_value(
        self,
        data: Dict[str, List[float]],
        key: str,
        period_index: int = -1,
    ) -> Optional[float]:
        """从解析数据中提取指定科目的值。

        Parameters
        ----------
        data : Dict[str, List[float]]
            已解析的科目数据。
        key : str
            标准科目键。
        period_index : int
            期间索引，-1 表示最后一期（当前期）。

        Returns
        -------
        float or None
            找到的值，不存在返回 None。
        """
        vals = data.get(key)
        if vals is None:
            return None
        try:
            return vals[period_index]
        except IndexError:
            return None

    def _safe_div(
        self, numerator: Optional[float], denominator: Optional[float]
    ) -> Optional[float]:
        """安全除法，分母为0或任一值为None时返回None。"""
        if numerator is None or denominator is None:
            return None
        if denominator == 0:
            return None
        return numerator / denominator

    # ── 1. 比率分析 ──────────────────────────────────────────────

    def compute_ratios(self) -> Dict[str, Any]:
        """计算 20+ 项财务比率并进行行业对标。

        从已解析的 BS/IS 数据中自动提取所需科目，计算各财务比率，
        并与内置行业参考值对比。

        Returns
        -------
        Dict[str, Any]
            {
                "ratios": [{"指标": str, "公式": str, "本期值": float, "行业参考": str, "评估": str}, ...],
                "benchmark_source": str,
            }
        """
        bs = self.bs_data
        pl = self.is_data
        prior_bs = self.prior_bs_data
        prior_pl = self.prior_is_data

        # 提取关键科目（当前期 = 最后一期）
        revenue = self._get_value(pl, "revenue")
        cogs = self._get_value(pl, "cogs")
        net_profit = self._get_value(pl, "net_profit")
        operating_profit = self._get_value(pl, "operating_profit")
        total_profit = self._get_value(pl, "total_profit")
        interest_expense = self._get_value(pl, "interest_expense")
        finance_expenses = self._get_value(pl, "finance_expenses")

        total_assets = self._get_value(bs, "total_assets")
        current_assets = self._get_value(bs, "current_assets")
        current_liabilities = self._get_value(bs, "current_liabilities")
        total_liabilities = self._get_value(bs, "total_liabilities")
        total_equity = self._get_value(bs, "total_equity")
        inventory = self._get_value(bs, "inventory")
        accounts_receivable = self._get_value(bs, "accounts_receivable")

        # 计算平均值所需的期初/上期数据
        # 若无独立 prior_input，则用 BS 的倒数第二期（如果有）
        def _get_avg(val_curr, val_prior):
            """计算两期平均值。"""
            if val_curr is None:
                return None
            if val_prior is None:
                return val_curr  # 退化为期末
            return (val_curr + val_prior) / 2.0

        # 期初值：优先 prior_input，否则同表倒数第二列
        total_assets_prev = self._get_value(prior_bs, "total_assets")
        if total_assets_prev is None and total_assets is not None:
            ta_list = bs.get("total_assets", [total_assets])
            if len(ta_list) >= 2:
                total_assets_prev = ta_list[-2]
        avg_total_assets = _get_avg(total_assets, total_assets_prev)

        total_equity_prev = self._get_value(prior_bs, "total_equity")
        if total_equity_prev is None and total_equity is not None:
            te_list = bs.get("total_equity", [total_equity])
            if len(te_list) >= 2:
                total_equity_prev = te_list[-2]
        avg_equity = _get_avg(total_equity, total_equity_prev)

        ar_prev = self._get_value(prior_bs, "accounts_receivable")
        if ar_prev is None and accounts_receivable is not None:
            ar_list = bs.get("accounts_receivable", [accounts_receivable])
            if len(ar_list) >= 2:
                ar_prev = ar_list[-2]
        avg_ar = _get_avg(accounts_receivable, ar_prev)

        inv_prev = self._get_value(prior_bs, "inventory")
        if inv_prev is None and inventory is not None:
            inv_list = bs.get("inventory", [inventory])
            if len(inv_list) >= 2:
                inv_prev = inv_list[-2]
        avg_inventory = _get_avg(inventory, inv_prev)

        # 利息保障倍数: EBIT / 利息费用，若无利息费用则用财务费用近似
        interest = interest_expense if interest_expense else finance_expenses
        if interest is not None and interest > 0:
            interest = abs(interest)
        elif interest is not None and interest == 0:
            interest = None  # 避免除零

        # ── 比率定义 ──
        ratio_defs: List[Dict[str, Any]] = [
            # (中文名, 计算公式函数, 公式描述)
            (
                "毛利率",
                lambda: self._safe_div(
                    (revenue - cogs)
                    if (revenue is not None and cogs is not None)
                    else None,
                    revenue,
                ),
                "(营业收入 - 营业成本) / 营业收入 × 100%",
            ),
            (
                "净利率",
                lambda: self._safe_div(net_profit, revenue),
                "净利润 / 营业收入 × 100%",
            ),
            (
                "ROA（总资产收益率）",
                lambda: self._safe_div(net_profit, avg_total_assets),
                "净利润 / 平均总资产 × 100%",
            ),
            (
                "ROE（净资产收益率）",
                lambda: self._safe_div(net_profit, avg_equity),
                "净利润 / 平均净资产 × 100%",
            ),
            (
                "营业利润率",
                lambda: self._safe_div(operating_profit, revenue),
                "营业利润 / 营业收入 × 100%",
            ),
            (
                "息税前利润率",
                lambda: self._safe_div(self._calc_ebit(pl, bs), revenue),
                "EBIT / 营业收入 × 100%",
            ),
            (
                "流动比率",
                lambda: self._safe_div(current_assets, current_liabilities),
                "流动资产 / 流动负债",
            ),
            (
                "速动比率",
                lambda: self._safe_div(
                    (current_assets - inventory)
                    if (current_assets is not None and inventory is not None)
                    else None,
                    current_liabilities,
                ),
                "(流动资产 - 存货) / 流动负债",
            ),
            (
                "现金比率",
                lambda: self._safe_div(
                    self._get_value(bs, "cash_and_equivalents"), current_liabilities
                ),
                "货币资金 / 流动负债",
            ),
            (
                "资产负债率",
                lambda: self._safe_div(total_liabilities, total_assets),
                "总负债 / 总资产 × 100%",
            ),
            (
                "权益乘数",
                lambda: self._safe_div(total_assets, total_equity),
                "总资产 / 净资产",
            ),
            (
                "产权比率",
                lambda: self._safe_div(total_liabilities, total_equity),
                "总负债 / 净资产",
            ),
            (
                "利息保障倍数",
                lambda: self._safe_div(self._calc_ebit(pl, bs), interest),
                "EBIT / 利息费用",
            ),
            (
                "应收账款周转率",
                lambda: self._safe_div(revenue, avg_ar),
                "营业收入 / 平均应收账款",
            ),
            (
                "应收账款周转天数",
                lambda: self._safe_div(365, self._safe_div(revenue, avg_ar)),
                "365 / 应收账款周转率",
            ),
            (
                "存货周转率",
                lambda: self._safe_div(cogs, avg_inventory),
                "营业成本 / 平均存货",
            ),
            (
                "存货周转天数",
                lambda: self._safe_div(365, self._safe_div(cogs, avg_inventory)),
                "365 / 存货周转率",
            ),
            (
                "总资产周转率",
                lambda: self._safe_div(revenue, avg_total_assets),
                "营业收入 / 平均总资产",
            ),
            (
                "流动资产周转率",
                lambda: self._safe_div(
                    revenue,
                    _get_avg(
                        current_assets, self._get_value(prior_bs, "current_assets")
                    ),
                ),
                "营业收入 / 平均流动资产",
            ),
            # 发展能力（需上期数据）
            (
                "收入增长率",
                lambda: (
                    self._safe_div(
                        revenue - self._get_value(prior_pl, "revenue"),
                        abs(self._get_value(prior_pl, "revenue") or 1),
                    )
                    if self._get_value(prior_pl, "revenue") is not None
                    else None
                ),
                "(本期收入 - 上期收入) / |上期收入| × 100%",
            ),
            (
                "净利润增长率",
                lambda: (
                    self._safe_div(
                        (net_profit or 0)
                        - (self._get_value(prior_pl, "net_profit") or 0),
                        abs(self._get_value(prior_pl, "net_profit") or 1),
                    )
                    if self._get_value(prior_pl, "net_profit") is not None
                    else None
                ),
                "(本期净利润 - 上期净利润) / |上期净利润| × 100%",
            ),
            (
                "总资产增长率",
                lambda: (
                    self._safe_div(
                        (total_assets or 0) - (total_assets_prev or 0),
                        abs(total_assets_prev or 1),
                    )
                    if total_assets_prev is not None
                    else None
                ),
                "(期末总资产 - 期初总资产) / |期初总资产| × 100%",
            ),
        ]

        # ── 计算 & 评估 ──
        benchmarks = INDUSTRY_BENCHMARKS.get(self.industry, INDUSTRY_BENCHMARKS["通用"])
        ratios = []

        for name, calc_fn, formula in ratio_defs:
            try:
                value = calc_fn()
            except (TypeError, ZeroDivisionError):
                value = None

            # 行业对标
            bm = benchmarks.get(name, benchmarks.get(name.split("（")[0], None))
            if bm is not None and value is not None:
                # 判断是否在合理区间
                if bm["low"] <= value <= bm["high"]:
                    assessment = "正常"
                    assessment_color = "green"
                elif value < bm["low"]:
                    assessment = f"偏低(<{bm['low']:.1%})"
                    assessment_color = "yellow"
                else:
                    assessment = f"偏高(>{bm['high']:.1%})"
                    assessment_color = "yellow"
                benchmark_str = f"{bm['low']:.1%} ~ {bm['high']:.1%}"
            else:
                assessment = "无参考"
                assessment_color = "gray"
                benchmark_str = "N/A"

            # 格式化值
            if value is not None:
                if "天数" in name:
                    display = f"{value:.1f}天"
                elif "倍数" in name or "乘数" in name or "比率" in name:
                    display = f"{value:.2f}"
                else:
                    display = f"{value:.2%}"
            else:
                display = "N/A"

            ratios.append(
                {
                    "指标": name,
                    "公式": formula,
                    "本期值": value,
                    "显示值": display,
                    "行业参考": benchmark_str,
                    "评估": assessment,
                    "评估颜色": assessment_color,
                }
            )

            # 收集异常
            if assessment_color == "yellow":
                self._anomalies.append(
                    {
                        "来源": "比率分析",
                        "指标": name,
                        "本期值": value,
                        "描述": f"{name} 偏离行业参考区间({benchmark_str})",
                        "严重程度": "警告",
                    }
                )

        self.results["ratios"] = {
            "ratios": ratios,
            "benchmark_source": f"{self.industry}行业参考值",
            "count": len(ratios),
        }
        return self.results["ratios"]

    def _calc_ebit(
        self, pl: Dict[str, List[float]], bs: Dict[str, List[float]]
    ) -> Optional[float]:
        """近似计算 EBIT（息税前利润）。

        优先使用 total_profit + interest_expense，
        其次用 operating_profit + investment_income + other_income。
        """
        total_profit = self._get_value(pl, "total_profit")
        interest_expense = self._get_value(pl, "interest_expense")
        finance_expenses = self._get_value(pl, "finance_expenses")
        operating_profit = self._get_value(pl, "operating_profit")

        if total_profit is not None:
            ie = interest_expense if interest_expense else finance_expenses
            if ie is not None and ie > 0:
                return total_profit + ie
            return total_profit
        if operating_profit is not None:
            return operating_profit
        return None

    # ── 2. 趋势分析 ──────────────────────────────────────────────

    def trend_analysis(self) -> Dict[str, Any]:
        """多期趋势分析。

        对每个科目计算期间变化率和同比变化率，
        标记超过波动阈值的异常变动。

        Returns
        -------
        Dict[str, Any]
            {
                "trends": {科目: {期间: {值, 变化率%, 环比%, 异常标记}}, ...},
                "periods": [期间标签...],
            }
        """
        # 合并 BS + IS 中所有有意义的科目
        all_accounts = set(self.bs_data.keys()) | set(self.is_data.keys())
        # 过滤：只保留合计/关键科目，排除过度明细
        key_accounts = {
            k
            for k in all_accounts
            if any(
                suffix in k
                for suffix in [
                    "revenue",
                    "cogs",
                    "net_profit",
                    "total_profit",
                    "operating_profit",
                    "total_assets",
                    "total_liabilities",
                    "total_equity",
                    "current_assets",
                    "current_liabilities",
                    "accounts_receivable",
                    "inventory",
                    "cash",
                    "selling_expenses",
                    "admin_expenses",
                    "finance_expenses",
                    "gross_profit",
                    "fixed_assets",
                ]
            )
        }

        periods = self.periods if self.periods else ["期间1"]
        trends: Dict[str, Dict[str, Any]] = {}

        for std_key in sorted(key_accounts):
            # 从 BS 或 IS 取数据
            values = self.bs_data.get(std_key) or self.is_data.get(std_key)
            if not values:
                continue

            # 补齐期间标签
            n_periods = min(len(values), len(periods))
            account_trend: Dict[str, Any] = {}

            for i in range(n_periods):
                p_label = periods[i]
                val = values[i]
                change_pct = None
                is_anomaly = False

                if i > 0 and values[i - 1] != 0:
                    # 环比变化率
                    change_pct = (val - values[i - 1]) / abs(values[i - 1]) * 100
                    if abs(change_pct) > self.fluctuation_threshold:
                        is_anomaly = True

                account_trend[p_label] = {
                    "值": val,
                    "环比变化率%": round(change_pct, 2)
                    if change_pct is not None
                    else None,
                    "异常标记": "⚠" if is_anomaly else "",
                }

                if is_anomaly:
                    cn_name = self._std_key_to_cn(std_key)
                    self._anomalies.append(
                        {
                            "来源": "趋势分析",
                            "指标": f"{cn_name}({p_label})",
                            "本期值": val,
                            "描述": (
                                f"{cn_name} 在 {p_label} 环比变动 {change_pct:.1f}%，"
                                f"超过波动阈值 {self.fluctuation_threshold}%"
                            ),
                            "严重程度": "警告",
                        }
                    )

            trends[std_key] = account_trend

        self.results["trends"] = {
            "trends": trends,
            "periods": periods[
                : max(
                    len(v)
                    for v in (list(self.bs_data.values()) + list(self.is_data.values()))
                    if v
                )
                or 1
            ],
            "threshold": self.fluctuation_threshold,
        }
        return self.results["trends"]

    def _std_key_to_cn(self, std_key: str) -> str:
        """标准键 → 中文名称（用于报告展示）。"""
        name_map = {
            "revenue": "营业收入",
            "cogs": "营业成本",
            "net_profit": "净利润",
            "total_profit": "利润总额",
            "operating_profit": "营业利润",
            "total_assets": "总资产",
            "total_liabilities": "总负债",
            "total_equity": "净资产",
            "current_assets": "流动资产",
            "current_liabilities": "流动负债",
            "accounts_receivable": "应收账款",
            "inventory": "存货",
            "cash_and_equivalents": "货币资金",
            "fixed_assets": "固定资产",
            "selling_expenses": "销售费用",
            "admin_expenses": "管理费用",
            "finance_expenses": "财务费用",
            "gross_profit": "毛利",
            "interest_expense": "利息费用",
            "investment_income": "投资收益",
            "income_tax": "所得税",
        }
        return name_map.get(std_key, std_key)

    # ── 3. Benford 定律检验 ──────────────────────────────────────

    def benford_test(
        self,
        target_data: Optional[List[float]] = None,
        sheet_type: str = "IS",
    ) -> Dict[str, Any]:
        """Benford 定律首位数字分布检验。

        对给定的数值列表，统计首位数字 (1-9) 的实际分布，
        与 Benford 理论分布对比，进行卡方拟合优度检验，
        并计算各首位数字的 Z-score。

        适用场景（CAS 1313 指引）：
          - 收入明细（自然生成的交易金额）
          - 费用明细
          - 应收账款/应付账款明细
          - 不适用于：人为设定数字（如折旧率）、有上下限的数据

        Parameters
        ----------
        target_data : List[float], optional
            待检验的数值序列。若不提供，则自动从 IS 数据中提取所有明细行。
        sheet_type : str
            自动提取数据的数据源，"BS" 或 "IS"，默认 "IS"。

        Returns
        -------
        Dict[str, Any]
            {
                "distribution": [{首位: int, 实际次数: int, 实际比例: float,
                                  理论比例: float, 期望次数: float, Z_score: float}, ...],
                "chi_square": float,
                "chi_critical_005": float,
                "is_conforming": bool,
                "total_count": int,
            }
        """
        # ── 数据提取 ──
        if target_data is None:
            source = self.is_data if sheet_type == "IS" else self.bs_data
            # 取所有明细值（排除合计行，但这里已经是标准键层级）
            target_data = []
            for key, vals in source.items():
                if vals:
                    target_data.extend(abs(v) for v in vals if v != 0 and abs(v) >= 1)
        else:
            target_data = [abs(v) for v in target_data if v != 0 and abs(v) >= 1]

        n = len(target_data)
        if n < 30:
            # 样本量不足，Benford 检验无统计意义
            self.results["benford"] = {
                "distribution": [],
                "chi_square": None,
                "chi_critical_005": CHI2_CRITICAL_005,
                "is_conforming": None,
                "total_count": n,
                "warning": f"样本量({n})不足，Benford检验需≥30个数据点",
            }
            return self.results["benford"]

        # ── 首位数字提取 ──
        def first_digit(x: float) -> int:
            """提取绝对值的第一位有效数字。"""
            x = abs(x)
            if x < 1:
                # 小数：乘10直到 >= 1
                while x < 1:
                    x *= 10
            elif x >= 10:
                while x >= 10:
                    x /= 10
            return int(x)

        # ── 统计实际分布 ──
        observed: Dict[int, int] = {d: 0 for d in range(1, 10)}
        for val in target_data:
            d = first_digit(val)
            if 1 <= d <= 9:
                observed[d] += 1

        # ── 卡方统计量 ──
        chi_square = 0.0
        distribution = []
        all_conforming = True

        for d in range(1, 10):
            o = observed[d]
            e_prob = BENFORD_EXPECTED[d]
            e_count = n * e_prob

            # Z-score: (O - E) / sqrt(n * p * (1-p))
            if n * e_prob * (1 - e_prob) > 0:
                z_score = (o - e_count) / math.sqrt(n * e_prob * (1 - e_prob))
            else:
                z_score = 0.0

            # 卡方分量
            if e_count > 0:
                chi_square += (o - e_count) ** 2 / e_count

            # Z-score > 2 或 < -2 视为显著偏离
            is_deviant = abs(z_score) > 2.0
            if is_deviant:
                all_conforming = False

            distribution.append(
                {
                    "首位数字": d,
                    "实际次数": o,
                    "实际比例": round(o / n, 4),
                    "理论比例": round(e_prob, 4),
                    "期望次数": round(e_count, 2),
                    "差值": round(o - e_count, 2),
                    "Z_score": round(z_score, 2),
                    "偏离标记": "⚠" if is_deviant else "",
                }
            )

        is_conforming = chi_square <= CHI2_CRITICAL_005

        # ── 收集异常 ──
        if not is_conforming:
            self._anomalies.append(
                {
                    "来源": "Benford检验",
                    "指标": f"整体分布 (n={n})",
                    "本期值": chi_square,
                    "描述": (
                        f"卡方统计量 {chi_square:.2f} > 临界值 {CHI2_CRITICAL_005:.2f}，"
                        f"数据不符合 Benford 分布，可能存在人工干预"
                    ),
                    "严重程度": "重大" if chi_square > CHI2_CRITICAL_001 else "警告",
                }
            )

        for item in distribution:
            if item["偏离标记"]:
                self._anomalies.append(
                    {
                        "来源": "Benford检验",
                        "指标": f"首位数字={item['首位数字']}",
                        "本期值": item["实际比例"],
                        "描述": (
                            f"首位数字 {item['首位数字']} 实际比例 {item['实际比例']:.2%} "
                            f"偏离理论 {item['理论比例']:.2%}，Z={item['Z_score']:.2f}"
                        ),
                        "严重程度": "重大" if abs(item["Z_score"]) > 3 else "警告",
                    }
                )

        self.results["benford"] = {
            "distribution": distribution,
            "chi_square": round(chi_square, 4),
            "chi_critical_005": CHI2_CRITICAL_005,
            "chi_critical_001": CHI2_CRITICAL_001,
            "is_conforming": is_conforming,
            "total_count": n,
        }
        return self.results["benford"]

    # ── 4. 期间对比 ──────────────────────────────────────────────

    def period_comparison(self) -> Dict[str, Any]:
        """本期 vs 上期对比分析。

        对 BS 和 IS 的主要科目逐项对比，计算变动额和变动率，
        超过重要性水平(materiality)或波动阈值的标记。

        Returns
        -------
        Dict[str, Any]
            {
                "comparisons": [{"科目": str, "本期值": float, "上期值": float,
                                 "变动额": float, "变动率%": float, "标记": str}, ...],
            }
        """
        if not self.prior_bs_data and not self.prior_is_data:
            # 无上期数据时尝试用同文件多期数据
            have_prior = False
            for vals in self.bs_data.values():
                if len(vals) >= 2:
                    have_prior = True
                    break
            for vals in self.is_data.values():
                if len(vals) >= 2:
                    have_prior = True
                    break
            if not have_prior:
                self.results["comparison"] = {
                    "comparisons": [],
                    "note": "无上期数据可用，请提供 --prior-input 或确保输入文件含多期数据",
                }
                return self.results["comparison"]

        comparisons = []
        # 合并所有可对比的科目
        all_keys = set(self.bs_data.keys()) | set(self.is_data.keys())
        # 优先显示重要科目
        priority_keys = [
            "revenue",
            "cogs",
            "net_profit",
            "total_profit",
            "operating_profit",
            "total_assets",
            "total_liabilities",
            "total_equity",
            "current_assets",
            "current_liabilities",
            "accounts_receivable",
            "inventory",
            "cash_and_equivalents",
            "selling_expenses",
            "admin_expenses",
            "finance_expenses",
            "interest_expense",
            "income_tax",
            "gross_profit",
            "fixed_assets",
        ]
        ordered_keys = [k for k in priority_keys if k in all_keys]
        ordered_keys += sorted(all_keys - set(priority_keys))

        for std_key in ordered_keys:
            curr_vals = self.bs_data.get(std_key) or self.is_data.get(std_key)
            if not curr_vals:
                continue

            curr = curr_vals[-1]  # 最后一期 = 本期

            # 上期值：优先 prior_input，其次同表倒数第二期
            prior = None
            if std_key in self.prior_bs_data:
                pv = self.prior_bs_data[std_key]
                prior = pv[-1] if pv else None
            elif std_key in self.prior_is_data:
                pv = self.prior_is_data[std_key]
                prior = pv[-1] if pv else None
            elif len(curr_vals) >= 2:
                prior = curr_vals[-2]

            if prior is None:
                continue

            variance = curr - prior
            if prior != 0:
                variance_pct = (variance / abs(prior)) * 100
            else:
                variance_pct = None

            # ── 标记逻辑 ──
            flag = ""
            if self.materiality > 0 and abs(variance) > self.materiality:
                flag = "重大"
            elif (
                variance_pct is not None
                and abs(variance_pct) > self.fluctuation_threshold
            ):
                flag = "异常"

            cn_name = self._std_key_to_cn(std_key)
            comparisons.append(
                {
                    "科目": cn_name,
                    "标准键": std_key,
                    "本期值": curr,
                    "上期值": prior,
                    "变动额": variance,
                    "变动率%": round(variance_pct, 2)
                    if variance_pct is not None
                    else None,
                    "标记": flag,
                }
            )

            if flag:
                self._anomalies.append(
                    {
                        "来源": "期间对比",
                        "指标": cn_name,
                        "本期值": curr,
                        "描述": (
                            f"{cn_name} 本期({curr:,.2f}) vs 上期({prior:,.2f})，"
                            f"变动 {variance:,.2f}"
                            + (
                                f" ({variance_pct:.1f}%)"
                                if variance_pct is not None
                                else ""
                            )
                            + (
                                f"，超过重要性水平 {self.materiality:,.2f}"
                                if flag == "重大"
                                else f"，超过波动阈值 {self.fluctuation_threshold}%"
                            )
                        ),
                        "严重程度": "重大" if flag == "重大" else "警告",
                    }
                )

        self.results["comparison"] = {"comparisons": comparisons}
        return self.results["comparison"]

    # ── 执行全部分析 ──────────────────────────────────────────────

    def run_all(self) -> Dict[str, Any]:
        """执行全部四项分析程序。

        Returns
        -------
        Dict[str, Any]
            包含 ratios, trends, benford, comparison 四个模块的完整结果。
        """
        self.parse()
        self._anomalies = []

        self.compute_ratios()
        self.trend_analysis()
        self.benford_test()
        self.period_comparison()

        # 汇总异常
        self.results["summary"] = {
            "total_anomalies": len(self._anomalies),
            "critical_count": sum(
                1 for a in self._anomalies if a["严重程度"] == "重大"
            ),
            "warning_count": sum(1 for a in self._anomalies if a["严重程度"] == "警告"),
            "anomalies": self._anomalies,
            "execution_time": datetime.now().isoformat(),
            "input_file": self.input_path,
            "materiality": self.materiality,
            "fluctuation_threshold": self.fluctuation_threshold,
            "industry": self.industry,
        }

        return self.results

    # ── 输出生成 ──────────────────────────────────────────────────

    def generate_output(
        self,
        output_path: str,
        results: Optional[Dict[str, Any]] = None,
        generate_charts: bool = False,
    ) -> str:
        """生成格式化的分析报告 Excel。

        Parameters
        ----------
        output_path : str
            输出 Excel 文件路径。
        results : Dict, optional
            分析结果（默认使用 self.results）。
        generate_charts : bool
            是否生成图表（使用 openpyxl 内置 LineChart/BarChart）。

        Returns
        -------
        str
            输出文件路径。
        """
        if results is None:
            results = self.results

        wb = openpyxl.Workbook()
        # 删除默认 sheet
        wb.remove(wb.active)

        # ── Sheet 1: 比率分析 ──
        self._write_ratios_sheet(wb, results.get("ratios", {}))

        # ── Sheet 2: 趋势分析 ──
        self._write_trend_sheet(wb, results.get("trends", {}), generate_charts)

        # ── Sheet 3: Benford 检验 ──
        self._write_benford_sheet(wb, results.get("benford", {}), generate_charts)

        # ── Sheet 4: 期间对比 ──
        self._write_comparison_sheet(wb, results.get("comparison", {}))

        # ── Sheet 5: 分析总结 ──
        self._write_summary_sheet(wb, results.get("summary", {}))

        # 保存
        wb.save(output_path)
        wb.close()
        return output_path

    def _apply_header_style(self, ws, row: int, max_col: int):
        """为表头行应用统一样式。"""
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

    def _apply_data_border(self, ws, start_row: int, end_row: int, max_col: int):
        """为数据区域应用边框。"""
        for r in range(start_row, end_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).border = THIN_BORDER

    def _auto_width(self, ws, min_width: int = 8, max_width: int = 40):
        """自动调整列宽。"""
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    # 中文字符约占2个英文字符宽度
                    val = str(cell.value)
                    char_len = 0
                    for ch in val:
                        if "\u4e00" <= ch <= "\u9fff" or "\u3000" <= ch <= "\u303f":
                            char_len += 2
                        else:
                            char_len += 1
                    max_len = max(max_len, char_len)
            adjusted = min(max(max_len + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted

    def _write_ratios_sheet(self, wb, ratios_data: Dict):
        """写入比率分析 sheet。"""
        ws = wb.create_sheet("比率分析")

        # 标题
        ws.merge_cells("A1:F1")
        ws["A1"] = f"财务比率分析 — {ratios_data.get('benchmark_source', '')}"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")

        # 表头
        headers = ["序号", "财务指标", "计算公式", "本期值", "行业参考区间", "评估"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=3, column=c, value=h)
        self._apply_header_style(ws, 3, len(headers))

        # 数据
        ratios = ratios_data.get("ratios", [])
        for i, r in enumerate(ratios):
            row = 4 + i
            ws.cell(row=row, column=1, value=i + 1).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=r["指标"]).font = NORMAL_FONT
            ws.cell(row=row, column=3, value=r["公式"]).font = Font(
                name="微软雅黑", size=9, color="666666"
            )
            ws.cell(row=row, column=4, value=r["显示值"]).font = NUMBER_FONT
            ws.cell(row=row, column=5, value=r["行业参考"]).font = NORMAL_FONT

            # 评估列 + 着色
            assess_cell = ws.cell(row=row, column=6, value=r["评估"])
            assess_cell.font = Font(name="微软雅黑", size=10, bold=True)
            color = r.get("评估颜色", "gray")
            if color == "yellow":
                assess_cell.fill = WARNING_FILL
            elif color == "green":
                assess_cell.fill = PASS_FILL

        data_end = 3 + len(ratios)
        self._apply_data_border(ws, 3, data_end, len(headers))
        self._auto_width(ws)

    def _write_trend_sheet(self, wb, trends_data: Dict, generate_charts: bool):
        """写入趋势分析 sheet。"""
        ws = wb.create_sheet("趋势分析")

        trends = trends_data.get("trends", {})
        periods = trends_data.get("periods", ["期间1"])

        if not trends:
            ws["A1"] = "趋势分析 — 无数据"
            ws["A1"].font = TITLE_FONT
            return

        # ── 布局：每个科目一个小节 ──
        current_row = 1
        chart_data_start = {}  # {科目: (start_row, end_row, periods_count)}

        for std_key, period_data in trends.items():
            cn_name = self._std_key_to_cn(std_key)

            # 节标题
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=len(periods) + 3,
            )
            ws.cell(row=current_row, column=1, value=f"▌ {cn_name} ({std_key})")
            ws.cell(row=current_row, column=1).font = SECTION_FONT
            current_row += 1

            # 表头
            sub_headers = ["期间", "值", "环比变化率(%)", "异常标记"]
            for c, h in enumerate(sub_headers, 1):
                ws.cell(row=current_row, column=c, value=h)
            self._apply_header_style(ws, current_row, len(sub_headers))
            header_row = current_row
            current_row += 1

            # 数据行
            data_start_row = current_row
            for p_label in periods:
                info = period_data.get(p_label)
                if info is None:
                    current_row += 1
                    continue
                ws.cell(row=current_row, column=1, value=p_label).font = NORMAL_FONT
                ws.cell(row=current_row, column=2, value=info["值"]).font = NUMBER_FONT
                ws.cell(row=current_row, column=2).number_format = "#,##0.00"
                change = info.get("环比变化率%")
                if change is not None:
                    ws.cell(row=current_row, column=3, value=change).font = NUMBER_FONT
                    ws.cell(row=current_row, column=3).number_format = '0.00"%"'
                    # 标记异常单元格
                    if abs(change) > trends_data.get("threshold", 30):
                        ws.cell(row=current_row, column=3).fill = WARNING_FILL
                ws.cell(
                    row=current_row, column=4, value=info.get("异常标记", "")
                ).font = NORMAL_FONT
                current_row += 1

            data_end_row = current_row - 1
            self._apply_data_border(ws, header_row, data_end_row, len(sub_headers))
            chart_data_start[std_key] = (data_start_row, data_end_row, len(periods))
            current_row += 1  # 空行分隔

        self._auto_width(ws)

        # ── 图表生成 ──
        if generate_charts and periods:
            chart_sheet = wb.create_sheet("趋势图表")
            chart_row = 1

            for std_key, (ds, de, n_p) in chart_data_start.items():
                if de - ds < 1:
                    continue
                cn_name = self._std_key_to_cn(std_key)

                chart = LineChart()
                chart.title = f"{cn_name} 趋势"
                chart.style = 10
                chart.y_axis.title = "金额"
                chart.x_axis.title = "期间"
                chart.width = 22
                chart.height = 13

                # 数据引用
                values_ref = Reference(ws, min_col=2, min_row=ds, max_row=de)
                cats_ref = Reference(ws, min_col=1, min_row=ds, max_row=de)

                chart.add_data(values_ref, titles_from_data=False)
                chart.set_categories(cats_ref)
                chart.series[0].title = openpyxl.chart.series.SeriesLabel(v=cn_name)

                # 数据标签
                chart.series[0].dLbls = DataLabelList()
                chart.series[0].dLbls.showVal = True

                chart_sheet.add_chart(chart, f"A{chart_row}")
                chart_row += 18

    def _write_benford_sheet(self, wb, benford_data: Dict, generate_charts: bool):
        """写入 Benford 检验 sheet。"""
        ws = wb.create_sheet("Benford检验")

        # 标题
        ws.merge_cells("A1:H1")
        ws["A1"] = "Benford 定律 — 首位数字分布检验"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")

        # 检验概要
        ws.merge_cells("A3:H3")
        chi2 = benford_data.get("chi_square")
        is_conf = benford_data.get("is_conforming")
        n = benford_data.get("total_count", 0)
        ws["A3"] = (
            f"样本量: {n}  |  卡方统计量 χ² = {chi2:.2f}"
            if chi2 is not None
            else "卡方统计量: N/A"
            + f"  |  临界值(α=0.05) = {CHI2_CRITICAL_005:.2f}"
            + f"  |  结论: {'✅ 符合 Benford 分布' if is_conf else '❌ 偏离 Benford 分布' if is_conf is False else '⚠ 样本量不足'}"
        )
        ws["A3"].font = Font(name="微软雅黑", size=10, bold=True)
        if is_conf is True:
            ws["A3"].fill = PASS_FILL
        elif is_conf is False:
            ws["A3"].fill = CRITICAL_FILL

        # 表头
        headers = [
            "首位数字",
            "实际次数",
            "实际比例",
            "理论比例",
            "期望次数",
            "差值",
            "Z-Score",
            "偏离标记",
        ]
        for c, h in enumerate(headers, 1):
            ws.cell(row=5, column=c, value=h)
        self._apply_header_style(ws, 5, len(headers))

        # 数据
        distribution = benford_data.get("distribution", [])
        for i, d in enumerate(distribution):
            row = 6 + i
            ws.cell(row=row, column=1, value=d["首位数字"]).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=d["实际次数"]).font = NUMBER_FONT
            ws.cell(row=row, column=3, value=d["实际比例"]).font = NUMBER_FONT
            ws.cell(row=row, column=3).number_format = "0.00%"
            ws.cell(row=row, column=4, value=d["理论比例"]).font = NUMBER_FONT
            ws.cell(row=row, column=4).number_format = "0.00%"
            ws.cell(row=row, column=5, value=d["期望次数"]).font = NUMBER_FONT
            ws.cell(row=row, column=6, value=d["差值"]).font = NUMBER_FONT
            ws.cell(row=row, column=7, value=d["Z_score"]).font = NUMBER_FONT

            # 标记异常 Z-score
            mark = d.get("偏离标记", "")
            ws.cell(row=row, column=8, value=mark).font = NORMAL_FONT
            if mark:
                z = d["Z_score"]
                if abs(z) > 3:
                    ws.cell(row=row, column=7).fill = CRITICAL_FILL
                else:
                    ws.cell(row=row, column=7).fill = WARNING_FILL

        data_end = 5 + len(distribution)
        self._apply_data_border(ws, 5, data_end, len(headers))
        self._auto_width(ws)

        # ── Benford 柱状图 ──
        if generate_charts and distribution:
            chart_sheet = wb["Benford检验"]
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.title = "Benford 定律 — 首位数字分布 vs 理论分布"
            chart.style = 10
            chart.width = 22
            chart.height = 13

            # 实际分布
            actual_ref = Reference(ws, min_col=2, min_row=5, max_row=data_end)
            cats_ref = Reference(ws, min_col=1, min_row=6, max_row=data_end)
            chart.add_data(actual_ref, titles_from_data=True)
            chart.set_categories(cats_ref)

            # 理论分布
            expected_ref = Reference(ws, min_col=5, min_row=5, max_row=data_end)
            chart.add_data(expected_ref, titles_from_data=True)

            chart.series[0].title = openpyxl.chart.series.SeriesLabel(v="实际")
            chart.series[1].title = openpyxl.chart.series.SeriesLabel(v="理论")

            chart_row = data_end + 3
            chart_sheet.add_chart(chart, f"A{chart_row}")

    def _write_comparison_sheet(self, wb, comparison_data: Dict):
        """写入期间对比 sheet。"""
        ws = wb.create_sheet("期间对比")

        ws.merge_cells("A1:G1")
        ws["A1"] = "期间对比分析 — 本期 vs 上期"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")

        comparisons = comparison_data.get("comparisons", [])
        if not comparisons:
            ws["A3"] = comparison_data.get("note", "无对比数据")
            ws["A3"].font = NORMAL_FONT
            return

        # 参数说明
        ws.merge_cells("A3:G3")
        params = []
        if self.materiality > 0:
            params.append(f"重要性水平: {self.materiality:,.2f}")
        params.append(f"波动阈值: {self.fluctuation_threshold}%")
        ws["A3"] = "  |  ".join(params)
        ws["A3"].font = Font(name="微软雅黑", size=9, color="666666")

        # 表头
        headers = ["序号", "科目", "本期值", "上期值", "变动额", "变动率(%)", "标记"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=5, column=c, value=h)
        self._apply_header_style(ws, 5, len(headers))

        # 数据
        for i, comp in enumerate(comparisons):
            row = 6 + i
            ws.cell(row=row, column=1, value=i + 1).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=comp["科目"]).font = NORMAL_FONT
            for ci, key in enumerate(["本期值", "上期值", "变动额"], 3):
                val = comp[key]
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font = NUMBER_FONT
                cell.number_format = "#,##0.00"

            var_pct = comp.get("变动率%")
            if var_pct is not None:
                ws.cell(row=row, column=6, value=var_pct).font = NUMBER_FONT
                ws.cell(row=row, column=6).number_format = '0.00"%"'

            flag = comp.get("标记", "")
            flag_cell = ws.cell(row=row, column=7, value=flag)
            flag_cell.font = Font(name="微软雅黑", size=10, bold=True)
            if flag == "重大":
                flag_cell.fill = CRITICAL_FILL
                # 整行轻微标红
                for c in range(1, 8):
                    if ws.cell(row=row, column=c).fill == PatternFill():
                        ws.cell(row=row, column=c).fill = PatternFill(
                            start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"
                        )
            elif flag == "异常":
                flag_cell.fill = WARNING_FILL

        data_end = 5 + len(comparisons)
        self._apply_data_border(ws, 5, data_end, len(headers))
        self._auto_width(ws)

    def _write_summary_sheet(self, wb, summary_data: Dict):
        """写入分析总结 sheet。"""
        ws = wb.create_sheet("分析总结")

        # 标题
        ws.merge_cells("A1:G1")
        ws["A1"] = "实质性分析程序 — 异常汇总报告"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")

        # 执行信息
        ws.merge_cells("A3:G3")
        ws["A3"] = (
            f"输入文件: {summary_data.get('input_file', 'N/A')}  |  "
            f"行业: {summary_data.get('industry', 'N/A')}  |  "
            f"执行时间: {summary_data.get('execution_time', 'N/A')}"
        )
        ws["A3"].font = Font(name="微软雅黑", size=9, color="666666")

        # 统计概要
        total = summary_data.get("total_anomalies", 0)
        critical = summary_data.get("critical_count", 0)
        warning = summary_data.get("warning_count", 0)

        ws["A5"] = "异常统计"
        ws["A5"].font = SECTION_FONT
        ws.merge_cells("A6:B6")
        ws["A6"] = f"异常总数: {total}"
        ws["A6"].font = Font(name="微软雅黑", size=11, bold=True)
        ws.merge_cells("A7:B7")
        ws["A7"] = f"重大异常: {critical}"
        ws["A7"].font = Font(name="微软雅黑", size=11, bold=True, color="FF0000")
        ws.merge_cells("A8:B8")
        ws["A8"] = f"警告异常: {warning}"
        ws["A8"].font = Font(name="微软雅黑", size=11, bold=True, color="FF8C00")

        # 异常明细表头
        headers = ["序号", "来源", "指标", "本期值", "描述", "严重程度"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=10, column=c, value=h)
        self._apply_header_style(ws, 10, len(headers))

        anomalies = summary_data.get("anomalies", [])
        # 排序：重大在前
        anomalies_sorted = sorted(
            anomalies, key=lambda a: 0 if a["严重程度"] == "重大" else 1
        )

        for i, a in enumerate(anomalies_sorted):
            row = 11 + i
            ws.cell(row=row, column=1, value=i + 1).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=a["来源"]).font = NORMAL_FONT
            ws.cell(row=row, column=3, value=a["指标"]).font = NORMAL_FONT
            ws.cell(row=row, column=4, value=a.get("本期值", "N/A")).font = NUMBER_FONT
            ws.cell(row=row, column=5, value=a["描述"]).font = Font(
                name="微软雅黑", size=9
            )
            ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)

            severity_cell = ws.cell(row=row, column=6, value=a["严重程度"])
            severity_cell.font = Font(name="微软雅黑", size=10, bold=True)
            if a["严重程度"] == "重大":
                severity_cell.fill = CRITICAL_FILL
            else:
                severity_cell.fill = WARNING_FILL

        data_end = 10 + len(anomalies)
        if anomalies:
            self._apply_data_border(ws, 10, data_end, len(headers))
        self._auto_width(ws)
        # 描述列更宽
        ws.column_dimensions["E"].width = 60


# ══════════════════════════════════════════════════════════════════════
# CLI 入口
# ══════════════════════════════════════════════════════════════════════


def main():
    """命令行入口。

    Examples
    --------
    # 执行全部分析
    python analytics_engine.py --input 财务报表.xlsx --output 分析报告.xlsx

    # 仅比率分析 + 指定行业
    python analytics_engine.py --input 报表.xlsx --method ratios --industry 电力

    # 包含上期数据 + 生成图表
    python analytics_engine.py --input 本期.xlsx --prior-input 上期.xlsx --chart

    # 自定义阈值和重要性水平
    python analytics_engine.py --input 报表.xlsx --materiality 100000 --fluctuation-threshold 20
    """
    parser = argparse.ArgumentParser(
        description="审计实质性分析程序引擎 — 符合 CAS 1313 (ISA 520)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
分析程序说明:
  ratios      比率分析 — 20+ 财务指标 + 行业对标
  trend       趋势分析 — 多期趋势 + 异常波动标记
  benford     Benford 定律 — 首位数字分布检验
  comparison  期间对比 — 本期 vs 上期重大波动
  all         全部四项分析 (默认)

行业选项:
  内置参考值: 电力, 能源, 通用

示例:
  python analytics_engine.py -i 财务报表.xlsx -o 分析报告.xlsx
  python analytics_engine.py -i 报表.xlsx -m ratios --industry 电力
  python analytics_engine.py -i 本期.xlsx -p 上期.xlsx --chart
        """,
    )

    parser.add_argument(
        "-i",
        "--input",
        required=True,
        help="输入 Excel 文件路径（含 BS 和 IS 工作表）",
    )
    parser.add_argument(
        "--sheet-bs",
        default="BS",
        help="资产负债表工作表名称 (默认: BS)",
    )
    parser.add_argument(
        "--sheet-is",
        default="IS",
        help="利润表工作表名称 (默认: IS)",
    )
    parser.add_argument(
        "-p",
        "--prior-input",
        default=None,
        help="上期数据 Excel 文件路径（期间对比用）",
    )
    parser.add_argument(
        "-m",
        "--method",
        default="all",
        choices=["all", "ratios", "trend", "benford", "comparison"],
        help="分析方法 (默认: all)",
    )
    parser.add_argument(
        "--materiality",
        type=float,
        default=0.0,
        help="重要性水平（绝对金额），超过此值标记为重大异常 (默认: 0)",
    )
    parser.add_argument(
        "--fluctuation-threshold",
        type=float,
        default=30.0,
        help="波动阈值百分比，超过此百分比标记为异常 (默认: 30)",
    )
    parser.add_argument(
        "--industry",
        default="电力",
        choices=["电力", "能源", "通用"],
        help="行业参考值 (默认: 电力)",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="输出 Excel 文件路径 (默认: 输入文件名_分析报告_时间戳.xlsx)",
    )
    parser.add_argument(
        "--chart",
        action="store_true",
        default=False,
        help="生成趋势图表和 Benford 分布图",
    )

    args = parser.parse_args()

    # 验证输入文件
    if not os.path.exists(args.input):
        print(f"[错误] 输入文件不存在: {args.input}", file=sys.stderr)
        sys.exit(1)

    # 自动生成输出路径
    if args.output is None:
        base = os.path.splitext(os.path.basename(args.input))[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = os.path.join(
            os.path.dirname(args.input) or ".",
            f"{base}_分析报告_{timestamp}.xlsx",
        )

    print(f"══════════════════════════════════════════════")
    print(f"  审计实质性分析程序引擎 v1.0")
    print(f"  符合 CAS 1313 / ISA 520")
    print(f"══════════════════════════════════════════════")
    print(f"  输入文件: {args.input}")
    print(f"  BS 工作表: {args.sheet_bs}")
    print(f"  IS 工作表: {args.sheet_is}")
    print(f"  分析方法: {args.method}")
    print(f"  行业参考: {args.industry}")
    print(f"  波动阈值: {args.fluctuation_threshold}%")
    print(f"  重要性水平: {args.materiality:,.2f}")
    if args.prior_input:
        print(f"  上期数据: {args.prior_input}")
    print(f"  生成图表: {'是' if args.chart else '否'}")
    print(f"══════════════════════════════════════════════")

    # 初始化引擎
    engine = AnalyticsEngine(
        input_path=args.input,
        sheet_bs=args.sheet_bs,
        sheet_is=args.sheet_is,
        prior_input=args.prior_input,
        materiality=args.materiality,
        fluctuation_threshold=args.fluctuation_threshold,
        industry=args.industry,
    )

    # 解析数据
    print("\n[1/3] 解析财务数据...")
    try:
        engine.parse()
        parsed_bs = len(engine.bs_data)
        parsed_is = len(engine.is_data)
        prior_bs = len(engine.prior_bs_data)
        prior_is = len(engine.prior_is_data)
        print(f"  BS 科目: {parsed_bs} 项, IS 科目: {parsed_is} 项")
        if prior_bs or prior_is:
            print(f"  上期 BS: {prior_bs} 项, 上期 IS: {prior_is} 项")
        print(f"  期间数: {len(engine.periods)}")
    except ValueError as e:
        print(f"\n[错误] {e}", file=sys.stderr)
        sys.exit(1)

    # 执行分析
    print("\n[2/3] 执行分析程序...")
    results = {}

    if args.method in ("all", "ratios"):
        print("  → 比率分析...")
        results["ratios"] = engine.compute_ratios()
        print(f"    计算 {results['ratios']['count']} 个财务比率")

    if args.method in ("all", "trend"):
        print("  → 趋势分析...")
        results["trends"] = engine.trend_analysis()
        trend_count = len(results["trends"].get("trends", {}))
        print(f"    分析 {trend_count} 个科目趋势")

    if args.method in ("all", "benford"):
        print("  → Benford 检验...")
        results["benford"] = engine.benford_test()
        bf = results["benford"]
        if bf.get("warning"):
            print(f"    {bf['warning']}")
        else:
            print(
                f"    样本量: {bf['total_count']}, χ²={bf['chi_square']:.2f}, "
                f"符合Benford: {'是' if bf['is_conforming'] else '否'}"
            )

    if args.method in ("all", "comparison"):
        print("  → 期间对比...")
        results["comparison"] = engine.period_comparison()
        comp_count = len(results["comparison"].get("comparisons", []))
        print(f"    对比 {comp_count} 个科目")

    # 汇总
    results["summary"] = {
        "total_anomalies": len(engine._anomalies),
        "critical_count": sum(1 for a in engine._anomalies if a["严重程度"] == "重大"),
        "warning_count": sum(1 for a in engine._anomalies if a["严重程度"] == "警告"),
        "anomalies": engine._anomalies,
        "execution_time": datetime.now().isoformat(),
        "input_file": args.input,
        "materiality": args.materiality,
        "fluctuation_threshold": args.fluctuation_threshold,
        "industry": args.industry,
    }

    # 生成输出
    print(f"\n[3/3] 生成分析报告: {args.output}")
    engine.results = results  # sync
    engine.generate_output(args.output, results, generate_charts=args.chart)

    # 最终统计
    s = results["summary"]
    print(f"\n══════════════════════════════════════════════")
    print(f"  分析完成")
    print(f"  异常总数: {s['total_anomalies']}")
    print(f"  重大异常: {s['critical_count']}")
    print(f"  警告异常: {s['warning_count']}")
    print(f"  报告文件: {args.output}")
    print(f"══════════════════════════════════════════════")

    # 退出码：有重大异常时返回 2
    if s["critical_count"] > 0:
        sys.exit(2)
    elif s["warning_count"] > 0:
        sys.exit(1)
    else:
        sys.exit(0)


if __name__ == "__main__":
    main()

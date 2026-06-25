#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
审计抽样引擎 — Audit Sampling Engine
======================================

实现四种审计抽样方法，参考：
  - 《中国注册会计师审计准则第1314号——审计抽样》（CSA 1314）
  - AICPA Audit Guide: Audit Sampling (2017)

抽样方法：
  1. PPS/MUS 货币单位抽样 — 实质性程序（变量抽样）
  2. 属性抽样            — 控制测试（属性抽样）
  3. 分层抽样            — 结合 Neyman 最优分配
  4. 简单随机抽样        — 无放回抽样

依赖: openpyxl, (可选) numpy
Python 3.9+ 兼容

Author: Audit Toolkit
Version: 1.0.0
"""

import argparse
import datetime
import os
import random
import sys
from typing import Any, Dict, List, Optional, Tuple, Union

# ---------------------------------------------------------------------------
# 可选依赖
# ---------------------------------------------------------------------------
try:
    import numpy as np

    _HAS_NUMPY = True
except ImportError:
    _HAS_NUMPY = False

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("缺少依赖 openpyxl，请执行: pip install openpyxl")

# ---------------------------------------------------------------------------
# 常量：可靠性因子 RF（零预期错报时）
#   来源: AICPA Audit Sampling Guide, Table 4-5 / C-3
#   对应过度信赖风险（误受风险）
# ---------------------------------------------------------------------------
RELIABILITY_FACTORS: Dict[int, float] = {
    1: 4.61,  # 风险 1%
    5: 3.00,  # 风险 5%
    10: 2.31,  # 风险 10%
    15: 1.90,  # 风险 15%
    20: 1.61,  # 风险 20%
    25: 1.39,  # 风险 25%
    30: 1.21,  # 风险 30%
}

# ---------------------------------------------------------------------------
# 常量：扩展因子 EF（预期错报 >0 时使用）
#   来源: AICPA Audit Sampling Guide, Table 4-5 / C-3
# ---------------------------------------------------------------------------
EXPANSION_FACTORS: Dict[int, float] = {
    1: 1.90,  # 风险 1%
    5: 1.60,  # 风险 5%
    10: 1.50,  # 风险 10%
    15: 1.40,  # 风险 15%
    20: 1.30,  # 风险 20%
    25: 1.25,  # 风险 25%
    30: 1.20,  # 风险 30%
}

# ---------------------------------------------------------------------------
# 属性抽样样本量表（节选）
#   来源: AICPA Audit Guide — Appendix A
#   置信水平 → 可容忍偏差率 → 样本量（预期偏差率=0.00~0.05）
# ---------------------------------------------------------------------------
ATTRIBUTE_SAMPLE_TABLE: Dict[str, Dict[str, Dict[float, int]]] = {
    # 置信水平 90%
    "90": {
        "2": {0.00: 114, 0.01: 194, 0.02: 194, 0.03: 194, 0.04: 194, 0.05: 194},
        "3": {0.00: 76, 0.01: 129, 0.02: 176, 0.03: 221, 0.04: 221, 0.05: 221},
        "4": {0.00: 57, 0.01: 96, 0.02: 132, 0.03: 166, 0.04: 198, 0.05: 262},
        "5": {0.00: 45, 0.01: 77, 0.02: 105, 0.03: 132, 0.04: 158, 0.05: 209},
        "6": {0.00: 38, 0.01: 64, 0.02: 88, 0.03: 110, 0.04: 132, 0.05: 174},
        "7": {0.00: 32, 0.01: 55, 0.02: 75, 0.03: 94, 0.04: 113, 0.05: 149},
        "8": {0.00: 28, 0.01: 48, 0.02: 65, 0.03: 82, 0.04: 98, 0.05: 130},
        "9": {0.00: 25, 0.01: 42, 0.02: 58, 0.03: 73, 0.04: 87, 0.05: 115},
        "10": {0.00: 22, 0.01: 38, 0.02: 52, 0.03: 65, 0.04: 78, 0.05: 103},
        "12": {0.00: 18, 0.01: 31, 0.02: 42, 0.03: 53, 0.04: 64, 0.05: 85},
        "14": {0.00: 15, 0.01: 26, 0.02: 35, 0.03: 44, 0.04: 53, 0.05: 70},
        "16": {0.00: 13, 0.01: 22, 0.02: 30, 0.03: 38, 0.04: 45, 0.05: 60},
        "18": {0.00: 11, 0.01: 19, 0.02: 26, 0.03: 33, 0.04: 39, 0.05: 52},
        "20": {0.00: 10, 0.01: 17, 0.02: 23, 0.03: 29, 0.04: 35, 0.05: 46},
    },
    # 置信水平 95%
    "95": {
        "2": {0.00: 149, 0.01: 236, 0.02: 313, 0.03: 386, 0.04: 458, 0.05: 590},
        "3": {0.00: 99, 0.01: 157, 0.02: 208, 0.03: 257, 0.04: 303, 0.05: 392},
        "4": {0.00: 74, 0.01: 117, 0.02: 156, 0.03: 192, 0.04: 227, 0.05: 294},
        "5": {0.00: 59, 0.01: 93, 0.02: 124, 0.03: 153, 0.04: 181, 0.05: 234},
        "6": {0.00: 49, 0.01: 78, 0.02: 103, 0.03: 127, 0.04: 150, 0.05: 195},
        "7": {0.00: 42, 0.01: 66, 0.02: 88, 0.03: 109, 0.04: 129, 0.05: 167},
        "8": {0.00: 36, 0.01: 57, 0.02: 77, 0.03: 95, 0.04: 112, 0.05: 146},
        "9": {0.00: 32, 0.01: 51, 0.02: 68, 0.03: 84, 0.04: 100, 0.05: 130},
        "10": {0.00: 29, 0.01: 46, 0.02: 61, 0.03: 76, 0.04: 89, 0.05: 116},
        "12": {0.00: 24, 0.01: 38, 0.02: 50, 0.03: 62, 0.04: 74, 0.05: 96},
        "14": {0.00: 20, 0.01: 31, 0.02: 42, 0.03: 52, 0.04: 61, 0.05: 80},
        "16": {0.00: 17, 0.01: 27, 0.02: 36, 0.03: 45, 0.04: 53, 0.05: 69},
        "18": {0.00: 15, 0.01: 24, 0.02: 32, 0.03: 39, 0.04: 47, 0.05: 61},
        "20": {0.00: 14, 0.01: 21, 0.02: 29, 0.03: 35, 0.04: 42, 0.05: 54},
    },
    # 置信水平 99%
    "99": {
        "2": {0.00: 228, 0.01: 368, 0.02: 531, 0.03: 594, 0.04: 727, 0.05: 927},
        "3": {0.00: 152, 0.01: 245, 0.02: 354, 0.03: 461, 0.04: 564, 0.05: 718},
        "4": {0.00: 114, 0.01: 184, 0.02: 265, 0.03: 346, 0.04: 423, 0.05: 539},
        "5": {0.00: 91, 0.01: 147, 0.02: 212, 0.03: 276, 0.04: 338, 0.05: 431},
        "6": {0.00: 76, 0.01: 122, 0.02: 176, 0.03: 230, 0.04: 281, 0.05: 358},
        "7": {0.00: 65, 0.01: 105, 0.02: 151, 0.03: 197, 0.04: 241, 0.05: 307},
        "8": {0.00: 57, 0.01: 92, 0.02: 132, 0.03: 173, 0.04: 211, 0.05: 269},
        "9": {0.00: 51, 0.01: 82, 0.02: 118, 0.03: 154, 0.04: 188, 0.05: 240},
        "10": {0.00: 46, 0.01: 74, 0.02: 106, 0.03: 139, 0.04: 169, 0.05: 216},
        "12": {0.00: 38, 0.01: 61, 0.02: 88, 0.03: 115, 0.04: 140, 0.05: 179},
        "14": {0.00: 32, 0.01: 52, 0.02: 75, 0.03: 98, 0.04: 119, 0.05: 153},
        "16": {0.00: 28, 0.01: 45, 0.02: 65, 0.03: 85, 0.04: 104, 0.05: 133},
        "18": {0.00: 25, 0.01: 40, 0.02: 58, 0.03: 75, 0.04: 92, 0.05: 118},
        "20": {0.00: 22, 0.01: 36, 0.02: 52, 0.03: 67, 0.04: 82, 0.05: 105},
    },
}


# ===================================================================
# 辅助函数
# ===================================================================


def _resolve_column_index(
    headers: List[str], col_spec: Union[str, int], sheet_context: str = ""
) -> int:
    """根据列名或列号解析列索引（0-based）。

    Parameters
    ----------
    headers : List[str]
        表头行列表
    col_spec : str 或 int
        列名（字符串匹配）或 1-based 列号
    sheet_context : str
        上下文信息，用于错误提示

    Returns
    -------
    int
        0-based 列索引

    Raises
    ------
    ValueError
        列名未在表头中找到时抛出
    """
    if isinstance(col_spec, int):
        # 用户输入1-based列号，转为0-based
        if col_spec < 1:
            raise ValueError(f"列号必须 >=1，当前值: {col_spec}（{sheet_context}）")
        return col_spec - 1
    col_spec_str = str(col_spec).strip()
    # 尝试精确匹配
    for i, h in enumerate(headers):
        if h is not None and str(h).strip() == col_spec_str:
            return i
    # 尝试模糊匹配（包含关系）
    for i, h in enumerate(headers):
        if h is not None and col_spec_str in str(h).strip():
            return i
    raise ValueError(f"未找到列'{col_spec_str}'（{sheet_context}），可用列: {headers}")


def _safe_float(value: Any) -> float:
    """安全转换为float，无法转换时返回0.0。"""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def _auto_column_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    """自适应列宽（基于表头内容长度）。

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        目标工作表
    min_width : int
        最小列宽
    max_width : int
        最大列宽
    """
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                # CJK 字符按 2 倍宽度估算
                val = str(cell.value)
                line_lengths = []
                for line in val.split("\n"):
                    length = 0
                    for ch in line:
                        length += (
                            2
                            if "\u4e00" <= ch <= "\u9fff" or "\u3000" <= ch <= "\u303f"
                            else 1
                        )
                    line_lengths.append(length)
                max_len = max(max_len, max(line_lengths))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _style_header(ws, row_idx: int = 1) -> None:
    """设置表头行样式（加粗、浅灰背景、冻结首行）。

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
    row_idx : int
        表头所在行号（1-based）
    """
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for cell in ws[row_idx]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws.freeze_panes = ws.cell(row=row_idx + 1, column=1)


# ===================================================================
# 核心类
# ===================================================================


class AuditSamplingEngine:
    """审计抽样引擎 — 支持 PPS/MUS、属性抽样、分层抽样、简单随机抽样。

    Usage
    -----
    >>> engine = AuditSamplingEngine(
    ...     input_path="data.xlsx",
    ...     method="pps",
    ...     sheet_name="凭证",
    ...     key_col="凭证号",
    ...     amount_col="金额",
    ...     tolerable_misstatement=50000,
    ...     risk=10,
    ...     seed=42,
    ... )
    >>> engine.run()
    >>> engine.export("output.xlsx")
    """

    # ---- PPS/MUS 支持的误受风险值 ----
    VALID_RISKS = (1, 5, 10, 15, 20, 25, 30)

    # ---- 属性抽样支持的置信水平 ----
    VALID_CONFIDENCES = (90, 95, 99)

    def __init__(
        self,
        input_path: str,
        method: str = "pps",
        sheet_name: Optional[str] = None,
        key_col: Optional[str] = None,
        amount_col: Optional[str] = None,
        header_row: int = 1,
        population_amount: Optional[float] = None,
        tolerable_misstatement: Optional[float] = None,
        expected_misstatement: float = 0.0,
        risk: int = 10,
        sample_size: Optional[int] = None,
        num_strata: int = 3,
        seed: int = 42,
        confidence: int = 95,
        tolerable_rate: float = 0.05,
        expected_rate: float = 0.01,
    ):
        """初始化抽样引擎。

        Parameters
        ----------
        input_path : str
            输入 Excel 文件路径
        method : str
            抽样方法: 'pps' / 'attribute' / 'stratified' / 'random'
        sheet_name : str 或 None
            工作表名，None 则取第一个 sheet
        key_col : str 或 None
            关键识别列名（如"凭证号"），None 则使用行号
        amount_col : str 或 None
            金额列名，PPS 和分层抽样必须
        header_row : int
            表头所在行号（1-based），默认为第1行
        population_amount : float 或 None
            总体金额，None 则从 amount_col 自动求和
        tolerable_misstatement : float 或 None
            可容忍错报（PPS 必需）
        expected_misstatement : float
            预期错报（默认 0）
        risk : int
            误受风险百分比（1/5/10/15/20/25/30），默认 10
        sample_size : int 或 None
            指定样本量，覆盖自动计算
        num_strata : int
            分层数量（分层抽样用），默认 3
        seed : int
            随机种子，默认 42
        confidence : int
            置信水平百分比（属性抽样用），默认 95
        tolerable_rate : float
            可容忍偏差率（属性抽样用），默认 0.05
        expected_rate : float
            预期偏差率（属性抽样用），默认 0.01
        """
        # ---- 参数校验 ----
        method_lower = method.lower().strip()
        if method_lower not in ("pps", "mus", "attribute", "stratified", "random"):
            raise ValueError(
                f"不支持的抽样方法 '{method}'，"
                f"可选: pps/mus, attribute, stratified, random"
            )
        self.method = "pps" if method_lower in ("pps", "mus") else method_lower

        if (
            self.method == "pps"
            and tolerable_misstatement is None
            and sample_size is None
        ):
            raise ValueError(
                "PPS 抽样必须提供 --tolerable（可容忍错报）或 --sample-size"
            )
        if self.method == "pps" and amount_col is None:
            raise ValueError("PPS 抽样必须提供 --amount-col（金额列）")

        if risk not in self.VALID_RISKS:
            raise ValueError(f"误受风险必须为 {self.VALID_RISKS} 之一，当前值: {risk}")

        self.input_path = input_path
        self.sheet_name = sheet_name
        self.key_col = key_col
        self.amount_col = amount_col
        self.header_row = header_row
        self.population_amount = population_amount
        self.tolerable_misstatement = tolerable_misstatement
        self.expected_misstatement = expected_misstatement
        self.risk = risk
        self.user_sample_size = sample_size
        self.num_strata = num_strata
        self.seed = seed
        self.confidence = confidence
        self.tolerable_rate = tolerable_rate
        self.expected_rate = expected_rate

        # ---- 运行时状态 ----
        self.data: List[List[Any]] = []  # 所有原始数据行（含表头之前的行）
        self.headers: List[str] = []  # 表头行
        self.records: List[List[Any]] = []  # 数据记录（不含表头）
        self.amount_values: List[float] = []  # 金额列对应的数值
        self.key_values: List[str] = []  # 关键识别的值
        self.sampled_indices: List[int] = []  # 抽样结果索引（records 中的索引）
        self.computed_sample_size: int = 0  # 自动计算的样本量
        self.total_population: float = 0.0  # 实际总体金额
        self.timestamp: str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.record_count: int = 0
        self.params_snapshot: Dict[str, Any] = {}  # 抽样参数存档

    # ------------------------------------------------------------------
    # 数据加载
    # ------------------------------------------------------------------

    def load_data(self) -> None:
        """从 Excel 加载数据，解析表头和金额列。

        加载后填充 self.headers, self.records, self.amount_values, self.key_values。
        """
        if not os.path.exists(self.input_path):
            raise FileNotFoundError(f"文件不存在: {self.input_path}")

        wb = openpyxl.load_workbook(self.input_path, data_only=True)
        if self.sheet_name:
            if self.sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"工作表 '{self.sheet_name}' 不存在，可用: {wb.sheetnames}"
                )
            ws = wb[self.sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]
            self.sheet_name = ws.title

        # 读取所有数据
        all_data = []
        for row in ws.iter_rows(values_only=True):
            all_data.append(list(row))

        wb.close()

        if len(all_data) < self.header_row + 1:
            raise ValueError(
                f"数据行数不足：共 {len(all_data)} 行，"
                f"表头行={self.header_row}，至少需要 {self.header_row + 1} 行"
            )

        self.data = all_data
        # 提取表头
        header_values = all_data[self.header_row - 1]
        self.headers = [str(h) if h is not None else "" for h in header_values]

        # 提取数据记录（表头之后的各行）
        self.records = all_data[self.header_row :]
        self.record_count = len(self.records)

        # 解析金额列
        if self.amount_col is not None:
            amt_idx = _resolve_column_index(
                self.headers,
                self.amount_col,
                f"文件={os.path.basename(self.input_path)}, sheet={self.sheet_name}",
            )
            self.amount_values = [
                _safe_float(r[amt_idx]) if len(r) > amt_idx else 0.0
                for r in self.records
            ]
            self.total_population = sum(self.amount_values)
        else:
            self.total_population = 0.0

        # 如果用户指定了总体金额，覆盖自动计算
        if self.population_amount is not None:
            self.total_population = self.population_amount

        # 解析关键识别列
        if self.key_col is not None:
            key_idx = _resolve_column_index(
                self.headers,
                self.key_col,
                f"文件={os.path.basename(self.input_path)}, sheet={self.sheet_name}",
            )
            self.key_values = [
                str(r[key_idx])
                if len(r) > key_idx and r[key_idx] is not None
                else f"Row_{self.header_row + 1 + i}"
                for i, r in enumerate(self.records)
            ]
        else:
            self.key_values = [
                f"Row_{self.header_row + 1 + i}" for i in range(self.record_count)
            ]

    # ------------------------------------------------------------------
    # 样本量计算
    # ------------------------------------------------------------------

    def _calc_pps_size(self) -> int:
        """计算 PPS/MUS 样本量。

        Formula
        -------
            n = (BV × RF) / (TM - EM × EF)

        其中:
            BV = 总体账面值（Book Value）
            RF = 可靠性因子（Reliability Factor，基于零预期错报）
            TM = 可容忍错报（Tolerable Misstatement）
            EM = 预期错报（Expected Misstatement）
            EF = 扩展因子（Expansion Factor，仅 EM > 0 时使用）

        当分母 ≤ 0 时（TM <= EM × EF），返回 None 表示需要全量检查。

        Returns
        -------
        int
            计算得到的样本量
        """
        bv = self.total_population
        tm = self.tolerable_misstatement
        em = self.expected_misstatement
        rf = RELIABILITY_FACTORS.get(self.risk, 2.31)  # 默认 10%
        ef = EXPANSION_FACTORS.get(self.risk, 1.50)

        if bv <= 0:
            raise ValueError(
                f"总体金额必须 >0，当前值: {bv}。请检查数据或手动指定 --population。"
            )
        if tm is None or tm <= 0:
            raise ValueError(f"可容忍错报必须 >0，当前值: {tm}。")

        if em > 0:
            # 含预期错报的扩展公式
            denominator = tm - em * ef
            if denominator <= 0:
                raise ValueError(
                    f"可容忍错报({tm}) ≤ 预期错报×扩展因子({em}×{ef}={em * ef:.2f})，"
                    f"建议进行全量实质性程序而非抽样。"
                )
            n = (bv * rf) / denominator
        else:
            # 零预期错报时的简化公式
            n = (bv * rf) / tm

        return max(1, int(round(n)))

    def _calc_attribute_size(self) -> Optional[int]:
        """查表获取属性抽样样本量。

        使用 AICPA 标准样本量表，按置信水平 > 可容忍偏差率 > 预期偏差率 查找。

        Returns
        -------
        int 或 None
            样本量，若查不到对应组合则返回 None
        """
        conf_key = str(self.confidence)
        if conf_key not in ATTRIBUTE_SAMPLE_TABLE:
            raise ValueError(
                f"属性抽样仅支持置信水平 {list(ATTRIBUTE_SAMPLE_TABLE.keys())}%"
            )

        tolerable_pct = int(self.tolerable_rate * 100)
        expected_pct = self.expected_rate

        table = ATTRIBUTE_SAMPLE_TABLE[conf_key]

        # 找到 ≤ 指定可容忍偏差率的最大条目（保守原则）
        tolerable_str = None
        for k in sorted(table.keys(), key=int, reverse=True):
            if int(k) <= tolerable_pct:
                tolerable_str = k
                break
        if tolerable_str is None:
            tolerable_str = sorted(table.keys(), key=int)[0]

        # 找到 ≥ 预期偏差率的条目
        er_keys = sorted(table[tolerable_str].keys())
        er_key = er_keys[0]
        for k in er_keys:
            if k >= expected_pct:
                er_key = k
                break
        else:
            er_key = er_keys[-1]

        sample_size = table[tolerable_str].get(er_key)
        return sample_size

    def _calc_stratified_size(self) -> int:
        """计算分层抽样的总样本量。

        默认按总体记录数的平方根取整（经验法则）。
        若用户用 --sample-size 覆盖则直接使用。

        Returns
        -------
        int
            推荐的总样本量
        """
        return max(1, int(round(self.record_count**0.5 * 0.8)))

    def _calc_random_size(self) -> int:
        """计算简单随机抽样的样本量（经验公式）。

        基于 AICPA 非统计抽样经验法则：
            n = √N × 0.8  (N = 总体记录数)

        Returns
        -------
        int
            推荐样本量
        """
        return max(1, int(round(self.record_count**0.5 * 0.8)))

    def _determine_sample_size(self) -> int:
        """根据方法确定样本量（用户指定优先于自动计算）。

        Returns
        -------
        int
            最终样本量
        """
        if self.user_sample_size is not None:
            if self.user_sample_size < 1:
                raise ValueError("样本量必须 >=1")
            if self.user_sample_size > self.record_count:
                raise ValueError(
                    f"样本量({self.user_sample_size}) 超过总体记录数({self.record_count})"
                )
            return self.user_sample_size

        if self.method == "pps":
            size = self._calc_pps_size()
        elif self.method == "attribute":
            size = self._calc_attribute_size()
            if size is None:
                raise ValueError(
                    "无法在属性抽样样本量表中找到匹配组合，"
                    "请使用 --sample-size 手动指定样本量。"
                )
        elif self.method == "stratified":
            size = self._calc_stratified_size()
        elif self.method == "random":
            size = self._calc_random_size()
        else:
            size = max(1, int(round(self.record_count**0.5)))

        # 确保样本量不超总体
        return min(size, self.record_count)

    # ------------------------------------------------------------------
    # 抽样方法实现
    # ------------------------------------------------------------------

    def _do_pps_selection(self, sample_size: int) -> List[int]:
        """PPS/MUS 货币单位抽样 — 系统性选样。

        原理：
        1. 计算抽样间隔 = BV / n
        2. 随机起点 ∈ [0, 抽样间隔)
        3. 按记录累计金额逐一扫描，累计金额跨越"起点+k×间隔"时选中该记录
        4. 一条记录可能被多次选中（金额大的概率高）

        Parameters
        ----------
        sample_size : int
            样本量

        Returns
        -------
        List[int]
            被选中记录在 self.records 中的索引列表
        """
        bv = self.total_population
        if bv <= 0:
            raise ValueError("总体金额必须 >0 才能执行 PPS 选样")

        interval = bv / sample_size
        rng = random.Random(self.seed)
        start = rng.uniform(0, interval)

        selection_points = [start + i * interval for i in range(sample_size)]
        selection_points.sort()

        # 按累计金额 systematic 选样
        selected: List[int] = []
        cumulative = 0.0
        sp_idx = 0

        for i, amt in enumerate(self.amount_values):
            if sp_idx >= len(selection_points):
                break
            cumulative += amt
            while (
                sp_idx < len(selection_points)
                and cumulative >= selection_points[sp_idx]
            ):
                selected.append(i)
                sp_idx += 1

        return selected

    def _do_attribute_selection(self, sample_size: int) -> List[int]:
        """属性抽样选样 — 无放回简单随机选样。

        Parameters
        ----------
        sample_size : int
            样本量

        Returns
        -------
        List[int]
            被选中记录索引列表
        """
        return self._random_select(sample_size)

    def _do_stratified_selection(self, sample_size: int) -> List[int]:
        """分层抽样选样 — Neyman 最优分配。

        步骤：
        1. 按金额排序后划分为 num_strata 层
        2. 计算每层金额标准差
        3. 按 Neyman 分配计算各层样本量: n_h = n × (N_h × σ_h) / Σ(N_i × σ_i)
        4. 各层内独立随机选样

        Parameters
        ----------
        sample_size : int
            总样本量

        Returns
        -------
        List[int]
            被选中记录索引列表
        """
        n_strata = min(self.num_strata, self.record_count)

        if n_strata <= 1 or sample_size <= 1:
            return self._random_select(sample_size)

        # 按金额值和原始索引排序
        indexed = sorted(
            enumerate(self.amount_values),
            key=lambda x: x[1],
        )

        # 等规模分层
        per_stratum = max(1, self.record_count // n_strata)
        strata: List[List[int]] = []  # 每层存放 (原始索引) 列表
        for s in range(n_strata):
            start = s * per_stratum
            if s == n_strata - 1:
                end = self.record_count
            else:
                end = start + per_stratum
            strata.append([idx for idx, _ in indexed[start:end]])

        # 过滤空层
        strata = [s for s in strata if s]

        # Neyman 分配
        n_st = len(strata)
        alloc = []
        weights = []
        for st in strata:
            amounts = [self.amount_values[idx] for idx in st]
            if _HAS_NUMPY:
                std = float(np.std(amounts, ddof=1)) if len(amounts) > 1 else 0.0
            else:
                # 纯 Python 标准差
                m = sum(amounts) / len(amounts)
                var = (
                    sum((x - m) ** 2 for x in amounts) / (len(amounts) - 1)
                    if len(amounts) > 1
                    else 0.0
                )
                std = var**0.5
            weight = len(st) * std
            alloc.append(len(st))
            weights.append(weight)

        total_weight = sum(weights)
        if total_weight > 0:
            stratum_sizes = [
                max(1, int(round(sample_size * w / total_weight))) for w in weights
            ]
        else:
            # 所有金额相等，等比例分配
            stratum_sizes = [
                max(1, int(round(sample_size * len(st) / self.record_count)))
                for st in strata
            ]

        # 调平（确保总和 ≤ sample_size）
        while sum(stratum_sizes) > sample_size:
            max_idx = max(range(n_st), key=lambda i: stratum_sizes[i])
            if stratum_sizes[max_idx] > 1:
                stratum_sizes[max_idx] -= 1
            else:
                break

        # 各层内随机选样
        rng = random.Random(self.seed)
        result = []
        for si, st in enumerate(strata):
            k = min(stratum_sizes[si], len(st))
            chosen = rng.sample(st, k)
            result.extend(chosen)

        return sorted(result)

    def _do_random_selection(self, sample_size: int) -> List[int]:
        """简单随机抽样 — 无放回。

        Parameters
        ----------
        sample_size : int
            样本量

        Returns
        -------
        List[int]
            被选中记录索引列表
        """
        return self._random_select(sample_size)

    def _random_select(self, sample_size: int) -> List[int]:
        """无放回随机选样（内部工具方法）。

        Parameters
        ----------
        sample_size : int
            样本量

        Returns
        -------
        List[int]
            被选中记录索引列表
        """
        rng = random.Random(self.seed)
        population = list(range(self.record_count))
        return sorted(rng.sample(population, min(sample_size, self.record_count)))

    # ------------------------------------------------------------------
    # 主流程
    # ------------------------------------------------------------------

    def run(self) -> List[List[Any]]:
        """执行抽样全流程：加载数据 → 计算样本量 → 选样。

        Returns
        -------
        List[List[Any]]
            抽样结果列表，每条为完整的原始记录 + 抽样标记
        """
        # 1. 加载数据
        self.load_data()

        # 2. 确定样本量
        self.computed_sample_size = self._determine_sample_size()

        # 3. 按方法选样
        if self.method == "pps":
            self.sampled_indices = self._do_pps_selection(self.computed_sample_size)
            # PPS 可能重复选中同一条记录
            seen: Dict[int, int] = {}
            for idx in self.sampled_indices:
                seen[idx] = seen.get(idx, 0) + 1
            unique_count = len(seen)
        elif self.method == "attribute":
            self.sampled_indices = self._do_attribute_selection(
                self.computed_sample_size
            )
            unique_count = len(self.sampled_indices)
        elif self.method == "stratified":
            self.sampled_indices = self._do_stratified_selection(
                self.computed_sample_size
            )
            unique_count = len(self.sampled_indices)
        elif self.method == "random":
            self.sampled_indices = self._do_random_selection(self.computed_sample_size)
            unique_count = len(self.sampled_indices)
        else:
            raise RuntimeError(f"内部错误：未知方法 '{self.method}'")

        # 4. 构建抽样结果
        result = []
        for idx in self.sampled_indices:
            rec = list(self.records[idx])
            result.append(rec)

        # 5. 存档参数
        self.params_snapshot = {
            "方法": self._method_display_name(),
            "输入文件": self.input_path,
            "工作表": self.sheet_name,
            "表头行": self.header_row,
            "关键列": self.key_col or "(未指定)",
            "金额列": self.amount_col or "(未指定)",
            "总体记录数": self.record_count,
            "总体金额": self.total_population,
            "计算样本量": self.computed_sample_size,
            "实际选中记录数": len(self.sampled_indices),
            "唯一样本数": unique_count,
            "随机种子": self.seed,
        }

        if self.method == "pps":
            self.params_snapshot["可容忍错报"] = self.tolerable_misstatement
            self.params_snapshot["预期错报"] = self.expected_misstatement
            self.params_snapshot["误受风险"] = f"{self.risk}%"
            self.params_snapshot["可靠性因子(RF)"] = RELIABILITY_FACTORS.get(
                self.risk, "N/A"
            )
            self.params_snapshot["扩展因子(EF)"] = EXPANSION_FACTORS.get(
                self.risk, "N/A"
            )
            if self.total_population > 0:
                self.params_snapshot["抽样间隔"] = round(
                    self.total_population / self.computed_sample_size, 2
                )

        if self.method == "attribute":
            self.params_snapshot["置信水平"] = f"{self.confidence}%"
            self.params_snapshot["可容忍偏差率"] = self.tolerable_rate
            self.params_snapshot["预期偏差率"] = self.expected_rate

        if self.method == "stratified":
            self.params_snapshot["分层数"] = self.num_strata
            self.params_snapshot["分配方式"] = "Neyman最优分配"

        self.params_snapshot["执行时间"] = self.timestamp

        return result

    # ------------------------------------------------------------------
    # 导出 Excel
    # ------------------------------------------------------------------

    def export(self, output_path: Optional[str] = None) -> str:
        """导出抽样结果到 Excel 工作簿。

        输出两个 Sheet：
          - "抽样清单": 选中的记录（原始数据 + 抽样标记）
          - "抽样参数": 审计轨迹（所有参数存档）

        Parameters
        ----------
        output_path : str 或 None
            输出路径，None 则自动生成为 input 同目录下 *_samples.xlsx

        Returns
        -------
        str
            实际输出文件路径
        """
        if output_path is None:
            base = os.path.splitext(os.path.basename(self.input_path))[0]
            out_dir = os.path.dirname(self.input_path) or "."
            output_path = os.path.join(out_dir, f"{base}_samples.xlsx")

        wb = openpyxl.Workbook()

        # ===================== Sheet 1: 抽样清单 =====================
        ws1 = wb.active
        ws1.title = "抽样清单"

        # 构建表头：原始列名 + 抽样标记 + 金额
        output_headers = list(self.headers) + ["抽样序号", "选中金额"]
        for ci, h in enumerate(output_headers, start=1):
            ws1.cell(row=1, column=ci, value=h)

        amt_idx = None
        if self.amount_col is not None:
            try:
                amt_idx = _resolve_column_index(self.headers, self.amount_col)
            except ValueError:
                pass

        for si, rec_idx in enumerate(self.sampled_indices, start=1):
            row_num = si + 1  # 1-based, 跳过表头
            rec = self.records[rec_idx]
            for ci, val in enumerate(rec, start=1):
                ws1.cell(row=row_num, column=ci, value=val)
            ws1.cell(row=row_num, column=len(self.headers) + 1, value=si)
            if amt_idx is not None and amt_idx < len(rec):
                ws1.cell(
                    row=row_num,
                    column=len(self.headers) + 2,
                    value=_safe_float(rec[amt_idx]),
                )

        _style_header(ws1)
        _auto_column_width(ws1)

        # ===================== Sheet 2: 抽样参数 =====================
        ws2 = wb.create_sheet("抽样参数")

        param_font = Font(bold=True)
        ws2.cell(row=1, column=1, value="审计抽样参数存档").font = Font(
            bold=True, size=14
        )
        ws2.cell(row=2, column=1, value=f"执行时间: {self.timestamp}").font = Font(
            italic=True, color="666666"
        )
        ws2.cell(row=3, column=1, value="")  # 空行

        row = 4
        for key, value in self.params_snapshot.items():
            ws2.cell(row=row, column=1, value=key).font = param_font
            ws2.cell(row=row, column=2, value=str(value) if value is not None else "-")
            row += 1

        # 方法公式说明
        row += 1
        ws2.cell(row=row, column=1, value="抽样方法说明").font = Font(
            bold=True, size=12
        )
        row += 1
        method_descriptions = {
            "pps": "PPS/MUS 货币单位抽样 — n = (BV×RF) / (TM - EM×EF)，系统性等距选样",
            "attribute": f"属性抽样 — AICPA 样本量表查表，置信水平 {self.confidence}%",
            "stratified": f"分层抽样 — Neyman 最优分配，{self.num_strata} 层",
            "random": "简单随机抽样 — 无放回随机选样",
        }
        ws2.cell(
            row=row, column=1, value=method_descriptions.get(self.method, self.method)
        )
        row += 2
        ws2.cell(row=row, column=1, value="参考标准").font = Font(bold=True)
        row += 1
        ws2.cell(
            row=row, column=1, value="《中国注册会计师审计准则第1314号——审计抽样》"
        )
        row += 1
        ws2.cell(row=row, column=1, value="AICPA Audit Guide: Audit Sampling (2017)")

        _auto_column_width(ws2, min_width=18, max_width=60)

        # 保存
        wb.save(output_path)
        return output_path

    # ------------------------------------------------------------------
    # 工具方法
    # ------------------------------------------------------------------

    def _method_display_name(self) -> str:
        """获取方法的显示名称。"""
        names = {
            "pps": "PPS/MUS 货币单位抽样",
            "attribute": "属性抽样",
            "stratified": "分层抽样",
            "random": "简单随机抽样",
        }
        return names.get(self.method, self.method)

    def summary(self) -> str:
        """生成抽样结果摘要。"""
        lines = [
            "=" * 60,
            f"  审计抽样结果摘要 — {self._method_display_name()}",
            "=" * 60,
            f"  输入文件:      {self.input_path}",
            f"  工作表:        {self.sheet_name}",
            f"  总体记录数:     {self.record_count}",
            f"  总体金额:       {self.total_population:,.2f}",
            f"  计算样本量:     {self.computed_sample_size}",
            f"  实际抽中:       {len(self.sampled_indices)} 条",
            f"  随机种子:       {self.seed}",
            f"  执行时间:       {self.timestamp}",
        ]
        if self.method == "pps":
            lines.append(f"  可容忍错报:     {self.tolerable_misstatement:,.2f}")
            lines.append(f"  预期错报:       {self.expected_misstatement:,.2f}")
            lines.append(f"  误受风险:       {self.risk}%")
            lines.append(f"  可靠性因子:     {RELIABILITY_FACTORS.get(self.risk)}")
        if self.method == "attribute":
            lines.append(f"  置信水平:       {self.confidence}%")
            lines.append(f"  可容忍偏差率:   {self.tolerable_rate}")
            lines.append(f"  预期偏差率:     {self.expected_rate}")
        if self.method == "stratified":
            lines.append(f"  分层数:         {self.num_strata}")
        lines.append("=" * 60)
        return "\n".join(lines)


# ===================================================================
# CLI 入口
# ===================================================================


def build_parser() -> argparse.ArgumentParser:
    """构建命令行参数解析器。"""
    parser = argparse.ArgumentParser(
        description="审计抽样引擎 — PPS/MUS · 属性抽样 · 分层抽样 · 随机抽样",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # PPS 货币单位抽样
  python sampling_engine.py --input data.xlsx --method pps \\
      --amount-col "借方金额" --tolerable 50000 --risk 10

  # 属性抽样（控制测试）
  python sampling_engine.py --input data.xlsx --method attribute \\
      --confidence 95 --tolerable-rate 0.05 --expected-rate 0.01

  # 分层抽样（5层）
  python sampling_engine.py --input data.xlsx --method stratified \\
      --amount-col "金额" --strata 5

  # 简单随机抽样，指定样本量
  python sampling_engine.py --input data.xlsx --method random \\
      --sample-size 100 --seed 12345

参考标准:
  《中国注册会计师审计准则第1314号——审计抽样》（CSA 1314）
  AICPA Audit Guide: Audit Sampling (2017)
        """,
    )

    # ---- 基础参数 ----
    parser.add_argument(
        "--input",
        "-i",
        required=True,
        help="输入 Excel 文件路径",
    )
    parser.add_argument(
        "--method",
        "-m",
        required=True,
        choices=["pps", "mus", "attribute", "stratified", "random"],
        help="抽样方法: pps/mus=货币单位抽样, attribute=属性抽样, "
        "stratified=分层抽样, random=简单随机抽样",
    )
    parser.add_argument(
        "--sheet",
        "-s",
        default=None,
        help="工作表名（默认第一个 sheet）",
    )
    parser.add_argument(
        "--key-col",
        "-k",
        default=None,
        help="关键识别列名（如'凭证号'）",
    )
    parser.add_argument(
        "--amount-col",
        "-a",
        default=None,
        help="金额列名（PPS 和分层抽样必须）",
    )
    parser.add_argument(
        "--header",
        "-H",
        type=int,
        default=1,
        help="表头所在行号（1-based），默认 1",
    )

    # ---- 总体参数 ----
    parser.add_argument(
        "--population",
        type=float,
        default=None,
        help="总体金额（可不填，自动从 amount-col 求和）",
    )

    # ---- PPS 参数 ----
    parser.add_argument(
        "--tolerable",
        "-t",
        type=float,
        default=None,
        help="可容忍错报（PPS 必须）",
    )
    parser.add_argument(
        "--expected",
        "-e",
        type=float,
        default=0.0,
        help="预期错报（PPS 用，默认 0）",
    )
    parser.add_argument(
        "--risk",
        "-r",
        type=int,
        default=10,
        choices=[1, 5, 10, 15, 20, 25, 30],
        help="误受风险百分比（PPS 用，默认 10）",
    )

    # ---- 属性抽样参数 ----
    parser.add_argument(
        "--confidence",
        "-c",
        type=int,
        default=95,
        choices=[90, 95, 99],
        help="置信水平百分比（属性抽样用，默认 95）",
    )
    parser.add_argument(
        "--tolerable-rate",
        type=float,
        default=0.05,
        help="可容忍偏差率（属性抽样用，默认 0.05）",
    )
    parser.add_argument(
        "--expected-rate",
        type=float,
        default=0.01,
        help="预期偏差率（属性抽样用，默认 0.01）",
    )

    # ---- 分层抽样参数 ----
    parser.add_argument(
        "--strata",
        type=int,
        default=3,
        help="分层数（分层抽样用，默认 3）",
    )

    # ---- 通用参数 ----
    parser.add_argument(
        "--sample-size",
        "-n",
        type=int,
        default=None,
        help="指定样本量（覆盖自动计算）",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=42,
        help="随机种子（默认 42，确保可复现）",
    )
    parser.add_argument(
        "--output",
        "-o",
        default=None,
        help="输出 Excel 路径（默认 input 同目录下加 _samples.xlsx）",
    )
    parser.add_argument(
        "--quiet",
        "-q",
        action="store_true",
        help="静默模式，只输出文件路径",
    )

    return parser


def main(argv: Optional[List[str]] = None) -> int:
    """CLI 主入口。

    Parameters
    ----------
    argv : list 或 None
        命令行参数，None 则使用 sys.argv

    Returns
    -------
    int
        退出码: 0=成功, 1=参数错误, 2=运行时错误
    """
    parser = build_parser()
    args = parser.parse_args(argv)

    # ---- 参数交叉校验 ----
    method = "pps" if args.method in ("pps", "mus") else args.method
    if method == "pps" and args.amount_col is None:
        parser.error("PPS 抽样必须提供 --amount-col（金额列）")
    if method == "pps" and args.tolerable is None and args.sample_size is None:
        parser.error("PPS 抽样必须提供 --tolerable（可容忍错报）或 --sample-size")

    # ---- 构建引擎 ----
    try:
        engine = AuditSamplingEngine(
            input_path=args.input,
            method=method,
            sheet_name=args.sheet,
            key_col=args.key_col,
            amount_col=args.amount_col,
            header_row=args.header,
            population_amount=args.population,
            tolerable_misstatement=args.tolerable,
            expected_misstatement=args.expected,
            risk=args.risk,
            sample_size=args.sample_size,
            num_strata=args.strata,
            seed=args.seed,
            confidence=args.confidence,
            tolerable_rate=args.tolerable_rate,
            expected_rate=args.expected_rate,
        )

        # ---- 执行抽样 ----
        engine.run()

        # ---- 导出结果 ----
        output_path = engine.export(args.output)

        if not args.quiet:
            print(engine.summary())
            print(f"\n输出文件: {output_path}")
        else:
            print(output_path)

        return 0

    except FileNotFoundError as e:
        print(f"[错误] 文件未找到: {e}", file=sys.stderr)
        return 1
    except ValueError as e:
        print(f"[错误] 参数无效: {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"[错误] 运行时异常: {e}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    sys.exit(main())

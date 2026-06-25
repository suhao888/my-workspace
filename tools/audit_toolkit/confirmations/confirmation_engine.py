#!/usr/bin/env python3
"""
审计函证管理引擎
=================
符合《中国注册会计师审计准则第1312号——函证》

功能模块:
    1. 函证项目自动筛选 — 金额阈值 + 覆盖率双重控制
    2. 函证控制表生成   — 15列标准化控制表
    3. 回函差异自动汇总 — 差异计算、原因分类、重大差异标红
    4. 统计摘要         — 发函率、回函率、覆盖率、差异分布

依赖:
    openpyxl (必须), numpy (可选, 用于数值加速)

用法:
    python confirmation_engine.py --input 往来明细.xlsx --output 函证控制表.xlsx \\
        --name-col "客户名称" --amount-col "期末余额" --type ar --threshold 10000

参考准则:
    - CAS 1312 函证 — 第十条至第十九条（函证决策、设计、实施、评价）
    - CAS 1312 应用指南 — 附录1-4（函证控制表范例、差异分析指引）

Author: 审计工具链项目组
Version: 1.0.0
License: Internal Use Only
"""

from __future__ import annotations

import argparse
import datetime
import os
import sys
import traceback
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

# ============================================================================
# 常量定义
# ============================================================================

# 函证类型前缀映射
CONFIRMATION_TYPE_PREFIX: Dict[str, str] = {
    "ar": "AR",  # 应收账款 (Accounts Receivable)
    "ap": "AP",  # 应付账款 (Accounts Payable)
    "other": "OT",  # 其他往来 (Other)
}
CONFIRMATION_TYPE_LABEL: Dict[str, str] = {
    "ar": "应收账款",
    "ap": "应付账款",
    "other": "其他往来",
}

# 控制表列名 (15列 + 1列特殊标记)
CONTROL_SHEET_HEADERS: List[str] = [
    "序号",
    "客户/供应商名称",
    "账面金额",
    "函证金额",
    "函证类型",  # 积极式 / 消极式
    "函证编号",
    "发函日期",
    "回函日期",
    "回函金额",
    "差异金额",
    "差异原因",
    "调节状态",
    "替代程序执行情况",
    "审计结论",
    "特殊标记",
    "备注",
]

# 差异汇总表列名
DIFF_SUMMARY_HEADERS: List[str] = [
    "序号",
    "客户/供应商名称",
    "账面金额",
    "回函金额",
    "差异金额",
    "差异率(%)",
    "差异原因",
    "是否重大差异",
    "备注",
]

# 统计摘要字段
SUMMARY_FIELDS: List[Tuple[str, str]] = [
    ("confirmation_type", "函证类型"),
    ("total_items", "筛选项目总数"),
    ("selected_items", "发函数量"),
    ("total_book_amount", "账面总金额"),
    ("selected_book_amount", "函证金额合计"),
    ("coverage_rate", "金额覆盖率(%)"),
    ("positive_count", "积极式函证数量"),
    ("negative_count", "消极式函证数量"),
    ("special_count", "特殊标记项目数量"),
    ("reply_count", "回函数量"),
    ("reply_rate", "回函率(%)"),
    ("total_difference", "差异金额合计"),
    ("material_diff_count", "重大差异数量(>重要性水平)"),
    ("avg_difference", "平均差异金额"),
    ("generated_at", "生成时间"),
]

# 样式定义
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=10)
RED_FONT = Font(name="微软雅黑", size=10, color="FF0000", bold=True)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
NUM_FORMAT = "#,##0.00"

# 默认参数
DEFAULT_COVERAGE = 0.70
DEFAULT_MIN_ITEMS = 10
DEFAULT_THRESHOLD = 0.0
DEFAULT_MATERIALITY = 0.0


# ============================================================================
# 数据类
# ============================================================================


@dataclass
class ConfirmationItem:
    """单条函证项目的数据结构"""

    index: int  # 原始行号
    name: str  # 客户/供应商名称
    book_amount: float  # 账面金额
    confirmation_amount: float = 0.0  # 函证金额（默认=账面金额）
    confirmation_type: str = "积极式"  # 积极式/消极式
    confirmation_number: str = ""  # 函证编号
    send_date: Optional[str] = None  # 发函日期
    reply_date: Optional[str] = None  # 回函日期
    reply_amount: Optional[float] = None  # 回函金额
    difference: Optional[float] = None  # 差异金额
    diff_pct: Optional[float] = None  # 差异率(%)
    diff_reason: str = ""  # 差异原因
    reconciliation_status: str = "待发函"  # 调节状态
    alt_procedures: str = ""  # 替代程序执行情况
    audit_conclusion: str = ""  # 审计结论
    special_flag: str = ""  # 特殊标记
    remarks: str = ""  # 备注
    is_special: bool = False  # 是否为特殊项目
    special_reason: str = ""  # 特殊项目原因
    is_material_diff: bool = False  # 是否重大差异

    def __post_init__(self):
        if not self.confirmation_amount:
            self.confirmation_amount = self.book_amount

    @property
    def abs_difference(self) -> float:
        """差异绝对值"""
        if self.difference is None:
            return 0.0
        return abs(self.difference)

    @property
    def has_replied(self) -> bool:
        """是否已回函"""
        return self.reply_amount is not None


# ============================================================================
# 主引擎
# ============================================================================


class ConfirmationEngine:
    """函证管理主引擎

    负责往来账款函证的筛选、控制表生成、回函差异追踪全流程。

    Attributes:
        input_path: 输入Excel路径（往来明细）
        output_path: 输出Excel路径
        sheet_name: 工作表名
        name_col: 名称列名
        amount_col: 金额列名
        confirmation_type_key: 函证类型键 (ar/ap/other)
        threshold: 金额阈值，低于此值不函证（特殊项目除外）
        coverage: 覆盖率目标，如0.7表示覆盖70%余额
        min_items: 最少函证数量
        reply_path: 回函数据路径（可选）
        reply_name_col: 回函名称列名
        reply_amount_col: 回函金额列名
        materiality: 重要性水平，差异超过此值标红
        items: 所有往来项目
        selected_items: 筛选后的函证项目
    """

    def __init__(
        self,
        input_path: str,
        output_path: str,
        sheet_name: str = "Sheet1",
        name_col: str = "客户名称",
        amount_col: str = "期末余额",
        confirmation_type_key: str = "ar",
        threshold: float = DEFAULT_THRESHOLD,
        coverage: float = DEFAULT_COVERAGE,
        min_items: int = DEFAULT_MIN_ITEMS,
        reply_path: Optional[str] = None,
        reply_name_col: str = "客户名称",
        reply_amount_col: str = "回函金额",
        materiality: float = DEFAULT_MATERIALITY,
        special_col: Optional[str] = None,
        related_party_col: Optional[str] = None,
        transaction_freq_col: Optional[str] = None,
        abnormal_col: Optional[str] = None,
    ):
        """初始化函证引擎

        Args:
            input_path: 往来明细Excel文件路径
            output_path: 输出Excel文件路径
            sheet_name: 工作表名
            name_col: 客户/供应商名称所在列名
            amount_col: 金额所在列名
            confirmation_type_key: ar(应收)/ap(应付)/other(其他)
            threshold: 金额阈值，低于此值不纳入函证范围
            coverage: 期望函证覆盖率，范围(0, 1]
            min_items: 最少函证数量
            reply_path: 回函数据Excel路径（可选）
            reply_name_col: 回函数据中名称列名
            reply_amount_col: 回函数据中金额列名
            materiality: 重要性水平（差异超过此值标红）
            special_col: 特殊标记列名（可选）
            related_party_col: 关联方列名（可选）
            transaction_freq_col: 交易频率列名（可选）
            abnormal_col: 异常标记列名（可选）
        """
        self.input_path = input_path
        self.output_path = output_path
        self.sheet_name = sheet_name
        self.name_col = name_col
        self.amount_col = amount_col
        self.confirmation_type_key = confirmation_type_key
        self.threshold = threshold
        self.coverage = coverage
        self.min_items = min_items
        self.reply_path = reply_path
        self.reply_name_col = reply_name_col
        self.reply_amount_col = reply_amount_col
        self.materiality = materiality
        self.special_col = special_col
        self.related_party_col = related_party_col
        self.transaction_freq_col = transaction_freq_col
        self.abnormal_col = abnormal_col

        # 运行时状态
        self.items: List[ConfirmationItem] = []
        self.selected_items: List[ConfirmationItem] = []
        self.reply_data: Dict[str, Dict[str, Any]] = {}
        self.summary: Dict[str, Any] = {}
        self._prefix = CONFIRMATION_TYPE_PREFIX.get(confirmation_type_key, "CF")
        self._type_label = CONFIRMATION_TYPE_LABEL.get(confirmation_type_key, "往来款")

        # 验证参数
        self._validate_params()

    def _validate_params(self) -> None:
        """验证初始化参数"""
        if not os.path.exists(self.input_path):
            raise FileNotFoundError(f"输入文件不存在: {self.input_path}")
        if not 0 < self.coverage <= 1:
            raise ValueError(f"覆盖率必须介于(0, 1]，当前值: {self.coverage}")
        if self.min_items < 1:
            raise ValueError(f"最少函证数量必须>=1，当前值: {self.min_items}")
        if self.threshold < 0:
            raise ValueError(f"金额阈值不能为负数，当前值: {self.threshold}")
        if self.materiality < 0:
            raise ValueError(f"重要性水平不能为负数，当前值: {self.materiality}")
        if self.confirmation_type_key not in CONFIRMATION_TYPE_PREFIX:
            raise ValueError(
                f"函证类型无效: '{self.confirmation_type_key}'，"
                f"可选值: {list(CONFIRMATION_TYPE_PREFIX.keys())}"
            )

    # ------------------------------------------------------------------
    # 1. 数据加载
    # ------------------------------------------------------------------

    def load_data(self) -> None:
        """从Excel加载往来明细数据"""
        print(f"[加载] 正在读取: {self.input_path}")
        print(f"[加载] 工作表: '{self.sheet_name}'")

        try:
            wb = openpyxl.load_workbook(self.input_path, data_only=True)
        except Exception as e:
            raise RuntimeError(f"无法打开Excel文件: {self.input_path}") from e

        if self.sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            raise ValueError(
                f"工作表 '{self.sheet_name}' 不存在。可用工作表: {available}"
            )

        ws = wb[self.sheet_name]

        # 定位表头行（第一行默认为表头）
        headers, header_row = self._read_headers(ws)

        # 验证必要列
        self._validate_columns(headers, header_row)

        # 读取数据行
        self.items = self._parse_data_rows(ws, headers, header_row)
        wb.close()

        print(f"[加载] 共读取 {len(self.items)} 条往来项目")

        if len(self.items) == 0:
            print("[警告] 未找到有效数据行，请检查工作表内容")

    def _read_headers(
        self, ws: openpyxl.worksheet.worksheet.Worksheet
    ) -> Tuple[Dict[str, int], int]:
        """读取并解析表头

        Returns:
            (列名→列索引映射, 表头所在行号)
        """
        headers: Dict[str, int] = {}
        header_row = 1

        for row_idx in range(1, min(ws.max_row + 1, 10)):  # 前10行中找表头
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    cleaned = (
                        str(cell_value).strip().replace("\n", "").replace("\r", "")
                    )
                    if cleaned:
                        headers[cleaned] = col_idx
            if headers:
                header_row = row_idx
                break

        print(
            f"[表头] 第{header_row}行，共{len(headers)}列: {list(headers.keys())[:10]}..."
        )
        return headers, header_row

    def _validate_columns(self, headers: Dict[str, int], header_row: int) -> None:
        """验证必要列是否存在"""
        missing = []
        for col_name in [self.name_col, self.amount_col]:
            if col_name not in headers:
                missing.append(col_name)
        if missing:
            raise ValueError(
                f"第{header_row}行表头中未找到必要列: {missing}。"
                f"请使用 --name-col 和 --amount-col 指定正确的列名。"
                f"现有列: {list(headers.keys())}"
            )

    def _parse_data_rows(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        headers: Dict[str, int],
        header_row: int,
    ) -> List[ConfirmationItem]:
        """解析数据行，构建ConfirmationItem列表"""
        items: List[ConfirmationItem] = []
        name_col_idx = headers[self.name_col]
        amount_col_idx = headers[self.amount_col]

        # 可选列索引
        special_col_idx = headers.get(self.special_col) if self.special_col else None
        related_col_idx = (
            headers.get(self.related_party_col) if self.related_party_col else None
        )
        freq_col_idx = (
            headers.get(self.transaction_freq_col)
            if self.transaction_freq_col
            else None
        )
        abnormal_col_idx = headers.get(self.abnormal_col) if self.abnormal_col else None

        for row_idx in range(header_row + 1, ws.max_row + 1):
            name_cell = ws.cell(row=row_idx, column=name_col_idx).value
            amount_cell = ws.cell(row=row_idx, column=amount_col_idx).value

            # 跳过空行
            if name_cell is None:
                continue

            name = str(name_cell).strip()
            if not name:
                continue

            # 跳过合计行
            if any(kw in name for kw in ["合计", "总计", "小计", "平均"]):
                continue

            # 解析金额
            try:
                book_amount = float(amount_cell) if amount_cell is not None else 0.0
            except (ValueError, TypeError):
                print(f"[警告] 第{row_idx}行金额无法解析: {amount_cell}，跳过")
                continue

            # 特殊项目判定
            is_special = False
            special_reasons = []

            # 检查显式特殊标记列
            if special_col_idx:
                sv = ws.cell(row=row_idx, column=special_col_idx).value
                if sv and str(sv).strip():
                    is_special = True
                    special_reasons.append(str(sv).strip())

            # 检查关联方列
            if related_col_idx:
                rv = ws.cell(row=row_idx, column=related_col_idx).value
                if rv and str(rv).strip().lower() in ("是", "yes", "y", "1", "true"):
                    is_special = True
                    special_reasons.append("关联方")

            # 余额为零但交易频繁
            if freq_col_idx:
                fv = ws.cell(row=row_idx, column=freq_col_idx).value
                try:
                    freq = float(fv) if fv is not None else 0
                except (ValueError, TypeError):
                    freq = 0
                if book_amount == 0 and freq > 0:
                    is_special = True
                    special_reasons.append("零余额高频交易")

            # 异常标记列
            if abnormal_col_idx:
                av = ws.cell(row=row_idx, column=abnormal_col_idx).value
                if av and str(av).strip():
                    is_special = True
                    special_reasons.append(f"异常:{str(av).strip()}")

            items.append(
                ConfirmationItem(
                    index=row_idx,
                    name=name,
                    book_amount=book_amount,
                    is_special=is_special,
                    special_reason="; ".join(special_reasons)
                    if special_reasons
                    else "",
                    special_flag="; ".join(special_reasons) if special_reasons else "",
                )
            )

        return items

    # ------------------------------------------------------------------
    # 2. 函证项目筛选
    # ------------------------------------------------------------------

    def screen_items(self) -> None:
        """筛选函证项目

        筛选逻辑:
            1. 标记特殊项目（关联方、零余额高频交易、异常项目）
            2. 将特殊项目强制纳入函证范围（不论金额大小）
            3. 非特殊项目按金额阈值过滤
            4. 金额降序排列
            5. 累进选取直到满足覆盖率目标或达到最少数量
            6. 分配函证类型（积极式/消极式）
        """
        if not self.items:
            raise RuntimeError("请先调用 load_data() 加载数据")

        print(f"\n[筛选] 开始函证项目筛选")
        print(f"[筛选] 金额阈值: {self.threshold:,.2f}")
        print(f"[筛选] 目标覆盖率: {self.coverage:.0%}")
        print(f"[筛选] 最少数量: {self.min_items}")

        # 按金额降序排列
        sorted_items = sorted(
            self.items, key=lambda x: abs(x.book_amount), reverse=True
        )

        # 分离特殊项目（特殊项目不受阈值限制）
        specials = [it for it in sorted_items if it.is_special]
        normals = [it for it in sorted_items if not it.is_special]

        total_book = sum(abs(it.book_amount) for it in sorted_items)
        print(f"[筛选] 总余额: {total_book:,.2f}，特殊项目: {len(specials)}个")

        # 先纳入所有特殊项目
        selected: List[ConfirmationItem] = list(specials)
        selected_names = {it.name for it in specials}
        accumulated = sum(abs(it.book_amount) for it in specials)

        # 再按阈值+覆盖率选常规项目
        for item in normals:
            if item.name in selected_names:
                continue
            if len(selected) >= self.min_items and total_book > 0:
                current_coverage = accumulated / total_book
                if current_coverage >= self.coverage:
                    break
            # 阈值筛选（特殊项目已绕过）
            if (
                abs(item.book_amount) < self.threshold
                and len(selected) >= self.min_items
            ):
                break
            selected.append(item)
            selected_names.add(item.name)
            accumulated += abs(item.book_amount)

        # 如果仍不满足最少数量，强制追加
        if len(selected) < self.min_items:
            for item in normals:
                if item.name in selected_names:
                    continue
                selected.append(item)
                selected_names.add(item.name)
                accumulated += abs(item.book_amount)
                if len(selected) >= self.min_items:
                    break

        self.selected_items = selected

        # 分配函证类型和编号
        self._assign_confirmation_types()
        self._assign_confirmation_numbers()

        # 输出筛选结果
        actual_coverage = accumulated / total_book if total_book > 0 else 0
        positive = sum(1 for it in selected if it.confirmation_type == "积极式")
        negative = sum(1 for it in selected if it.confirmation_type == "消极式")

        print(f"[筛选] 选定 {len(selected)} 项")
        print(f"[筛选] 函证金额合计: {accumulated:,.2f}")
        print(f"[筛选] 实际覆盖率: {actual_coverage:.2%}")
        print(f"[筛选] 积极式: {positive}, 消极式: {negative}, 特殊: {len(specials)}")

    def _assign_confirmation_types(self) -> None:
        """分配函证类型（积极式/消极式）

        规则:
            - 特殊项目 → 积极式
            - 金额 > 阈值*2 → 积极式
            - 其余 → 消极式
        """
        for item in self.selected_items:
            if item.is_special:
                item.confirmation_type = "积极式"
            elif abs(item.book_amount) >= self.threshold * 2:
                item.confirmation_type = "积极式"
            else:
                item.confirmation_type = "消极式"

    def _assign_confirmation_numbers(self) -> None:
        """生成函证编号

        格式: {类型前缀}-{日期}-{序号:04d}
        例如: AR-20260603-0001
        """
        today = datetime.date.today().strftime("%Y%m%d")
        for i, item in enumerate(self.selected_items, 1):
            item.confirmation_number = f"{self._prefix}-{today}-{i:04d}"
            item.send_date = datetime.date.today().strftime("%Y-%m-%d")

    # ------------------------------------------------------------------
    # 3. 回函数据处理
    # ------------------------------------------------------------------

    def process_replies(self) -> None:
        """处理回函数据，计算差异"""
        if not self.reply_path:
            return
        if not self.selected_items:
            raise RuntimeError("请先调用 screen_items() 筛选项目")

        print(f"\n[回函] 正在读取回函数据: {self.reply_path}")
        if not os.path.exists(self.reply_path):
            print(f"[警告] 回函文件不存在，跳过差异分析: {self.reply_path}")
            return

        # 加载回函数据
        reply_map: Dict[str, float] = {}
        reply_date_map: Dict[str, str] = {}
        try:
            wb = openpyxl.load_workbook(self.reply_path, data_only=True)
            ws = wb.active

            # 找表头
            headers, header_row = self._read_headers(ws)
            name_idx = headers.get(self.reply_name_col)
            amount_idx = headers.get(self.reply_amount_col)

            if not name_idx:
                raise ValueError(
                    f"回函文件中未找到名称列'{self.reply_name_col}'。"
                    f"可用列: {list(headers.keys())}"
                )
            if not amount_idx:
                raise ValueError(
                    f"回函文件中未找到金额列'{self.reply_amount_col}'。"
                    f"可用列: {list(headers.keys())}"
                )

            # 尝试找回函日期列
            date_idx = None
            for possible_date_col in [
                "回函日期",
                "签复日期",
                "回复日期",
                "确认日期",
                "日期",
            ]:
                if possible_date_col in headers:
                    date_idx = headers[possible_date_col]
                    break

            for row_idx in range(header_row + 1, ws.max_row + 1):
                name = ws.cell(row=row_idx, column=name_idx).value
                amount = ws.cell(row=row_idx, column=amount_idx).value
                if name is None:
                    continue
                name_str = str(name).strip()
                if not name_str:
                    continue
                try:
                    reply_map[name_str] = float(amount) if amount is not None else 0.0
                except (ValueError, TypeError):
                    print(f"[警告] 回函第{row_idx}行金额无法解析: {amount}，跳过")
                if date_idx:
                    dv = ws.cell(row=row_idx, column=date_idx).value
                    if dv:
                        if isinstance(dv, datetime.datetime):
                            reply_date_map[name_str] = dv.strftime("%Y-%m-%d")
                        else:
                            reply_date_map[name_str] = str(dv)[:10]

            wb.close()
            print(f"[回函] 加载 {len(reply_map)} 条回函记录")
        except Exception as e:
            print(f"[错误] 读取回函数据失败: {e}")
            raise

        # 匹配回函到函证项目
        matched = 0
        for item in self.selected_items:
            if item.name in reply_map:
                item.reply_amount = reply_map[item.name]
                item.reply_date = reply_date_map.get(item.name)
                item.difference = item.book_amount - item.reply_amount
                if item.book_amount != 0:
                    item.diff_pct = round(
                        item.difference / abs(item.book_amount) * 100, 2
                    )
                else:
                    item.diff_pct = None if item.difference == 0 else float("inf")

                # 判定重大差异
                item.is_material_diff = (
                    self.materiality > 0 and abs(item.difference) > self.materiality
                )

                # 更新状态
                item.reconciliation_status = (
                    "已回函待调节" if item.difference != 0 else "已回函一致"
                )
                matched += 1
            else:
                item.reconciliation_status = "未回函"

        print(f"[回函] 匹配成功: {matched}/{len(self.selected_items)}")

    def _classify_diff_reasons(self) -> None:
        """对差异原因进行自动分类（可选扩展）

        当前版本使用简单分类规则，后续可按语义模型升级。
        """
        for item in self.selected_items:
            if item.difference is None or item.difference == 0:
                continue
            if item.diff_reason:
                continue  # 已有原因，不覆盖

            # 简单自动分类
            abs_diff = abs(item.difference)
            abs_book = abs(item.book_amount)
            if abs_book > 0 and abs_diff / abs_book < 0.01:
                item.diff_reason = "小额尾差"
            elif item.difference > 0:
                item.diff_reason = "待查-对方少记"
            else:
                item.diff_reason = "待查-对方多记"

    # ------------------------------------------------------------------
    # 4. 统计摘要
    # ------------------------------------------------------------------

    def generate_summary(self) -> Dict[str, Any]:
        """生成统计摘要

        Returns:
            统计字典，包含发函数量、覆盖率、回函率、差异统计等
        """
        if not self.items:
            raise RuntimeError("请先调用 load_data() 加载数据")

        total_book = sum(abs(it.book_amount) for it in self.items)
        selected_book = sum(abs(it.book_amount) for it in self.selected_items)
        positive_count = sum(
            1 for it in self.selected_items if it.confirmation_type == "积极式"
        )
        negative_count = sum(
            1 for it in self.selected_items if it.confirmation_type == "消极式"
        )
        special_count = sum(1 for it in self.selected_items if it.is_special)
        replied = [it for it in self.selected_items if it.has_replied]
        reply_count = len(replied)
        reply_rate = (
            reply_count / len(self.selected_items) if self.selected_items else 0
        )
        coverage_rate = selected_book / total_book if total_book > 0 else 0

        diffs = [it.difference for it in replied if it.difference is not None]
        total_diff = sum(diffs)
        material_diffs = sum(1 for it in replied if it.is_material_diff)
        avg_diff = total_diff / len(diffs) if diffs else 0

        self.summary = {
            "confirmation_type": self._type_label,
            "total_items": len(self.items),
            "selected_items": len(self.selected_items),
            "total_book_amount": total_book,
            "selected_book_amount": selected_book,
            "coverage_rate": round(coverage_rate * 100, 2),
            "positive_count": positive_count,
            "negative_count": negative_count,
            "special_count": special_count,
            "reply_count": reply_count,
            "reply_rate": round(reply_rate * 100, 2),
            "total_difference": round(total_diff, 2),
            "material_diff_count": material_diffs,
            "avg_difference": round(avg_diff, 2),
            "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        return self.summary

    # ------------------------------------------------------------------
    # 5. Excel输出
    # ------------------------------------------------------------------

    def export(self) -> str:
        """导出Excel（函证控制表 + 差异汇总 + 统计摘要）

        Returns:
            输出文件路径
        """
        if not self.selected_items:
            raise RuntimeError("请先调用 screen_items() 筛选项目")

        print(f"\n[导出] 正在生成: {self.output_path}")

        wb = openpyxl.Workbook()

        # 删除默认Sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Sheet1: 函证控制表
        self._write_control_sheet(wb)

        # Sheet2: 差异汇总（仅在有回函数据时）
        if any(it.has_replied for it in self.selected_items):
            self._write_diff_summary(wb)

        # Sheet3: 统计摘要
        self._write_summary_sheet(wb)

        # 保存
        wb.save(self.output_path)
        print(f"[导出] 完成: {self.output_path}")
        print(f"[导出] 工作表: {wb.sheetnames}")
        return self.output_path

    def _write_control_sheet(self, wb: openpyxl.Workbook) -> None:
        """写入函证控制表"""
        ws = wb.create_sheet("函证控制表")

        # 标题行
        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=len(CONTROL_SHEET_HEADERS),
        )
        title_cell = ws.cell(row=1, column=1, value=f"{self._type_label}函证控制表")
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER_ALIGN

        # 表头行（第2行）
        for col_idx, header in enumerate(CONTROL_SHEET_HEADERS, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # 数据行
        for row_idx, item in enumerate(self.selected_items, 1):
            row = row_idx + 2  # 从第3行开始
            data_values = [
                row_idx,
                item.name,
                item.book_amount,
                item.confirmation_amount,
                item.confirmation_type,
                item.confirmation_number,
                item.send_date or "",
                item.reply_date or "",
                item.reply_amount if item.reply_amount is not None else "",
                item.difference if item.difference is not None else "",
                item.diff_reason,
                item.reconciliation_status,
                item.alt_procedures,
                item.audit_conclusion,
                item.special_flag,
                item.remarks,
            ]

            for col_idx, value in enumerate(data_values, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                if col_idx in (1, 5, 6, 7, 8, 12, 13, 14, 15):
                    cell.alignment = CENTER_ALIGN
                else:
                    cell.alignment = LEFT_ALIGN
                # 金额格式
                if col_idx in (3, 4, 9, 10) and isinstance(value, (int, float)):
                    cell.number_format = NUM_FORMAT

                # 重大差异标红
                if col_idx == 10 and item.is_material_diff:
                    cell.font = RED_FONT

        # 冻结表头
        ws.freeze_panes = "A3"
        # 自动筛选
        ws.auto_filter.ref = f"A2:{get_column_letter(len(CONTROL_SHEET_HEADERS))}{len(self.selected_items) + 2}"
        # 列宽自适应
        self._auto_fit_columns(ws, len(self.selected_items) + 2)

    def _write_diff_summary(self, wb: openpyxl.Workbook) -> None:
        """写入差异汇总表"""
        ws = wb.create_sheet("差异汇总")

        # 标题
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=len(DIFF_SUMMARY_HEADERS)
        )
        title_cell = ws.cell(row=1, column=1, value=f"{self._type_label}回函差异汇总表")
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER_ALIGN

        # 表头（第2行）
        for col_idx, header in enumerate(DIFF_SUMMARY_HEADERS, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # 仅输出已回函且有差异的项目
        replied_with_diff = [
            it
            for it in self.selected_items
            if it.has_replied and it.difference is not None
        ]
        # 也包含已回函一致的（差异为0）
        all_replied = [it for it in self.selected_items if it.has_replied]
        # 按差异绝对值降序
        all_replied.sort(key=lambda x: abs(x.difference or 0), reverse=True)

        for row_idx, item in enumerate(all_replied, 1):
            row = row_idx + 2
            data_values = [
                row_idx,
                item.name,
                item.book_amount,
                item.reply_amount,
                item.difference,
                item.diff_pct if item.diff_pct is not None else "",
                item.diff_reason,
                "是" if item.is_material_diff else "否",
                item.remarks,
            ]

            for col_idx, value in enumerate(data_values, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                if col_idx in (1, 8):
                    cell.alignment = CENTER_ALIGN
                else:
                    cell.alignment = LEFT_ALIGN
                if col_idx in (3, 4, 5, 6) and isinstance(value, (int, float)):
                    cell.number_format = NUM_FORMAT

                # 重大差异标红
                if col_idx == 5 and item.is_material_diff:
                    cell.font = RED_FONT

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = (
            f"A2:{get_column_letter(len(DIFF_SUMMARY_HEADERS))}{len(all_replied) + 2}"
        )
        self._auto_fit_columns(ws, len(all_replied) + 2)

    def _write_summary_sheet(self, wb: openpyxl.Workbook) -> None:
        """写入统计摘要"""
        ws = wb.create_sheet("统计摘要")

        # 确保摘要已生成
        if not self.summary:
            self.generate_summary()

        # 标题
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        title_cell = ws.cell(row=1, column=1, value=f"{self._type_label}函证统计摘要")
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER_ALIGN

        # 参数信息
        param_header_font = Font(name="微软雅黑", size=11, bold=True, color="333333")
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
        ws.cell(row=3, column=1, value="━━━ 筛选参数 ━━━").font = param_header_font
        ws.cell(row=3, column=1).alignment = CENTER_ALIGN

        param_rows = [
            ("金额阈值", f"{self.threshold:,.2f}"),
            ("目标覆盖率", f"{self.coverage:.0%}"),
            ("最少数量", f"{self.min_items}"),
            ("重要性水平", f"{self.materiality:,.2f}"),
        ]
        for i, (label, value) in enumerate(param_rows, 4):
            cell_a = ws.cell(row=i, column=1, value=label)
            cell_a.font = Font(name="微软雅黑", size=10, bold=True)
            cell_a.alignment = LEFT_ALIGN
            cell_b = ws.cell(row=i, column=2, value=value)
            cell_b.font = DATA_FONT
            cell_b.alignment = LEFT_ALIGN

        # 统计信息
        start = 9
        ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=2)
        ws.cell(row=start, column=1, value="━━━ 函证统计 ━━━").font = param_header_font
        ws.cell(row=start, column=1).alignment = CENTER_ALIGN

        key_stats = [
            ("筛选项目总数", self.summary["total_items"]),
            ("发函数量", self.summary["selected_items"]),
            ("账面总金额", f"{self.summary['total_book_amount']:,.2f}"),
            ("函证金额合计", f"{self.summary['selected_book_amount']:,.2f}"),
            ("金额覆盖率", f"{self.summary['coverage_rate']}%"),
            ("积极式函证", self.summary["positive_count"]),
            ("消极式函证", self.summary["negative_count"]),
            ("特殊标记项目", self.summary["special_count"]),
        ]
        for i, (label, value) in enumerate(key_stats, start + 1):
            cell_a = ws.cell(row=i, column=1, value=label)
            cell_a.font = Font(name="微软雅黑", size=10, bold=True)
            cell_a.alignment = LEFT_ALIGN
            cell_b = ws.cell(row=i, column=2, value=value)
            cell_b.font = DATA_FONT
            cell_b.alignment = LEFT_ALIGN

        # 回函统计（如果有回函数据）
        if self.reply_path:
            start2 = start + 1 + len(key_stats) + 1
            ws.merge_cells(
                start_row=start2, start_column=1, end_row=start2, end_column=2
            )
            ws.cell(
                row=start2, column=1, value="━━━ 回函统计 ━━━"
            ).font = param_header_font
            ws.cell(row=start2, column=1).alignment = CENTER_ALIGN

            reply_stats = [
                ("回函数量", self.summary["reply_count"]),
                ("回函率", f"{self.summary['reply_rate']}%"),
                ("差异金额合计", f"{self.summary['total_difference']:,.2f}"),
                ("重大差异数量", self.summary["material_diff_count"]),
                ("平均差异金额", f"{self.summary['avg_difference']:,.2f}"),
            ]
            for i, (label, value) in enumerate(reply_stats, start2 + 1):
                cell_a = ws.cell(row=i, column=1, value=label)
                cell_a.font = Font(name="微软雅黑", size=10, bold=True)
                cell_a.alignment = LEFT_ALIGN
                cell_b = ws.cell(row=i, column=2, value=value)
                cell_b.font = DATA_FONT
                cell_b.alignment = LEFT_ALIGN

        # 生成时间
        last_row = ws.max_row + 2
        ws.cell(
            row=last_row, column=1, value=f"生成时间: {self.summary['generated_at']}"
        ).font = Font(name="微软雅黑", size=9, italic=True, color="888888")

        # 列宽
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 28

    # ------------------------------------------------------------------
    # 6. 格式化辅助
    # ------------------------------------------------------------------

    def _auto_fit_columns(
        self, ws: openpyxl.worksheet.worksheet.Worksheet, max_row: int
    ) -> None:
        """自适应列宽

        Args:
            ws: 工作表对象
            max_row: 最大数据行号
        """
        for col_idx in range(1, ws.max_column + 1):
            max_width = 0
            for row_idx in range(2, max_row + 2):  # 从表头行开始
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    # 估算中文字符宽度（中文约2个字符宽）
                    val_str = str(cell.value)
                    char_width = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    max_width = max(max_width, char_width)
            # 设置列宽，最小8，最大40
            width = min(max(max_width + 4, 8), 40)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ------------------------------------------------------------------
    # 7. 主流程
    # ------------------------------------------------------------------

    def run(self) -> str:
        """执行完整函证管理流程

        流程:
            load_data → screen_items → process_replies → generate_summary → export

        Returns:
            输出文件路径
        """
        print("=" * 60)
        print(f"  审计函证管理引擎 v1.0.0")
        print(f"  函证类型: {self._type_label}")
        print(f"  输入文件: {self.input_path}")
        print("=" * 60)

        try:
            # 第1步：加载数据
            self.load_data()

            # 第2步：筛选项目
            self.screen_items()

            # 第3步：处理回函（如有）
            if self.reply_path:
                self.process_replies()
                self._classify_diff_reasons()

            # 第4步：生成摘要
            self.generate_summary()

            # 第5步：导出Excel
            output = self.export()

            # 打印摘要
            self._print_summary()

            return output

        except Exception as e:
            print(f"\n[严重错误] 执行失败: {e}")
            traceback.print_exc()
            raise

    def _print_summary(self) -> None:
        """打印摘要到控制台"""
        s = self.summary
        print("\n" + "=" * 60)
        print("  执行摘要")
        print("=" * 60)
        print(f"  函证类型:     {s['confirmation_type']}")
        print(f"  筛选总数:     {s['total_items']} 项")
        print(f"  发函数量:     {s['selected_items']} 项")
        print(f"  金额覆盖率:   {s['coverage_rate']}%")
        print(f"  积极式:       {s['positive_count']} 份")
        print(f"  消极式:       {s['negative_count']} 份")
        print(f"  特殊标记:     {s['special_count']} 项")
        if self.reply_path:
            print(f"  回函数量:     {s['reply_count']} 份 (回函率 {s['reply_rate']}%)")
            print(f"  差异合计:     {s['total_difference']:,.2f}")
            print(f"  重大差异:     {s['material_diff_count']} 项")
        print(f"  输出文件:     {self.output_path}")
        print("=" * 60)


# ============================================================================
# CLI 入口
# ============================================================================


def build_parser() -> argparse.ArgumentParser:
    """构建命令行参数解析器"""
    parser = argparse.ArgumentParser(
        prog="confirmation_engine",
        description="审计函证管理引擎 — 符合《中国注册会计师审计准则第1312号——函证》",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 基础筛选：应收账款函证，阈值1万，覆盖70%
  python confirmation_engine.py --input 往来明细.xlsx --output 函证控制表.xlsx \\
      --name-col "客户名称" --amount-col "期末余额" --type ar --threshold 10000

  # 含回函差异分析
  python confirmation_engine.py --input 往来明细.xlsx --output 函证控制表.xlsx \\
      --name-col "供应商名称" --amount-col "应付余额" --type ap --threshold 5000 \\
      --replies 回函数据.xlsx --reply-amount-col "确认金额" --materiality 50000

  # 其他往来款函证 + 特殊标记列
  python confirmation_engine.py --input 往来明细.xlsx --output 函证控制表.xlsx \\
      --name-col "对方单位" --amount-col "余额" --type other --coverage 0.8 \\
      --special-col "特殊标记" --related-party-col "是否关联方" --min-items 20
        """,
    )

    # 必选参数
    parser.add_argument(
        "--input", "-i", required=True, help="输入Excel路径（往来明细表）"
    )
    parser.add_argument(
        "--output", "-o", required=True, help="输出Excel路径（函证控制表）"
    )

    # 列名配置
    parser.add_argument("--sheet", default="Sheet1", help="工作表名（默认: Sheet1）")
    parser.add_argument(
        "--name-col", default="客户名称", help="名称列名（默认: 客户名称）"
    )
    parser.add_argument(
        "--amount-col", default="期末余额", help="金额列名（默认: 期末余额）"
    )

    # 函证策略
    parser.add_argument(
        "--type",
        "-t",
        choices=["ar", "ap", "other"],
        default="ar",
        help="函证类型: ar=应收账款, ap=应付账款, other=其他往来（默认: ar）",
    )
    parser.add_argument(
        "--threshold",
        type=float,
        default=DEFAULT_THRESHOLD,
        help=f"金额阈值，低于此值不函证（默认: {DEFAULT_THRESHOLD}）",
    )
    parser.add_argument(
        "--coverage",
        "-c",
        type=float,
        default=DEFAULT_COVERAGE,
        help=f"目标覆盖率（默认: {DEFAULT_COVERAGE}，即70%%）",
    )
    parser.add_argument(
        "--min-items",
        "-m",
        type=int,
        default=DEFAULT_MIN_ITEMS,
        help=f"最少函证数量（默认: {DEFAULT_MIN_ITEMS}）",
    )

    # 回函相关
    parser.add_argument(
        "--replies", "-r", help="回函数据Excel路径（可选，用于差异分析）"
    )
    parser.add_argument(
        "--reply-name-col", default="客户名称", help="回函数据中名称列名"
    )
    parser.add_argument(
        "--reply-amount-col", default="回函金额", help="回函数据中金额列名"
    )
    parser.add_argument(
        "--materiality",
        type=float,
        default=DEFAULT_MATERIALITY,
        help=f"重要性水平，差异超过此值标红（默认: {DEFAULT_MATERIALITY}）",
    )

    # 特殊标记列
    parser.add_argument("--special-col", help="特殊标记列名（可选）")
    parser.add_argument(
        "--related-party-col", help="关联方标记列名（可选，值为'是'时标记）"
    )
    parser.add_argument(
        "--transaction-freq-col", help="交易频率列名（可选，零余额+高频→特殊）"
    )
    parser.add_argument("--abnormal-col", help="异常标记列名（可选）")

    return parser


def main(argv: Optional[List[str]] = None) -> int:
    """CLI主入口

    Args:
        argv: 命令行参数列表（None=使用sys.argv）

    Returns:
        退出码 (0=成功, 1=失败)
    """
    parser = build_parser()
    args = parser.parse_args(argv)

    try:
        engine = ConfirmationEngine(
            input_path=args.input,
            output_path=args.output,
            sheet_name=args.sheet,
            name_col=args.name_col,
            amount_col=args.amount_col,
            confirmation_type_key=args.type,
            threshold=args.threshold,
            coverage=args.coverage,
            min_items=args.min_items,
            reply_path=args.replies,
            reply_name_col=args.reply_name_col,
            reply_amount_col=args.reply_amount_col,
            materiality=args.materiality,
            special_col=args.special_col,
            related_party_col=args.related_party_col,
            transaction_freq_col=args.transaction_freq_col,
            abnormal_col=args.abnormal_col,
        )
        engine.run()
        return 0
    except FileNotFoundError as e:
        print(f"[错误] 文件不存在: {e}", file=sys.stderr)
        return 1
    except ValueError as e:
        print(f"[错误] 参数无效: {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"[错误] 执行异常: {e}", file=sys.stderr)
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

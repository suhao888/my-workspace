# -*- coding: utf-8 -*-
import sys
import os

filepath = r"D:\Users\12844\Desktop\10-梳理-----0528\基础数据\5、浙江油田-附件1：合同结算流程调查表（浙江油田分公司）.et"
print(f"文件路径: {filepath}")

# 先检查文件头
with open(filepath, 'rb') as f:
    header = f.read(16)
    print(f"文件头(hex): {header.hex()}")
    print(f"文件头(raw): {header}")

fsize = os.path.getsize(filepath)
print(f"文件大小: {fsize} bytes")

# 判断文件类型
if header[:4] == b'\xd0\xcf\x11\xe0':
    print("文件类型: OLE2 (老版Excel/ET格式)")
elif header[:4] == b'PK\x03\x04':
    print("文件类型: ZIP (xlsx/新版ET格式)")
else:
    print(f"文件类型: 未知, 前4字节={header[:4].hex()}")

# 方法1: xlrd (支持xls格式)
print("\n===== 方法1: xlrd =====")
try:
    import xlrd
    wb = xlrd.open_workbook(filepath)
    print(f"Sheet数量: {wb.nsheets}")
    for idx in range(wb.nsheets):
        ws = wb.sheet_by_index(idx)
        print(f"\nSheet {idx}: {ws.name}, 行数={ws.nrows}, 列数={ws.ncols}")
        for r in range(min(10, ws.nrows)):
            row_data = [ws.cell_value(r, c) for c in range(ws.ncols)]
            print(f"  行{r}: {row_data}")
except Exception as e:
    print(f"xlrd失败: {type(e).__name__}: {e}")

# 方法2: openpyxl (支持xlsx格式)
print("\n===== 方法2: openpyxl =====")
try:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(f"Sheet名称: {wb.sheetnames}")
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\nSheet: {sname}, 行数={ws.max_row}, 列数={ws.max_column}")
        for r_idx, row in enumerate(ws.iter_rows(max_row=min(10, ws.max_row), values_only=True)):
            print(f"  行{r_idx}: {list(row)}")
except Exception as e:
    print(f"openpyxl失败: {type(e).__name__}: {e}")

# 方法3: xlrd with formatting_info
print("\n===== 方法3: xlrd (formatting_info=True) =====")
try:
    import xlrd
    wb = xlrd.open_workbook(filepath, formatting_info=True)
    print(f"Sheet数量: {wb.nsheets}")
    for idx in range(wb.nsheets):
        ws = wb.sheet_by_index(idx)
        print(f"\nSheet {idx}: {ws.name}, 行数={ws.nrows}, 列数={ws.ncols}")
        for r in range(min(10, ws.nrows)):
            row_data = [ws.cell_value(r, c) for c in range(ws.ncols)]
            print(f"  行{r}: {row_data}")
except Exception as e:
    print(f"xlrd+formatting失败: {type(e).__name__}: {e}")

# 方法4: olefile 解析OLE结构
print("\n===== 方法4: olefile =====")
try:
    import olefile
    ole = olefile.OleFileIO(filepath)
    print(f"OLE streams: {ole.listdir()}")
    for stream_path in ole.listdir():
        stream_name = '/'.join(stream_path)
        size = ole.get_size(stream_name)
        print(f"  Stream: {stream_name}, size={size}")
    ole.close()
except Exception as e:
    print(f"olefile失败: {type(e).__name__}: {e}")

# 方法5: 如果是ZIP格式,尝试解压读取XML
print("\n===== 方法5: ZIP解压 =====")
try:
    import zipfile
    if zipfile.is_zipfile(filepath):
        with zipfile.ZipFile(filepath, 'r') as zf:
            print(f"ZIP内容: {zf.namelist()}")
            for name in zf.namelist():
                print(f"  {name}, size={zf.getinfo(name).file_size}")
    else:
        print("不是ZIP格式文件")
except Exception as e:
    print(f"ZIP解压失败: {type(e).__name__}: {e}")

print("\n===== 完成 =====")

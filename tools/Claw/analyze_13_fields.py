"""
分析13家油田合同结算流程调查表 - 提取核心差异指标
"""
import os, sys, re
import xlrd

base = "D:/Users/12844/Desktop/10-梳理-----0528/基础数据"
files = sorted(os.listdir(base))

# 油田名称映射
name_map = {}
for f in files:
    num = f.split('、')[0] if '、' in f else f[:2]
    name = f.split('、')[1] if '、' in f else f
    # 提取油田名
    for kw in ['大港', '辽河', '西南', '塔里木', '浙江', '玉门', '青海', '大庆', '吉林', '长庆', '冀东', '华北', '新疆']:
        if kw in f:
            name_map[f] = kw + '油田'
            break
    else:
        name_map[f] = f[:20]

results = {}

for f in files:
    fpath = os.path.join(base, f)
    ext = os.path.splitext(f)[1].lower()
    name = name_map.get(f, f)
    
    print(f"\n{'='*80}")
    print(f"【{name}】文件: {f}")
    
    if ext == '.et':
        print("  [跳过ET格式]")
        continue
    
    try:
        if ext == '.xls':
            wb = xlrd.open_workbook(fpath)
            sheets = wb.sheet_names()
        elif ext == '.xlsx':
            import openpyxl
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            sheets = wb.sheetnames
        else:
            continue
            
        main_sheet = sheets[0]
        
        if ext == '.xls':
            sh = wb.sheet_by_name(main_sheet)
            nrows, ncols = sh.nrows, sh.ncols
            
            # 读取所有数据行（从行5开始）
            contract_types = set()
            platforms = set()
            payment_tiers = set()
            approval_depths = []  # 付款审批链长度
            has_parallel = False
            internal_settle_special = False
            online_offline = set()
            
            for r in range(5, nrows):
                row = [str(sh.cell_value(r, c)).strip() for c in range(ncols)]
                
                # 一级类别 (col 1)
                if row[1] and row[1] != 'nan':
                    contract_types.add(row[1][:30])
                
                # 操作平台 (col 9)
                if row[9] and row[9] != 'nan':
                    platforms.add(row[9][:50])
                
                # 付款审批金额档次 (col 22)
                if row[22] and row[22] != 'nan' and row[22] != '':
                    payment_tiers.add(row[22][:60])
                
                # 付款审批流程长度 (col 23)
                if row[23] and row[23] != 'nan' and '→' in row[23]:
                    chain = row[23].split('→')
                    approval_depths.append(len(chain))
                
                # 合同签订是否有并行 (col 5)
                if '并行' in row[5]:
                    has_parallel = True
                
                # 内部/关联交易特殊审批 (col 16-17)
                combined = row[16] + row[17]
                if '关联交易' in combined or '内部合同' in combined or '内部结算' in combined:
                    internal_settle_special = True
                
                # 线上OR线下 (col 26 或附近)
                for c in [26, 27, 30]:
                    if c < ncols and row[c] and '线下' in row[c]:
                        online_offline.add('线下')
                    if c < ncols and row[c] and '线上' in row[c]:
                        online_offline.add('线上')
            
            avg_depth = sum(approval_depths) / len(approval_depths) if approval_depths else 0
            
            print(f"  数据行数: {nrows - 5}")
            print(f"  合同类别数: {len(contract_types)} → {sorted(contract_types)}")
            print(f"  操作平台: {platforms}")
            print(f"  付款审批金额档次 ({len(payment_tiers)}档): {sorted(payment_tiers)[:10]}")
            print(f"  付款审批平均层级: {avg_depth:.1f} (样本{len(approval_depths)}条)")
            print(f"  合同签订有并行审查: {has_parallel}")
            print(f"  内部/关联交易有特殊流程: {internal_settle_special}")
            print(f"  线上线下: {online_offline}")
            
            # 特别分析：检查"内部结算无审批"等风险点
            for r in range(5, nrows):
                row = [str(sh.cell_value(r, c)).strip() for c in range(ncols)]
                combined = ''.join(row)
                if '无审批' in combined or '自动生成' in combined:
                    print(f"  ⚠️ 风险发现(行{r}): {row[1]} - {row[2]} : 包含'无审批/自动生成'")
                if '自行简化' in combined:
                    print(f"  ⚠️ 风险发现(行{r}): 包含'自行简化审批'")
            
        else:  # xlsx
            sh = wb[main_sheet]
            rows_data = []
            for row in sh.iter_rows(values_only=True):
                rows_data.append([str(v).strip() if v is not None else '' for v in row])
            
            nrows = len(rows_data)
            contract_types = set()
            platforms = set()
            payment_tiers = set()
            approval_depths = []
            has_parallel = False
            internal_settle_special = False
            online_offline = set()
            
            for r in range(5, nrows):
                row = rows_data[r]
                ncols_row = len(row)
                
                if ncols_row > 1 and row[1] and row[1] != 'None':
                    contract_types.add(row[1][:30])
                
                if ncols_row > 9 and row[9] and row[9] != 'None':
                    platforms.add(row[9][:50])
                
                if ncols_row > 22 and row[22] and row[22] != 'None' and row[22] != '':
                    payment_tiers.add(row[22][:60])
                
                if ncols_row > 23 and row[23] and row[23] != 'None' and '→' in row[23]:
                    chain = row[23].split('→')
                    approval_depths.append(len(chain))
                
                if ncols_row > 5 and '并行' in row[5]:
                    has_parallel = True
                
                combined = ''
                if ncols_row > 17:
                    combined = row[16] + row[17] if ncols_row > 17 else row[16] if ncols_row > 16 else ''
                if '关联交易' in combined or '内部合同' in combined or '内部结算' in combined:
                    internal_settle_special = True
                
                for c in [26, 27, 30]:
                    if c < ncols_row and row[c] and '线下' in row[c]:
                        online_offline.add('线下')
                    if c < ncols_row and row[c] and '线上' in row[c]:
                        online_offline.add('线上')
                
                combined_all = ''.join(row)
                if '无审批' in combined_all or '自动生成' in combined_all:
                    print(f"  ⚠️ 风险发现(行{r}): {row[1] if ncols_row>1 else ''} : 包含'无审批/自动生成'")
                if '自行简化' in combined_all:
                    print(f"  ⚠️ 风险发现(行{r}): 包含'自行简化审批'")
            
            avg_depth = sum(approval_depths) / len(approval_depths) if approval_depths else 0
            
            print(f"  数据行数: {nrows - 5}")
            print(f"  合同类别数: {len(contract_types)} → {sorted(contract_types)}")
            print(f"  操作平台: {platforms}")
            print(f"  付款审批金额档次 ({len(payment_tiers)}档): {sorted(payment_tiers)[:10]}")
            print(f"  付款审批平均层级: {avg_depth:.1f} (样本{len(approval_depths)}条)")
            print(f"  合同签订有并行审查: {has_parallel}")
            print(f"  内部/关联交易有特殊流程: {internal_settle_special}")
            print(f"  线上线下: {online_offline}")
        
        if ext == '.xlsx':
            wb.close()
            
    except Exception as e:
        print(f"  错误: {e}")

print("\n\n" + "="*80)
print("分析完成!")

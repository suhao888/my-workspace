# -*- coding: utf-8 -*-
"""
资金流水来源与去向统计表生成脚本
从4份已整理的银行/微信流水中提取交易数据，按对手方统计资金来源（收入）和资金去向（支出）。
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from collections import defaultdict, OrderedDict
import os

# ============================================================
# 文件配置
# ============================================================
BASE_DIR = "G:/老联想电脑资料/苏浩个人/原工作/铜峰会计师事务所/侯红/侯红审计报告及附件9.12添加数据修改"

FILES = [
    {
        "name": "李凯",
        "path": os.path.join(BASE_DIR, "5、李凯微信银行流水.xlsx"),
        "sheet": "李凯银行、微信流水",
        "amount_col": "进-正数",   # 李凯用"进-正数"/"出-负数"
        "amount_out_col": "出-负数",
        "dir_col": "进/出",
        "counterparty_col": "对手户名",
        "category_col": "交易类别",
        "date_col": "日期",
        "channel_col": "渠道",
        "summary_sheet": "分类汇总-结果-1",
        "flow_sheet": None,
    },
    {
        "name": "侯红",
        "path": os.path.join(BASE_DIR, "2、侯红微信银行流水.xlsx"),
        "sheet": "侯红银行、微信流水",
        "amount_col": "进",
        "amount_out_col": "出（负数）",
        "dir_col": "进/出",
        "counterparty_col": "对手户名",
        "category_col": "收支类别",
        "date_col": "日期",
        "channel_col": "渠道",
        "summary_sheet": "分类汇总-结果",
        "flow_sheet": None,
    },
    {
        "name": "尚喜营",
        "path": os.path.join(BASE_DIR, "3、尚喜营微信银行流水.xlsx"),
        "sheet": "尚喜营银行、微信流水",
        "amount_col": "进-正数",
        "amount_out_col": "出-负数",
        "dir_col": "进/出",
        "counterparty_col": "对手户名",
        "category_col": "交易类别",
        "date_col": "日期",
        "channel_col": "渠道",
        "summary_sheet": "分类汇总-结果-1",
        "flow_sheet": "尚喜营主要资金流向",
    },
    {
        "name": "尚明豪",
        "path": os.path.join(BASE_DIR, "4、尚明豪微信银行流水.xlsx"),
        "sheet": "尚明豪银行微信流水",
        "amount_col": "进-正数",
        "amount_out_col": "出-负数",
        "dir_col": "进/出",
        "counterparty_col": "对手户名",
        "category_col": "交易类别",
        "date_col": "日期",
        "channel_col": "渠道",
        "summary_sheet": None,
        "flow_sheet": "尚明豪资金流向",
    },
]

# 家庭成员名单（内部互转标记用）
FAMILY_MEMBERS = ["侯红", "尚喜营", "尚明豪", "李凯"]

# ============================================================
# 数据读取
# ============================================================
def clean_str(val):
    """清洗字符串值"""
    if val is None:
        return ""
    s = str(val).strip().replace("\t", "").replace("\n", "").replace("\r", "")
    return s

def parse_amount(val):
    """解析金额"""
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(",", "").replace(" ", "").replace("\t", ""))
    except (ValueError, TypeError):
        return 0.0

def read_transactions(cfg):
    """从流水sheet读取所有交易记录"""
    wb = openpyxl.load_workbook(cfg["path"], data_only=True, read_only=True)
    ws = wb[cfg["sheet"]]

    # 获取列索引
    headers = {}
    for i, cell in enumerate(ws[1], 1):
        val = clean_str(cell.value)
        if val:
            headers[val] = i

    transactions = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        direction = clean_str(row[headers.get(cfg["dir_col"], 0) - 1].value) if cfg["dir_col"] in headers else ""
        amount_in = 0.0
        amount_out = 0.0

        if direction == "进":
            amount_in = abs(parse_amount(row[headers.get(cfg["amount_col"], 0) - 1].value)) if cfg["amount_col"] in headers else 0.0
        elif direction == "出":
            amount_out = abs(parse_amount(row[headers.get(cfg["amount_out_col"], 0) - 1].value)) if cfg["amount_out_col"] in headers else 0.0

        counterparty = clean_str(row[headers.get(cfg["counterparty_col"], 0) - 1].value) if cfg["counterparty_col"] in headers else ""
        category = clean_str(row[headers.get(cfg["category_col"], 0) - 1].value) if cfg["category_col"] in headers else ""
        date_val = row[headers.get(cfg["date_col"], 0) - 1].value if cfg["date_col"] in headers else None
        channel = clean_str(row[headers.get(cfg["channel_col"], 0) - 1].value) if cfg["channel_col"] in headers else ""

        # 跳过无金额记录
        if amount_in == 0 and amount_out == 0:
            continue

        transactions.append({
            "direction": direction,
            "amount_in": amount_in,
            "amount_out": amount_out,
            "counterparty": counterparty,
            "category": category,
            "date": date_val,
            "channel": channel,
            "holder": cfg["name"],
        })

    wb.close()
    return transactions

def read_existing_summary(cfg):
    """读取已有的分类汇总表"""
    if not cfg["summary_sheet"]:
        return []
    wb = openpyxl.load_workbook(cfg["path"], data_only=True, read_only=True)
    ws = wb[cfg["summary_sheet"]]

    # 尝试检测格式
    # 格式1: 对手户名 | 进-正数 | 出-负数 | 净额 | 涉及人员 | 备注
    # 格式2: 渠道 | 卡类型 | 开户行 | 账户名称 | 账号 | 收支类别 | 交易金额
    headers = [clean_str(cell.value) for cell in ws[1]]
    summary_rows = []

    if "对手户名" in headers:
        # 格式1: 按对手方汇总
        cp_idx = headers.index("对手户名")
        in_idx = headers.index("进-正数") if "进-正数" in headers else -1
        out_idx = headers.index("出-负数") if "出-负数" in headers else -1
        net_idx = headers.index("净额") if "净额" in headers else -1
        remark_idx = headers.index("涉及人员") if "涉及人员" in headers else -1

        for row in ws.iter_rows(min_row=2, values_only=True):
            cp = clean_str(row[cp_idx])
            if not cp:
                continue
            amount_in = parse_amount(row[in_idx]) if in_idx >= 0 and in_idx < len(row) else 0.0
            amount_out = abs(parse_amount(row[out_idx])) if out_idx >= 0 and out_idx < len(row) else 0.0
            net = parse_amount(row[net_idx]) if net_idx >= 0 and net_idx < len(row) else 0.0
            remark = clean_str(row[remark_idx]) if remark_idx >= 0 and remark_idx < len(row) else ""
            summary_rows.append({
                "counterparty": cp,
                "amount_in": amount_in,
                "amount_out": amount_out,
                "net": net,
                "remark": remark,
            })

    elif "收支类别" in headers:
        # 格式2: 按收支类别汇总
        cat_idx = headers.index("收支类别")
        amt_idx = headers.index("交易金额") if "交易金额" in headers else -1
        ch_idx = headers.index("渠道") if "渠道" in headers else -1

        for row in ws.iter_rows(min_row=2, values_only=True):
            cat = clean_str(row[cat_idx])
            if not cat:
                continue
            amt = abs(parse_amount(row[amt_idx])) if amt_idx >= 0 and amt_idx < len(row) else 0.0
            ch = clean_str(row[ch_idx]) if ch_idx >= 0 and ch_idx < len(row) else ""
            summary_rows.append({
                "category": cat,
                "amount": amt,
                "channel": ch,
            })

    wb.close()
    return summary_rows

def read_existing_flow(cfg):
    """读取已有的资金流向表（按对手方汇总）"""
    if not cfg.get("flow_sheet"):
        return []
    wb = openpyxl.load_workbook(cfg["path"], data_only=True, read_only=True)
    ws = wb[cfg["flow_sheet"]]

    headers = [clean_str(cell.value) for cell in ws[1]]
    if "对手户名" not in headers:
        wb.close()
        return []

    cp_idx = headers.index("对手户名")
    in_idx = headers.index("进-正数") if "进-正数" in headers else -1
    out_idx = headers.index("出-负数") if "出-负数" in headers else -1
    net_idx = headers.index("净额") if "净额" in headers else -1
    remark_idx = -1
    for i, h in enumerate(headers):
        if h and "涉及" in h:
            remark_idx = i
            break

    flow_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cp = clean_str(row[cp_idx])
        if not cp:
            continue
        amount_in = parse_amount(row[in_idx]) if in_idx >= 0 and in_idx < len(row) else 0.0
        amount_out = abs(parse_amount(row[out_idx])) if out_idx >= 0 and out_idx < len(row) else 0.0
        net = parse_amount(row[net_idx]) if net_idx >= 0 and net_idx < len(row) else 0.0
        remark = clean_str(row[remark_idx]) if remark_idx >= 0 and remark_idx < len(row) else ""
        flow_rows.append({
            "counterparty": cp,
            "amount_in": amount_in,
            "amount_out": amount_out,
            "net": net,
            "remark": remark,
        })

    wb.close()
    return flow_rows

# ============================================================
# 统计分析
# ============================================================
def aggregate_by_counterparty(transactions):
    """按对手方聚合收入/支出"""
    stats = defaultdict(lambda: {"in": 0.0, "out": 0.0, "count_in": 0, "count_out": 0, "channel": set()})
    for t in transactions:
        cp = t["counterparty"]
        if not cp:
            cp = "(无对手方)"
        if t["amount_in"] > 0:
            stats[cp]["in"] += t["amount_in"]
            stats[cp]["count_in"] += 1
        if t["amount_out"] > 0:
            stats[cp]["out"] += t["amount_out"]
            stats[cp]["count_out"] += 1
        if t["channel"]:
            stats[cp]["channel"].add(t["channel"])
    return stats

def aggregate_by_category(transactions):
    """按收支类别聚合"""
    stats = defaultdict(lambda: {"in": 0.0, "out": 0.0, "count_in": 0, "count_out": 0})
    for t in transactions:
        cat = t["category"] if t["category"] else "(未分类)"
        if t["amount_in"] > 0:
            stats[cat]["in"] += t["amount_in"]
            stats[cat]["count_in"] += 1
        if t["amount_out"] > 0:
            stats[cat]["out"] += t["amount_out"]
            stats[cat]["count_out"] += 1
    return stats

def aggregate_by_channel(transactions):
    """按渠道聚合"""
    stats = defaultdict(lambda: {"in": 0.0, "out": 0.0, "count_in": 0, "count_out": 0})
    for t in transactions:
        ch = t["channel"] if t["channel"] else "(未知)"
        if t["amount_in"] > 0:
            stats[ch]["in"] += t["amount_in"]
            stats[ch]["count_in"] += 1
        if t["amount_out"] > 0:
            stats[ch]["out"] += t["amount_out"]
            stats[ch]["count_out"] += 1
    return stats

# ============================================================
# 样式定义
# ============================================================
FONT_TITLE = Font(name="微软雅黑", size=14, bold=True)
FONT_HEADER = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
FONT_NORMAL = Font(name="微软雅黑", size=10)
FONT_BOLD = Font(name="微软雅黑", size=10, bold=True)
FONT_RED = Font(name="微软雅黑", size=10, color="FF0000")
FONT_GREEN = Font(name="微软雅黑", size=10, color="008000")
FONT_SUM = Font(name="微软雅黑", size=10, bold=True, color="003366")

FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_LIGHT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_SUM = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FAMILY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def format_number(ws, row, col, value):
    """写入带千分位的数字"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = '#,##0.00'
    cell.alignment = ALIGN_RIGHT
    return cell

def apply_border_range(ws, min_row, max_row, min_col, max_col):
    """给范围加边框"""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = BORDER_THIN

# ============================================================
# 生成Excel
# ============================================================
def generate_report(all_data, output_path):
    wb = openpyxl.Workbook()

    # --------------------------------------------------------
    # Sheet 1: 总览
    # --------------------------------------------------------
    ws1 = wb.active
    ws1.title = "资金流水总览"
    ws1.sheet_properties.tabColor = "4472C4"

    # 标题
    ws1.merge_cells("A1:I1")
    ws1.cell(1, 1, "资金流水统计总览").font = FONT_TITLE
    ws1.cell(1, 1).alignment = ALIGN_CENTER

    # 总览表头
    overview_headers = ["姓名", "交易总笔数", "收入笔数", "支出笔数",
                        "收入合计(元)", "支出合计(元)", "净额(元)",
                        "涉及对手方数", "主要渠道"]
    for i, h in enumerate(overview_headers, 1):
        c = ws1.cell(row=3, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER

    row = 4
    grand_total_in = 0
    grand_total_out = 0
    for holder_info in all_data:
        name = holder_info["name"]
        txns = holder_info["transactions"]
        total_in = sum(t["amount_in"] for t in txns)
        total_out = sum(t["amount_out"] for t in txns)
        count_in = sum(1 for t in txns if t["amount_in"] > 0)
        count_out = sum(1 for t in txns if t["amount_out"] > 0)
        cp_stats = holder_info["cp_stats"]
        ch_stats = holder_info["ch_stats"]
        channels = ", ".join(sorted(ch_stats.keys()))
        if len(channels) > 20:
            channels = channels[:20] + "..."

        grand_total_in += total_in
        grand_total_out += total_out

        ws1.cell(row, 1, name).font = FONT_BOLD
        ws1.cell(row, 1).alignment = ALIGN_CENTER
        ws1.cell(row, 2, len(txns)).alignment = ALIGN_CENTER
        ws1.cell(row, 3, count_in).alignment = ALIGN_CENTER
        ws1.cell(row, 4, count_out).alignment = ALIGN_CENTER
        format_number(ws1, row, 5, total_in)
        format_number(ws1, row, 6, total_out)
        net_cell = format_number(ws1, row, 7, total_in - total_out)
        if total_in - total_out < 0:
            net_cell.font = FONT_RED
        elif total_in - total_out > 0:
            net_cell.font = FONT_GREEN
        ws1.cell(row, 8, len(cp_stats)).alignment = ALIGN_CENTER
        ws1.cell(row, 9, channels).font = FONT_NORMAL
        ws1.cell(row, 9).alignment = ALIGN_LEFT

        # 交替行底色
        if (row - 4) % 2 == 1:
            for c in range(1, 10):
                ws1.cell(row, c).fill = FILL_LIGHT
        row += 1

    # 合计行
    for c in range(1, 10):
        ws1.cell(row, c).fill = FILL_SUM
        ws1.cell(row, c).font = FONT_SUM
        ws1.cell(row, c).alignment = ALIGN_CENTER
    ws1.cell(row, 1, "合计")
    ws1.cell(row, 2, sum(len(h["transactions"]) for h in all_data))
    ws1.cell(row, 3, sum(sum(1 for t in h["transactions"] if t["amount_in"] > 0) for h in all_data))
    ws1.cell(row, 4, sum(sum(1 for t in h["transactions"] if t["amount_out"] > 0) for h in all_data))
    format_number(ws1, row, 5, grand_total_in)
    format_number(ws1, row, 6, grand_total_out)
    format_number(ws1, row, 7, grand_total_in - grand_total_out)

    apply_border_range(ws1, 3, row, 1, 9)

    # 设置列宽
    col_widths = [12, 12, 10, 10, 18, 18, 18, 14, 25]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # --------------------------------------------------------
    # Sheet 2-5: 各人资金来源与去向
    # --------------------------------------------------------
    for holder_info in all_data:
        name = holder_info["name"]
        txns = holder_info["transactions"]
        cp_stats = holder_info["cp_stats"]
        cat_stats = holder_info["cat_stats"]
        ch_stats = holder_info["ch_stats"]

        # 排序：先按净额绝对值降序
        sorted_cp = sorted(cp_stats.items(), key=lambda x: abs(x[1]["in"] + x[1]["out"]), reverse=True)

        sheet_name = f"{name}-来源去向"
        ws = wb.create_sheet(title=sheet_name[:31])

        # 标题
        ws.merge_cells("A1:H1")
        ws.cell(1, 1, f"{name} — 资金来源与去向统计").font = FONT_TITLE
        ws.cell(1, 1).alignment = ALIGN_CENTER

        # 汇总行
        total_in = sum(t["amount_in"] for t in txns)
        total_out = sum(t["amount_out"] for t in txns)
        count_in = sum(1 for t in txns if t["amount_in"] > 0)
        count_out = sum(1 for t in txns if t["amount_out"] > 0)

        ws.cell(2, 1, f"收入合计: {total_in:,.2f} 元（{count_in}笔）  |  支出合计: {total_out:,.2f} 元（{count_out}笔）  |  净额: {total_in - total_out:,.2f} 元").font = Font(name="微软雅黑", size=10, color="333333")

        # --- 上半部分：资金来源（按对手方）---
        current_row = 4
        ws.merge_cells(f"A{current_row}:H{current_row}")
        ws.cell(current_row, 1, "一、资金来源（收入）— 按对手方汇总").font = Font(name="微软雅黑", size=11, bold=True, color="008000")
        current_row += 1

        source_headers = ["序号", "对手方", "是否家庭成员", "收入金额(元)", "收入笔数", "占收入比例", "渠道", "备注"]
        for i, h in enumerate(source_headers, 1):
            c = ws.cell(row=current_row, column=i, value=h)
            c.font = FONT_HEADER
            c.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
            c.alignment = ALIGN_CENTER
        current_row += 1

        # 只取有收入的对手方
        income_cps = [(cp, v) for cp, v in sorted_cp if v["in"] > 0]
        income_cps.sort(key=lambda x: x[1]["in"], reverse=True)
        seq = 0
        for cp, v in income_cps:
            seq += 1
            is_family = "是" if cp in FAMILY_MEMBERS else ""
            channels = ", ".join(sorted(v["channel"]))

            fill = FAMILY_FILL if cp in FAMILY_MEMBERS else (FILL_WHITE if seq % 2 == 0 else FILL_LIGHT)

            ws.cell(current_row, 1, seq).alignment = ALIGN_CENTER
            ws.cell(current_row, 1).fill = fill
            ws.cell(current_row, 1).font = FONT_NORMAL

            ws.cell(current_row, 2, cp).fill = fill
            ws.cell(current_row, 2).font = FONT_BOLD if cp in FAMILY_MEMBERS else FONT_NORMAL
            ws.cell(current_row, 2).alignment = ALIGN_LEFT

            ws.cell(current_row, 3, is_family).fill = fill
            ws.cell(current_row, 3).alignment = ALIGN_CENTER
            if is_family:
                ws.cell(current_row, 3).font = FONT_BOLD

            format_number(ws, current_row, 4, v["in"]).fill = fill

            ws.cell(current_row, 5, v["count_in"]).fill = fill
            ws.cell(current_row, 5).alignment = ALIGN_CENTER

            pct = v["in"] / total_in if total_in > 0 else 0
            pct_cell = ws.cell(current_row, 6, pct)
            pct_cell.number_format = '0.00%'
            pct_cell.alignment = ALIGN_CENTER
            pct_cell.fill = fill

            ws.cell(current_row, 7, channels).fill = fill
            ws.cell(current_row, 7).alignment = ALIGN_LEFT
            ws.cell(current_row, 7).font = FONT_NORMAL

            ws.cell(current_row, 8, "").fill = fill

            current_row += 1

        # 收入合计行
        for c in range(1, 9):
            ws.cell(current_row, c).fill = FILL_SUM
            ws.cell(current_row, c).font = FONT_SUM
            ws.cell(current_row, c).alignment = ALIGN_CENTER
        ws.cell(current_row, 1, "")
        ws.cell(current_row, 2, "收入合计")
        ws.cell(current_row, 3, "")
        format_number(ws, current_row, 4, total_in)
        ws.cell(current_row, 5, count_in)
        ws.cell(current_row, 6, 1.0).number_format = '0.00%'
        current_row += 2

        # --- 下半部分：资金去向（支出）— 按对手方汇总 ---
        ws.merge_cells(f"A{current_row}:H{current_row}")
        ws.cell(current_row, 1, "二、资金去向（支出）— 按对手方汇总").font = Font(name="微软雅黑", size=11, bold=True, color="C00000")
        current_row += 1

        dest_headers = ["序号", "对手方", "是否家庭成员", "支出金额(元)", "支出笔数", "占支出比例", "渠道", "备注"]
        for i, h in enumerate(dest_headers, 1):
            c = ws.cell(row=current_row, column=i, value=h)
            c.font = FONT_HEADER
            c.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            c.alignment = ALIGN_CENTER
        current_row += 1

        outcome_cps = [(cp, v) for cp, v in sorted_cp if v["out"] > 0]
        outcome_cps.sort(key=lambda x: x[1]["out"], reverse=True)
        seq = 0
        for cp, v in outcome_cps:
            seq += 1
            is_family = "是" if cp in FAMILY_MEMBERS else ""
            channels = ", ".join(sorted(v["channel"]))

            fill = FAMILY_FILL if cp in FAMILY_MEMBERS else (FILL_WHITE if seq % 2 == 0 else FILL_LIGHT)

            ws.cell(current_row, 1, seq).alignment = ALIGN_CENTER
            ws.cell(current_row, 1).fill = fill
            ws.cell(current_row, 1).font = FONT_NORMAL

            ws.cell(current_row, 2, cp).fill = fill
            ws.cell(current_row, 2).font = FONT_BOLD if cp in FAMILY_MEMBERS else FONT_NORMAL
            ws.cell(current_row, 2).alignment = ALIGN_LEFT

            ws.cell(current_row, 3, is_family).fill = fill
            ws.cell(current_row, 3).alignment = ALIGN_CENTER
            if is_family:
                ws.cell(current_row, 3).font = FONT_BOLD

            format_number(ws, current_row, 4, v["out"]).fill = fill

            ws.cell(current_row, 5, v["count_out"]).fill = fill
            ws.cell(current_row, 5).alignment = ALIGN_CENTER

            pct = v["out"] / total_out if total_out > 0 else 0
            pct_cell = ws.cell(current_row, 6, pct)
            pct_cell.number_format = '0.00%'
            pct_cell.alignment = ALIGN_CENTER
            pct_cell.fill = fill

            ws.cell(current_row, 7, channels).fill = fill
            ws.cell(current_row, 7).alignment = ALIGN_LEFT
            ws.cell(current_row, 7).font = FONT_NORMAL

            ws.cell(current_row, 8, "").fill = fill

            current_row += 1

        # 支出合计行
        for c in range(1, 9):
            ws.cell(current_row, c).fill = FILL_SUM
            ws.cell(current_row, c).font = FONT_SUM
            ws.cell(current_row, c).alignment = ALIGN_CENTER
        ws.cell(current_row, 1, "")
        ws.cell(current_row, 2, "支出合计")
        ws.cell(current_row, 3, "")
        format_number(ws, current_row, 4, total_out)
        ws.cell(current_row, 5, count_out)
        ws.cell(current_row, 6, 1.0).number_format = '0.00%'
        current_row += 2

        # --- 第三部分：按收支类别汇总 ---
        ws.merge_cells(f"A{current_row}:F{current_row}")
        ws.cell(current_row, 1, "三、按收支类别汇总").font = Font(name="微软雅黑", size=11, bold=True, color="4472C4")
        current_row += 1

        cat_headers = ["序号", "收支类别", "收入金额(元)", "支出金额(元)", "净额(元)", "备注"]
        for i, h in enumerate(cat_headers, 1):
            c = ws.cell(row=current_row, column=i, value=h)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = ALIGN_CENTER
        current_row += 1

        sorted_cats = sorted(cat_stats.items(), key=lambda x: x[1]["in"] + x[1]["out"], reverse=True)
        for seq, (cat, v) in enumerate(sorted_cats, 1):
            fill = FILL_WHITE if seq % 2 == 0 else FILL_LIGHT
            ws.cell(current_row, 1, seq).alignment = ALIGN_CENTER
            ws.cell(current_row, 1).fill = fill
            ws.cell(current_row, 2, cat).fill = fill
            ws.cell(current_row, 2).alignment = ALIGN_LEFT
            format_number(ws, current_row, 3, v["in"]).fill = fill
            format_number(ws, current_row, 4, v["out"]).fill = fill
            net = v["in"] - v["out"]
            nc = format_number(ws, current_row, 5, net).fill = fill
            ws.cell(current_row, 6, "").fill = fill
            current_row += 1

        # 列宽
        col_widths_personal = [6, 30, 14, 18, 10, 12, 20, 15]
        for i, w in enumerate(col_widths_personal, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # --------------------------------------------------------
    # Sheet 6: 家庭成员内部互转汇总
    # --------------------------------------------------------
    ws6 = wb.create_sheet(title="家庭成员互转汇总")
    ws6.sheet_properties.tabColor = "70AD47"

    ws6.merge_cells("A1:G1")
    ws6.cell(1, 1, "家庭成员内部资金互转汇总").font = FONT_TITLE
    ws6.cell(1, 1).alignment = ALIGN_CENTER

    ws6.cell(2, 1, f"家庭成员: {', '.join(FAMILY_MEMBERS)}  |  以下仅统计家庭成员之间的转账记录").font = Font(name="微软雅黑", size=10, color="666666")

    # 构建成员间转账矩阵
    matrix = defaultdict(lambda: {"in": 0.0, "out": 0.0, "count": 0})
    for holder_info in all_data:
        holder_name = holder_info["name"]
        for t in holder_info["transactions"]:
            cp = t["counterparty"]
            if cp in FAMILY_MEMBERS and cp != holder_name:
                key = (holder_name, cp)
                if t["amount_in"] > 0:
                    matrix[key]["in"] += t["amount_in"]
                    matrix[key]["count"] += 1
                if t["amount_out"] > 0:
                    matrix[key]["out"] += t["amount_out"]
                    matrix[key]["count"] += 1

    # 表头
    current_row = 4
    matrix_headers = ["序号", "转出方", "转入方", "转出金额(元)", "转入金额(元)", "净额(元)", "笔数"]
    for i, h in enumerate(matrix_headers, 1):
        c = ws6.cell(row=current_row, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        c.alignment = ALIGN_CENTER
    current_row += 1

    sorted_matrix = sorted(matrix.items(), key=lambda x: x[1]["in"] + x[1]["out"], reverse=True)
    for seq, ((from_p, to_p), v) in enumerate(sorted_matrix, 1):
        fill = FILL_WHITE if seq % 2 == 0 else FILL_LIGHT
        ws6.cell(current_row, 1, seq).alignment = ALIGN_CENTER
        ws6.cell(current_row, 1).fill = fill
        ws6.cell(current_row, 2, from_p).fill = fill
        ws6.cell(current_row, 2).font = FONT_BOLD
        ws6.cell(current_row, 3, to_p).fill = fill
        ws6.cell(current_row, 3).font = FONT_BOLD
        ws6.cell(current_row, 3).alignment = ALIGN_CENTER

        # 转出 = from_p 的支出
        # 转入 = from_p 的收入（from_p 收到的来自 to_p 的钱）
        format_number(ws6, current_row, 4, v["out"]).fill = fill  # from_p -> to_p 支出
        format_number(ws6, current_row, 5, v["in"]).fill = fill   # from_p <- to_p 收入
        net = v["in"] - v["out"]
        format_number(ws6, current_row, 6, net).fill = fill
        ws6.cell(current_row, 7, v["count"]).fill = fill
        ws6.cell(current_row, 7).alignment = ALIGN_CENTER
        current_row += 1

    # 列宽
    for i, w in enumerate([6, 12, 12, 18, 18, 18, 8], 1):
        ws6.column_dimensions[get_column_letter(i)].width = w

    # --------------------------------------------------------
    # Sheet 7: 四人资金来源去向汇总对比
    # --------------------------------------------------------
    ws7 = wb.create_sheet(title="四人资金对比汇总")
    ws7.sheet_properties.tabColor = "ED7D31"

    ws7.merge_cells("A1:I1")
    ws7.cell(1, 1, "四人资金来源去向对比汇总").font = FONT_TITLE
    ws7.cell(1, 1).alignment = ALIGN_CENTER

    # 收入对比表
    current_row = 3
    ws7.merge_cells(f"A{current_row}:I{current_row}")
    ws7.cell(current_row, 1, "一、资金来源（收入）对比").font = Font(name="微软雅黑", size=11, bold=True, color="008000")
    current_row += 1

    comp_headers_in = ["序号", "对手方", "李凯(元)", "侯红(元)", "尚喜营(元)", "尚明豪(元)", "四人合计(元)", "涉及人数", "备注"]
    for i, h in enumerate(comp_headers_in, 1):
        c = ws7.cell(row=current_row, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        c.alignment = ALIGN_CENTER
    current_row += 1

    # 收集所有对手方
    all_counterparties_in = set()
    for h in all_data:
        for cp, v in h["cp_stats"].items():
            if v["in"] > 0:
                all_counterparties_in.add(cp)

    holder_names = [h["name"] for h in all_data]
    sorted_all_cp_in = sorted(all_counterparties_in, key=lambda cp: sum(
        h["cp_stats"].get(cp, {"in": 0})["in"] for h in all_data), reverse=True)

    for seq, cp in enumerate(sorted_all_cp_in, 1):
        fill = FAMILY_FILL if cp in FAMILY_MEMBERS else (FILL_WHITE if seq % 2 == 0 else FILL_LIGHT)
        ws7.cell(current_row, 1, seq).alignment = ALIGN_CENTER
        ws7.cell(current_row, 1).fill = fill
        ws7.cell(current_row, 2, cp).fill = fill
        ws7.cell(current_row, 2).font = FONT_BOLD if cp in FAMILY_MEMBERS else FONT_NORMAL

        amounts = []
        involved = 0
        for i, h in enumerate(all_data, 3):
            val = h["cp_stats"].get(cp, {}).get("in", 0)
            format_number(ws7, current_row, i, val).fill = fill
            if val > 0:
                involved += 1
            amounts.append(val)

        total = sum(amounts)
        format_number(ws7, current_row, 7, total).fill = fill
        ws7.cell(current_row, 8, involved).fill = fill
        ws7.cell(current_row, 8).alignment = ALIGN_CENTER
        ws7.cell(current_row, 9, "").fill = fill
        current_row += 1

    # 收入合计
    for c in range(1, 10):
        ws7.cell(current_row, c).fill = FILL_SUM
        ws7.cell(current_row, c).font = FONT_SUM
        ws7.cell(current_row, c).alignment = ALIGN_CENTER
    ws7.cell(current_row, 2, "合计")
    for i, h in enumerate(all_data, 3):
        total_in_h = sum(t["amount_in"] for t in h["transactions"])
        format_number(ws7, current_row, i, total_in_h)
    format_number(ws7, current_row, 7, grand_total_in)
    current_row += 2

    # 支出对比表
    ws7.merge_cells(f"A{current_row}:I{current_row}")
    ws7.cell(current_row, 1, "二、资金去向（支出）对比").font = Font(name="微软雅黑", size=11, bold=True, color="C00000")
    current_row += 1

    comp_headers_out = ["序号", "对手方", "李凯(元)", "侯红(元)", "尚喜营(元)", "尚明豪(元)", "四人合计(元)", "涉及人数", "备注"]
    for i, h in enumerate(comp_headers_out, 1):
        c = ws7.cell(row=current_row, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        c.alignment = ALIGN_CENTER
    current_row += 1

    all_counterparties_out = set()
    for h in all_data:
        for cp, v in h["cp_stats"].items():
            if v["out"] > 0:
                all_counterparties_out.add(cp)

    sorted_all_cp_out = sorted(all_counterparties_out, key=lambda cp: sum(
        h["cp_stats"].get(cp, {"out": 0})["out"] for h in all_data), reverse=True)

    for seq, cp in enumerate(sorted_all_cp_out, 1):
        fill = FAMILY_FILL if cp in FAMILY_MEMBERS else (FILL_WHITE if seq % 2 == 0 else FILL_LIGHT)
        ws7.cell(current_row, 1, seq).alignment = ALIGN_CENTER
        ws7.cell(current_row, 1).fill = fill
        ws7.cell(current_row, 2, cp).fill = fill
        ws7.cell(current_row, 2).font = FONT_BOLD if cp in FAMILY_MEMBERS else FONT_NORMAL

        amounts = []
        involved = 0
        for i, h in enumerate(all_data, 3):
            val = h["cp_stats"].get(cp, {}).get("out", 0)
            format_number(ws7, current_row, i, val).fill = fill
            if val > 0:
                involved += 1
            amounts.append(val)

        total = sum(amounts)
        format_number(ws7, current_row, 7, total).fill = fill
        ws7.cell(current_row, 8, involved).fill = fill
        ws7.cell(current_row, 8).alignment = ALIGN_CENTER
        ws7.cell(current_row, 9, "").fill = fill
        current_row += 1

    # 支出合计
    for c in range(1, 10):
        ws7.cell(current_row, c).fill = FILL_SUM
        ws7.cell(current_row, c).font = FONT_SUM
        ws7.cell(current_row, c).alignment = ALIGN_CENTER
    ws7.cell(current_row, 2, "合计")
    for i, h in enumerate(all_data, 3):
        total_out_h = sum(t["amount_out"] for t in h["transactions"])
        format_number(ws7, current_row, i, total_out_h)
    format_number(ws7, current_row, 7, grand_total_out)

    # 列宽
    for i, w in enumerate([6, 35, 18, 18, 18, 18, 18, 10, 15], 1):
        ws7.column_dimensions[get_column_letter(i)].width = w

    # 保存
    wb.save(output_path)
    print(f"报告已生成: {output_path}")


# ============================================================
# 主程序
# ============================================================
def main():
    print("开始解析流水数据...")
    all_data = []

    for cfg in FILES:
        name = cfg["name"]
        print(f"\n读取 {name} 的流水数据...")
        try:
            txns = read_transactions(cfg)
            print(f"  交易记录数: {len(txns)}")
            print(f"  收入笔数: {sum(1 for t in txns if t['amount_in'] > 0)}")
            print(f"  支出笔数: {sum(1 for t in txns if t['amount_out'] > 0)}")
            print(f"  收入合计: {sum(t['amount_in'] for t in txns):,.2f}")
            print(f"  支出合计: {sum(t['amount_out'] for t in txns):,.2f}")

            cp_stats = aggregate_by_counterparty(txns)
            cat_stats = aggregate_by_category(txns)
            ch_stats = aggregate_by_channel(txns)
            print(f"  对手方数: {len(cp_stats)}")
            print(f"  渠道: {sorted(ch_stats.keys())}")

            all_data.append({
                "name": name,
                "transactions": txns,
                "cp_stats": cp_stats,
                "cat_stats": cat_stats,
                "ch_stats": ch_stats,
            })
        except Exception as e:
            print(f"  错误: {e}")
            import traceback
            traceback.print_exc()

    # 生成报告
    output_path = os.path.join("C:/Users/12844/WorkBuddy/2026-05-13-task-3", "资金流水统计表.xlsx")
    print(f"\n正在生成统计报告...")
    generate_report(all_data, output_path)
    print("完成!")

if __name__ == "__main__":
    main()
